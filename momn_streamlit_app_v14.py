"""
momn_streamlit_app_v14.py
=========================
Momentum Screener + Portfolio Rebalancer — v14

UI Redesign vs v13  (business logic 100% unchanged):
  • Full CSS overhaul → premium fintech SaaS aesthetic (Syne + JetBrains Mono)
  • Top horizontal navigation bar (Screener Workflow | Strategy Tearsheet)
  • Sidebar: settings ONLY — step navigation moved to top nav & step bar
  • Design tokens: crisp blue/slate palette, precise shadows, 8-px grid
  • Metric cards: clean glass-card look with sharp left accent
  • Section headers: tighter, typography-driven (no background wash)
  • Step progress bar: pill-based, cleaner active states
  • Tables: zebra rows via CSS, subtle row hover
  • Buttons: clear hierarchy (primary / ghost)
  • Integrated Strategy Tearsheet tab (from strategy-tearsheet.py)
  • App footer with version + author
"""

import io
import time
import datetime
import warnings

import numpy as np
import pandas as pd
import streamlit as st
import requests

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side

warnings.filterwarnings("ignore")

# ── Local modules ──────────────────────────────────────────────
try:
    from calculations import build_dfStats, apply_filters
    _CALCS_AVAILABLE = True
except ImportError:
    _CALCS_AVAILABLE = False

import yfinance as yf

_DS_AVAILABLE   = False
_DS_IMPORT_ERR  = ""
try:
    from data_service import fetch_data
    _DS_AVAILABLE = True
except Exception as _e:
    _DS_IMPORT_ERR = str(_e)

_UPSTOX_AVAILABLE = False
try:
    from upstox_auth import get_upstox_access_token
    _UPSTOX_AVAILABLE = True
except Exception:
    pass

_ANGEL_AVAILABLE = False
try:
    from angelone_auth import get_angelone_client
    _ANGEL_AVAILABLE = True
except Exception:
    pass

_CACHE_AVAILABLE = False
try:
    from cache_loader import load_cache, get_cache_meta, get_cache_age_days, get_cache_status_html
    _CACHE_AVAILABLE = True
except ImportError:
    pass

_CACHE_UPSTOX_AVAILABLE = False
try:
    from cache_loader_upstox import (
        load_cache          as load_cache_upstox,
        get_cache_meta      as get_cache_meta_upstox,
        get_cache_age_days  as get_cache_age_days_upstox,
        get_cache_status_html as get_cache_status_html_upstox,
    )
    _CACHE_UPSTOX_AVAILABLE = True
except ImportError:
    pass

# ── Inline YFinance fetcher (unchanged from v13) ───────────────
def _fetch_yfinance_inline(symbols_ns, start_date, end_date,
                            progress_bar, status_text, chunk_size=15):
    """Pure yfinance fetch — no data_service dependency."""
    close_chunks, high_chunks, vol_chunks = [], [], []
    failed = []
    total  = len(symbols_ns)
    for k in range(0, total, chunk_size):
        chunk = symbols_ns[k:k + chunk_size]
        pct   = min((k + chunk_size) / total, 1.0)
        status_text.markdown(f"⏳ **Fetching {k+1}–{min(k+chunk_size, total)} / {total}**")
        progress_bar.progress(pct * 0.88)
        try:
            raw = yf.download(chunk, start=start_date, end=end_date,
                              progress=False, auto_adjust=True, threads=True,
                              multi_level_index=False)
            if not raw.empty:
                close_chunks.append(raw["Close"])
                high_chunks.append(raw["High"])
                vol_val = raw["Close"].multiply(raw.get("Volume", 1))
                vol_chunks.append(vol_val)
        except Exception as e:
            failed.extend(chunk)
        time.sleep(0.5)

    if not close_chunks:
        return None, None, None, failed

    close  = pd.concat(close_chunks,  axis=1)
    high   = pd.concat(high_chunks,   axis=1)
    volume = pd.concat(vol_chunks,    axis=1)
    close  = close.loc[:,  ~close.columns.duplicated()]
    high   = high.loc[:,   ~high.columns.duplicated()]
    volume = volume.loc[:, ~volume.columns.duplicated()]
    return close, high, volume, failed


# ═══════════════════════════════════════════════════════════════
# PAGE CONFIG
# ═══════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="Momn Screener v14",
    page_icon="📈",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ═══════════════════════════════════════════════════════════════
# ▸ CSS — v14 DESIGN SYSTEM
#   Aesthetic: Refined fintech minimal (Syne + JetBrains Mono)
#   Color: crisp slate+blue, zero gradients on data surfaces
# ═══════════════════════════════════════════════════════════════
st.markdown("""
<style>
/* ── Google Fonts ── */
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;500;600;700;800&family=JetBrains+Mono:wght@400;500&family=Inter:wght@400;500;600&display=swap');

/* ────────────────────────────────────────
   DESIGN TOKENS
──────────────────────────────────────── */
:root {
    /* Core palette */
    --ink:         #0a0f1e;
    --ink-mid:     #1e293b;
    --ink-soft:    #334155;
    --muted:       #64748b;
    --subtle:      #94a3b8;
    --border:      #e2e8f0;
    --border-soft: #f1f5f9;
    --surface:     #ffffff;
    --canvas:      #f8fafc;

    /* Brand */
    --blue:        #2563eb;
    --blue-light:  #3b82f6;
    --blue-bg:     #eff6ff;
    --blue-bdr:    #bfdbfe;

    /* Semantic */
    --green:       #16a34a;
    --green-bg:    #f0fdf4;
    --green-bdr:   #bbf7d0;
    --red:         #dc2626;
    --red-bg:      #fff1f2;
    --red-bdr:     #fecaca;
    --violet:      #7c3aed;
    --violet-bg:   #faf5ff;
    --violet-bdr:  #ddd6fe;
    --amber:       #d97706;
    --amber-bg:    #fffbeb;
    --amber-bdr:   #fde68a;
    --teal:        #0891b2;
    --teal-bg:     #ecfeff;
    --teal-bdr:    #a5f3fc;

    /* Typography */
    --font-ui:     'Inter', system-ui, sans-serif;
    --font-head:   'Syne', 'Inter', sans-serif;
    --font-mono:   'JetBrains Mono', 'Fira Code', monospace;

    /* Spacing */
    --r-sm:  6px;
    --r-md:  10px;
    --r-lg:  14px;
    --r-xl:  20px;

    /* Shadows */
    --sh-xs: 0 1px 2px rgba(0,0,0,.05);
    --sh-sm: 0 1px 4px rgba(0,0,0,.06), 0 1px 2px rgba(0,0,0,.04);
    --sh-md: 0 4px 12px rgba(0,0,0,.08), 0 2px 4px rgba(0,0,0,.04);
    --sh-lg: 0 8px 24px rgba(0,0,0,.10), 0 4px 8px rgba(0,0,0,.06);
}

/* ────────────────────────────────────────
   GLOBAL RESET
──────────────────────────────────────── */
html, body, [class*="css"] {
    font-family: var(--font-ui);
    color: var(--ink);
    -webkit-font-smoothing: antialiased;
}

/* Remove Streamlit default top padding */
.block-container { padding-top: 0.75rem !important; }

/* Mono class for numbers */
.mono { font-family: var(--font-mono) !important; }

/* ────────────────────────────────────────
   APP HEADER (top bar)
──────────────────────────────────────── */
.app-header {
    background: var(--ink);
    border-bottom: 1px solid rgba(255,255,255,.06);
    padding: 0 24px;
    margin: -0.75rem -1rem 0 -1rem;
    height: 56px;
    display: flex;
    align-items: center;
    justify-content: space-between;
    gap: 16px;
    position: sticky;
    top: 0;
    z-index: 100;
}
.app-logo {
    display: flex;
    align-items: center;
    gap: 10px;
    flex-shrink: 0;
}
.app-logo-icon {
    width: 30px;
    height: 30px;
    background: var(--blue);
    border-radius: 8px;
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 15px;
}
.app-logo-text {
    font-family: var(--font-head);
    font-size: 16px;
    font-weight: 800;
    color: #f1f5f9;
    letter-spacing: -.2px;
}
.app-logo-text span { color: #60a5fa; }
.app-badge {
    background: rgba(96,165,250,.15);
    border: 1px solid rgba(96,165,250,.25);
    color: #60a5fa;
    font-family: var(--font-mono);
    font-size: 9px;
    font-weight: 500;
    padding: 2px 7px;
    border-radius: 20px;
    letter-spacing: .5px;
    margin-left: 8px;
}
.app-header-right {
    display: flex;
    align-items: center;
    gap: 12px;
}
.hdr-date {
    font-size: 11.5px;
    color: #94a3b8;
    font-family: var(--font-mono);
}
.hdr-user {
    background: rgba(255,255,255,.07);
    border: 1px solid rgba(255,255,255,.10);
    color: #cbd5e1;
    font-size: 11.5px;
    font-weight: 500;
    padding: 5px 12px;
    border-radius: 20px;
}

/* ────────────────────────────────────────
   TOP NAV TABS
──────────────────────────────────────── */
.top-nav {
    background: var(--surface);
    border-bottom: 1px solid var(--border);
    padding: 0 24px;
    margin: 0 -1rem 1.5rem -1rem;
    display: flex;
    align-items: stretch;
    gap: 0;
    overflow-x: auto;
}
.nav-tab {
    display: inline-flex;
    align-items: center;
    gap: 6px;
    padding: 13px 18px;
    font-size: 13px;
    font-weight: 600;
    color: var(--muted);
    border-bottom: 2.5px solid transparent;
    cursor: pointer;
    white-space: nowrap;
    transition: color .15s, border-color .15s;
    text-decoration: none !important;
}
.nav-tab:hover { color: var(--ink-mid); }
.nav-tab.active {
    color: var(--blue);
    border-bottom-color: var(--blue);
}
.nav-tab-icon { font-size: 14px; }

/* ────────────────────────────────────────
   STEP PROGRESS BAR
──────────────────────────────────────── */
.step-bar {
    display: flex;
    align-items: center;
    gap: 0;
    margin-bottom: 1.5rem;
    overflow-x: auto;
    padding-bottom: 2px;
}
.step-item {
    display: flex;
    align-items: center;
    gap: 8px;
    padding: 9px 16px;
    border-radius: var(--r-lg);
    font-size: 12.5px;
    font-weight: 600;
    white-space: nowrap;
    transition: all .2s;
    cursor: default;
}
.step-item.done {
    background: var(--green-bg);
    color: var(--green);
    border: 1px solid var(--green-bdr);
}
.step-item.active {
    background: var(--blue-bg);
    color: var(--blue);
    border: 1.5px solid var(--blue-bdr);
    box-shadow: 0 0 0 3px rgba(37,99,235,.08);
}
.step-item.pending {
    color: var(--subtle);
    border: 1px solid var(--border-soft);
    background: var(--canvas);
}
.step-circle {
    width: 22px;
    height: 22px;
    border-radius: 50%;
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 10px;
    font-weight: 800;
    flex-shrink: 0;
    font-family: var(--font-head);
}
.done   .step-circle { background: var(--green); color: #fff; }
.active .step-circle { background: var(--blue);  color: #fff; }
.pending .step-circle { background: var(--border); color: var(--muted); }
.step-connector {
    width: 28px;
    height: 1.5px;
    background: var(--border);
    flex-shrink: 0;
    margin: 0 4px;
    border-radius: 2px;
}
.step-connector.done-line { background: var(--green-bdr); }

/* ────────────────────────────────────────
   METRIC CARDS
──────────────────────────────────────── */
.metric-row {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(140px, 1fr));
    gap: 12px;
    margin: 12px 0 20px;
}
.metric-card {
    background: var(--surface);
    border: 1px solid var(--border);
    border-left: 3px solid var(--ink-soft);
    border-radius: var(--r-md);
    padding: 14px 18px;
    box-shadow: var(--sh-xs);
    transition: box-shadow .2s, transform .15s;
    position: relative;
    overflow: hidden;
}
.metric-card::before {
    content: '';
    position: absolute;
    top: 0; right: 0; bottom: 0;
    width: 60px;
    background: linear-gradient(to left, rgba(248,250,252,.8), transparent);
    pointer-events: none;
}
.metric-card:hover {
    box-shadow: var(--sh-md);
    transform: translateY(-1px);
}
.metric-card.green  { border-left-color: var(--green); }
.metric-card.red    { border-left-color: var(--red); }
.metric-card.blue   { border-left-color: var(--blue); }
.metric-card.violet { border-left-color: var(--violet); }
.metric-card.amber  { border-left-color: var(--amber); }
.metric-card.teal   { border-left-color: var(--teal); }
.metric-label {
    font-size: 9.5px;
    color: var(--muted);
    text-transform: uppercase;
    letter-spacing: .8px;
    font-weight: 600;
    margin-bottom: 5px;
}
.metric-value {
    font-family: var(--font-head);
    font-size: 24px;
    font-weight: 800;
    color: var(--ink);
    letter-spacing: -.4px;
    line-height: 1;
}
.metric-value.green  { color: var(--green); }
.metric-value.red    { color: var(--red); }
.metric-value.blue   { color: var(--blue); }
.metric-value.violet { color: var(--violet); }
.metric-value.amber  { color: var(--amber); }
.metric-value.teal   { color: var(--teal); }

/* ────────────────────────────────────────
   SECTION HEADERS
──────────────────────────────────────── */
.section-hdr {
    font-family: var(--font-head);
    font-size: 13px;
    font-weight: 700;
    color: var(--ink-mid);
    text-transform: uppercase;
    letter-spacing: .6px;
    border-left: 3px solid var(--blue);
    padding: 5px 0 5px 12px;
    margin: 1.6rem 0 1rem;
}

/* ────────────────────────────────────────
   CARD CONTAINERS
──────────────────────────────────────── */
.card {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: var(--r-lg);
    padding: 20px 22px;
    box-shadow: var(--sh-sm);
    margin-bottom: 16px;
}
.card-sm {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: var(--r-md);
    padding: 12px 16px;
    box-shadow: var(--sh-xs);
}

/* ────────────────────────────────────────
   CHIPS (SELL / BUY / HOLD)
──────────────────────────────────────── */
.chip {
    display: inline-block;
    padding: 3px 9px;
    border-radius: 20px;
    font-size: 11px;
    font-weight: 600;
    margin: 2px 2px;
    letter-spacing: .2px;
    border: 1px solid transparent;
    font-family: var(--font-mono);
}
.chip-sell  { background: var(--red-bg);    color: var(--red);    border-color: var(--red-bdr); }
.chip-buy   { background: var(--green-bg);  color: var(--green);  border-color: var(--green-bdr); }
.chip-hold  { background: var(--canvas);    color: var(--ink-soft); border-color: var(--border); }

/* ────────────────────────────────────────
   REBALANCE STRIP
──────────────────────────────────────── */
.reb-strip {
    display: flex;
    gap: 0;
    flex-wrap: wrap;
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: var(--r-md);
    overflow: hidden;
    margin: 12px 0;
    box-shadow: var(--sh-xs);
}
.reb-stat {
    flex: 1;
    min-width: 100px;
    padding: 12px 16px;
    border-right: 1px solid var(--border-soft);
    text-align: center;
}
.reb-stat:last-child { border-right: none; }
.reb-stat .label {
    font-size: 9px;
    color: var(--muted);
    text-transform: uppercase;
    letter-spacing: .8px;
    font-weight: 600;
    margin-bottom: 5px;
}
.reb-stat .val {
    font-family: var(--font-head);
    font-size: 20px;
    font-weight: 800;
    letter-spacing: -.2px;
    color: var(--ink);
}
.reb-stat .val.r { color: var(--red); }
.reb-stat .val.g { color: var(--green); }
.reb-stat .val.b { color: var(--blue); }
.reb-stat .val.p { color: var(--violet); }

/* ────────────────────────────────────────
   NSE LINK BOX
──────────────────────────────────────── */
.nse-link-box {
    background: var(--blue-bg);
    border: 1px solid var(--blue-bdr);
    border-radius: var(--r-md);
    padding: 14px 18px;
    display: flex;
    align-items: center;
    gap: 12px;
    margin: 12px 0;
}
.nse-link-box a { color: var(--blue); font-weight: 600; font-size: 13px; text-decoration: none; }
.nse-link-box a:hover { text-decoration: underline; }
.nse-link-box .hint { font-size: 11.5px; color: var(--muted); margin-top: 3px; }

/* ────────────────────────────────────────
   WORKFLOW BOX
──────────────────────────────────────── */
.workflow-box {
    background: var(--canvas);
    border: 1px solid var(--border);
    border-left: 3px solid var(--blue);
    border-radius: var(--r-md);
    padding: 16px 20px;
    margin: 12px 0;
    font-size: 13px;
    line-height: 2.2;
    color: var(--ink-soft);
}
.step-tag {
    display: inline-flex;
    align-items: center;
    justify-content: center;
    background: var(--blue);
    color: #fff;
    border-radius: 50%;
    width: 20px;
    height: 20px;
    font-size: 10px;
    font-weight: 800;
    margin-right: 8px;
    flex-shrink: 0;
    font-family: var(--font-head);
}

/* ────────────────────────────────────────
   LOGIN PAGE
──────────────────────────────────────── */
.login-wrap {
    min-height: 80vh;
    display: flex;
    align-items: center;
    justify-content: center;
    padding: 40px 16px;
}
.login-card {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: 20px;
    box-shadow: var(--sh-lg);
    padding: 44px 48px 36px;
    width: 100%;
    max-width: 380px;
}
.login-icon {
    width: 52px; height: 52px;
    background: var(--blue);
    border-radius: 14px;
    display: flex; align-items: center; justify-content: center;
    font-size: 26px;
    margin: 0 auto 18px;
}
.login-title {
    font-family: var(--font-head);
    font-size: 22px;
    font-weight: 800;
    color: var(--ink);
    text-align: center;
    margin-bottom: 4px;
}
.login-sub {
    font-size: 12.5px;
    color: var(--muted);
    text-align: center;
    margin-bottom: 28px;
}

/* ────────────────────────────────────────
   QUICK LINK BUTTONS
──────────────────────────────────────── */
.qlink-btn {
    display: block;
    text-align: center;
    font-weight: 700;
    font-size: 13px;
    padding: 13px 20px;
    border-radius: var(--r-md);
    text-decoration: none !important;
    margin: 6px 0;
    transition: opacity .2s, transform .15s;
    color: #ffffff !important;
}
.qlink-btn:hover { opacity: .88; transform: translateY(-1px); color: #ffffff !important; }
.qlink-rebalancer { background: #1e3a8a; }
.qlink-dashboard  { background: #5b21b6; }

/* ────────────────────────────────────────
   SIDEBAR
──────────────────────────────────────── */
section[data-testid="stSidebar"] {
    background: var(--canvas) !important;
}
section[data-testid="stSidebar"] .block-container {
    padding-top: 1.2rem;
}
.sidebar-section-label {
    font-family: var(--font-head);
    font-size: 10px;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 1px;
    color: var(--muted);
    margin: 0 0 8px 2px;
}

/* ────────────────────────────────────────
   DATAFRAME TABLE TWEAKS
──────────────────────────────────────── */
[data-testid="stDataFrame"] table {
    font-family: var(--font-mono);
    font-size: 12px;
}
[data-testid="stDataFrame"] thead tr th {
    background: var(--canvas) !important;
    color: var(--muted) !important;
    font-size: 10px !important;
    text-transform: uppercase !important;
    letter-spacing: .6px !important;
    font-weight: 700 !important;
    border-bottom: 1px solid var(--border) !important;
}
[data-testid="stDataFrame"] tbody tr:nth-child(even) td {
    background: var(--canvas) !important;
}
[data-testid="stDataFrame"] tbody tr:hover td {
    background: var(--blue-bg) !important;
}

/* ────────────────────────────────────────
   STATUS BADGES
──────────────────────────────────────── */
.badge {
    display: inline-block;
    padding: 3px 10px;
    border-radius: 20px;
    font-size: 11px;
    font-weight: 600;
    border: 1px solid transparent;
}
.badge-blue   { background: var(--blue-bg);   color: var(--blue);   border-color: var(--blue-bdr); }
.badge-green  { background: var(--green-bg);  color: var(--green);  border-color: var(--green-bdr); }
.badge-red    { background: var(--red-bg);    color: var(--red);    border-color: var(--red-bdr); }
.badge-violet { background: var(--violet-bg); color: var(--violet); border-color: var(--violet-bdr); }
.badge-amber  { background: var(--amber-bg);  color: var(--amber);  border-color: var(--amber-bdr); }

/* ────────────────────────────────────────
   FOOTER
──────────────────────────────────────── */
.app-footer {
    margin-top: 40px;
    border-top: 1px solid var(--border);
    padding: 16px 0 8px;
    display: flex;
    justify-content: space-between;
    align-items: center;
    font-size: 11px;
    color: var(--subtle);
}
.footer-brand {
    font-family: var(--font-head);
    font-weight: 700;
    color: var(--muted);
}
.footer-brand span { color: var(--blue); }

/* ────────────────────────────────────────
   UNIVERSE INFO CARD
──────────────────────────────────────── */
.universe-card {
    background: var(--surface);
    border: 1px solid var(--border);
    border-left: 3px solid var(--teal);
    border-radius: var(--r-md);
    padding: 14px 18px;
    display: flex;
    align-items: center;
    gap: 14px;
    box-shadow: var(--sh-xs);
}
.universe-card .u-icon { font-size: 26px; }
.universe-card .u-name {
    font-family: var(--font-head);
    font-size: 15px;
    font-weight: 800;
    color: var(--ink);
}
.universe-card .u-meta { font-size: 12px; color: var(--muted); margin-top: 2px; }

/* ────────────────────────────────────────
   INFO STRIP (Step 2 run info)
──────────────────────────────────────── */
.run-info-strip {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: var(--r-md);
    padding: 12px 18px;
    font-size: 12.5px;
    color: var(--ink-soft);
    line-height: 2;
    box-shadow: var(--sh-xs);
}
.run-info-strip b { color: var(--ink); }

/* ────────────────────────────────────────
   TEARSHEET PAGE
──────────────────────────────────────── */
.tearsheet-header {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: var(--r-lg);
    padding: 20px 24px;
    margin-bottom: 20px;
    box-shadow: var(--sh-sm);
}
.tearsheet-title {
    font-family: var(--font-head);
    font-size: 18px;
    font-weight: 800;
    color: var(--ink);
    margin-bottom: 4px;
}
.tearsheet-sub {
    font-size: 12.5px;
    color: var(--muted);
}

/* ────────────────────────────────────────
   HIDE STREAMLIT BRANDING
──────────────────────────────────────── */
#MainMenu  { visibility: hidden; }
footer     { visibility: hidden; }
.stDeployButton { display: none; }
header     { visibility: hidden; }
</style>
""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════
# CONSTANTS  (unchanged from v13)
# ═══════════════════════════════════════════════════════════════
UNIVERSES    = ['Nifty50','Nifty100','Nifty200','Nifty250','Nifty500','N750','AllNSE']
API_OPTIONS  = ["📦 Pre-cached YFinance", "📦 Pre-cached Upstox", "YFinance", "Upstox", "Angel One"]
RANKING_MAP  = {
    "AvgZScore 12M/6M/3M":    "avgZScore12_6_3",
    "AvgZScore 12M/9M/6M/3M": "avgZScore12_9_6_3",
    "AvgSharpe 12M/6M/3M":    "avgSharpe12_6_3",
    "AvgSharpe 9M/6M/3M":     "avgSharpe9_6_3",
    "AvgSharpe 12M/9M/6M/3M": "avg_All",
    "Sharpe12M":               "sharpe12M",
    "Sharpe3M":                "sharpe3M",
}
PORTFOLIO_CSV_URL = (
    "https://docs.google.com/spreadsheets/d/e/"
    "2PACX-1vS4HDgiell4n1kd08OnlzOQobfPzeDtVyWJ8gETFlYbz27qhOmfqKZOoIXZItRQEq5ANATYIcZJm0gk"
    "/pub?output=csv"
)
APPS_SCRIPT_URL = (
    "https://script.google.com/macros/s/"
    "AKfycbwUNaPd82fIyQXBrPguLBZBv4tLA94Y_Uw4g-8_W77qRvmpQgJvK6_huvWcjVy0XRkc/exec"
)
GITHUB_BASE = "https://raw.githubusercontent.com/prayan2702/Streamlit_Momn_v13_Cached_DB/refs/heads/main"

TEARSHEET_CSV_URL = (
    "https://docs.google.com/spreadsheets/d/e/"
    "2PACX-1vTuyGRVZuafIk2s7moScIn5PAUcPYEyYIOOYJj54RXYUeugWmOP0iIToljSEMhHrg_Zp8Vab6YvBJDV"
    "/pub?output=csv"
)

_auth = st.secrets.get("auth", {})
USERNAME = _auth.get("username", "")
PASSWORD = _auth.get("password", "")

# ═══════════════════════════════════════════════════════════════
# SESSION STATE  (extended with main_view)
# ═══════════════════════════════════════════════════════════════
_defaults = {
    "logged_in":       False,
    "current_step":    1,
    "main_view":       "screener",   # "screener" | "tearsheet"
    "universe":        "AllNSE",
    "symbols":         None,
    "eq_df":           None,
    "dfStats":         None,
    "dfFiltered":      None,
    "failed_blank":    [],
    "reb_portfolio":   None,
    "sell_list":       None,
    "buy_list":        None,
    "rebalance_table": None,
    "lookback_date":   datetime.date.today(),
    "ranking_method":  "avgZScore12_6_3",
    "data_source":     "YFinance",
    "top_n_rank":      100,
    "screener_done":   False,
    "rebalance_done":  False,
}
for k, v in _defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v


# ═══════════════════════════════════════════════════════════════
# HELPERS  (unchanged from v13)
# ═══════════════════════════════════════════════════════════════
def fmt_inr(v):
    if pd.isna(v): return "—"
    v = int(round(v))
    if abs(v) >= 10_000_000: return f"₹{v/10_000_000:.1f}Cr"
    if abs(v) >= 100_000:    return f"₹{v/100_000:.1f}L"
    return f"₹{v:,}"

def metric_card(label, value, color=""):
    cls     = f"metric-card {color}" if color else "metric-card"
    val_cls = f"metric-value {color}" if color else "metric-value"
    return f'<div class="{cls}"><div class="metric-label">{label}</div><div class="{val_cls}">{value}</div></div>'

def parse_equity_csv(f) -> pd.DataFrame:
    df = pd.read_csv(f, skipinitialspace=True)
    df.columns = [c.strip() for c in df.columns]
    if 'SERIES' in df.columns:
        df = df[df['SERIES'].str.strip() == 'EQ'].copy()
    df['SYMBOL'] = df['SYMBOL'].str.strip().str.upper()
    return df.reset_index(drop=True)

def load_symbols_from_github(universe: str) -> list:
    if universe == 'N750':
        url = f"{GITHUB_BASE}/ind_niftytotalmarket_list.csv"
    else:
        url = f"{GITHUB_BASE}/ind_{universe.lower()}list.csv"
    df = pd.read_csv(url)
    df['Yahoo_Symbol'] = df['Symbol'].astype(str).str.strip() + '.NS'
    return df['Yahoo_Symbol'].tolist()

EXTRA_SYMBOLS = ["GOLDBEES.NS", "SILVERBEES.NS"]

def add_extra_symbols(syms: list) -> list:
    result = list(syms)
    for s in EXTRA_SYMBOLS:
        if s not in result:
            result.append(s)
    return result

def build_dates(end_date: datetime.date) -> dict:
    from dateutil.relativedelta import relativedelta
    end = datetime.datetime.combine(end_date, datetime.time())
    return {
        'startDate': datetime.datetime(2000, 1, 1),
        'endDate':   end,
        'date12M':   end - relativedelta(months=12),
        'date9M':    end - relativedelta(months=9),
        'date6M':    end - relativedelta(months=6),
        'date3M':    end - relativedelta(months=3),
        'date1M':    end - relativedelta(months=1),
    }


# ── v10-identical Excel formatting (unchanged) ────────────────
def format_excel_unfiltered(file_name, universe, top_n):
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"),  bottom=Side(style="thin"))
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border    = thin
            cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = 'A2'
    hdr_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")
    for col in range(1, ws.max_column + 1):
        c = ws.cell(row=1, column=col)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center")
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    no_fill   = PatternFill(start_color="d6b4fc", end_color="d6b4fc", fill_type="solid")
    bold_font = Font(bold=True)
    headers   = [c.value for c in ws[1]]

    def ci(name): return headers.index(name) + 1 if name in headers else None

    rank_threshold = top_n
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    idx = {k: ci(k) for k in ['volm_cr','Close','dma200d','AWAY_ATH','roc12M',
                                'circuit','roc1M','circuit5','Ticker','Rank']}

    for row in range(2, ws.max_row + 1):
        failed = False
        def v(col): return ws.cell(row=row, column=col).value if col else None
        def mark(col):
            nonlocal failed
            ws.cell(row=row, column=col).fill = no_fill
            ws.cell(row=row, column=col).font = bold_font
            failed = True
        if (vol := v(idx['volm_cr']))  is not None and vol < 1:           mark(idx['volm_cr'])
        cl = v(idx['Close']); dm = v(idx['dma200d'])
        if cl is not None and dm is not None and cl <= dm:                 mark(idx['Close'])
        if (aa := v(idx['AWAY_ATH']))  is not None and aa <= -25:         mark(idx['AWAY_ATH'])
        roc = v(idx['roc12M'])
        if roc is not None and roc <= 5.5:                                 mark(idx['roc12M'])
        if (ci_ := v(idx['circuit']))  is not None and ci_ >= 20:         mark(idx['circuit'])
        if cl is not None and cl <= 30:                                    mark(idx['Close'])
        if (c5 := v(idx['circuit5']))  is not None and c5 > 10:           mark(idx['circuit5'])
        if roc is not None and roc > 1000:                                 mark(idx['roc12M'])
        if failed and idx['Ticker']:
            ws.cell(row=row, column=idx['Ticker']).fill = no_fill
        if idx['Rank'] and (rk := v(idx['Rank'])) is not None and rk <= rank_threshold:
            ws.cell(row=row, column=idx['Rank']).fill = green_fill

    ath_col = ci('ATH')
    if ath_col:
        for r in range(2, ws.max_row + 1):
            c = ws.cell(row=r, column=ath_col)
            if isinstance(c.value, (int, float)):
                c.value = round(c.value)
    wb.save(file_name)


def format_excel_filtered(file_name, universe, top_n):
    wb = openpyxl.load_workbook(file_name)
    ws = wb["Filtered Stocks"]
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"),  bottom=Side(style="thin"))
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border    = thin
            cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = 'A2'
    hdr_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")
    for col in range(1, ws.max_column + 1):
        c = ws.cell(row=1, column=col)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center")
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == "ATH":
            for r in range(2, ws.max_row + 1):
                c = ws.cell(row=r, column=col)
                if isinstance(c.value, (int, float)):
                    c.value = round(c.value)
            break
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == "AWAY_ATH":
            for r in range(2, ws.max_row + 1):
                c = ws.cell(row=r, column=col)
                if isinstance(c.value, (int, float)):
                    c.value = f"{c.value}%"
            break

    rank_threshold = top_n
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == "Rank":
            rank_75_count = 0
            for r in range(2, ws.max_row + 1):
                c = ws.cell(row=r, column=col)
                if isinstance(c.value, (int, float)) and c.value <= rank_threshold:
                    c.fill = green_fill
                    rank_75_count += 1
            total_filtered = ws.max_row - 1
            ws.append([])
            ws.append(["Summary"])
            summary_start = ws.max_row
            ws.append([f"Total Filtered Stocks: {total_filtered}"])
            ws.append([f"Number of Stocks within {rank_threshold} Rank: {rank_75_count}"])
            for r in ws.iter_rows(min_row=summary_start, max_row=ws.max_row, min_col=1, max_col=1):
                for cell in r:
                    cell.font = Font(bold=True)
            break
    wb.save(file_name)


def format_simple_sheet(file_name, sheet_name):
    wb = openpyxl.load_workbook(file_name)
    if sheet_name not in wb.sheetnames:
        wb.save(file_name); return
    ws = wb[sheet_name]
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"),  bottom=Side(style="thin"))
    hdr_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border    = thin
            cell.alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, ws.max_column + 1):
        c = ws.cell(row=1, column=col)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center")
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
    ws.freeze_panes = 'A2'

    if sheet_name == "Portfolio Rebalancing":
        headers  = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        sell_col = headers.index('Sell Stocks') + 1 if 'Sell Stocks' in headers else None
        buy_col  = headers.index('Buy Stocks')  + 1 if 'Buy Stocks'  in headers else None
        sell_fill = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")
        buy_fill  = PatternFill(start_color="D7FFD7", end_color="D7FFD7", fill_type="solid")
        for r in range(2, ws.max_row + 1):
            if sell_col: ws.cell(row=r, column=sell_col).fill = sell_fill
            if buy_col:  ws.cell(row=r, column=buy_col).fill  = buy_fill

    if sheet_name == "Failed Downloads":
        headers   = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        stock_col = headers.index('Failed Stock') + 1 if 'Failed Stock' in headers else None
        org_fill  = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
        if stock_col:
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=stock_col).fill = org_fill
    wb.save(file_name)


# ═══════════════════════════════════════════════════════════════
# ▸ UI COMPONENTS
# ═══════════════════════════════════════════════════════════════

def render_step_bar(current: int):
    """Render step progress bar HTML."""
    steps = [(1, "Universe"), (2, "Screener"), (3, "Rebalance"), (4, "Export")]
    html = '<div class="step-bar">'
    for i, (n, label) in enumerate(steps):
        cls = "done" if n < current else ("active" if n == current else "pending")
        sym = "✓"   if n < current else str(n)
        html += f'<div class="step-item {cls}"><div class="step-circle">{sym}</div>{label}</div>'
        if i < len(steps) - 1:
            lc = "done-line" if n < current else ""
            html += f'<div class="step-connector {lc}"></div>'
    html += '</div>'
    return html


def render_app_header():
    """Render top app header bar."""
    today = datetime.date.today().strftime("%d %b %Y")
    st.markdown(f"""
    <div class="app-header">
      <div class="app-logo">
        <div class="app-logo-icon">📈</div>
        <div>
          <span class="app-logo-text"><span>Momn</span> Screener</span>
          <span class="app-badge">v14</span>
        </div>
      </div>
      <div class="app-header-right">
        <span class="hdr-date">📅 {today}</span>
        <span class="hdr-user">👤 prayan2702</span>
      </div>
    </div>
    """, unsafe_allow_html=True)


def render_top_nav():
    """Render top navigation tabs using Streamlit buttons styled as nav tabs."""
    # We use Streamlit columns + buttons to drive session state navigation
    # The active tab is stored in st.session_state.main_view
    tabs_def = [
        ("screener",   "🚀", "Screener Workflow"),
        ("tearsheet",  "📋", "Strategy Tearsheet"),
    ]
    # Build HTML nav (visual only — actual switching done via st.button below)
    nav_html = '<div class="top-nav">'
    for key, icon, label in tabs_def:
        active_cls = "active" if st.session_state.main_view == key else ""
        nav_html += f'<div class="nav-tab {active_cls}"><span class="nav-tab-icon">{icon}</span>{label}</div>'
    # External links inside nav
    nav_html += f'<a class="nav-tab" href="{APPS_SCRIPT_URL}" target="_blank"><span class="nav-tab-icon">⚖️</span>Rebalancer ↗</a>'
    nav_html += '<a class="nav-tab" href="https://prayan2702.github.io/momn-dashboard/" target="_blank"><span class="nav-tab-icon">📈</span>Dashboard ↗</a>'
    nav_html += '</div>'
    st.markdown(nav_html, unsafe_allow_html=True)

    # Actual click-able tab buttons (hidden visually via zero-height trick isn't possible in Streamlit,
    # so we render them in a single compact row beneath the visual nav)
    col_tabs = st.columns([1, 1, 3])
    with col_tabs[0]:
        if st.button("🚀 Workflow", use_container_width=True,
                     type="primary" if st.session_state.main_view == "screener" else "secondary",
                     key="nav_btn_screener"):
            st.session_state.main_view = "screener"; st.rerun()
    with col_tabs[1]:
        if st.button("📋 Tearsheet", use_container_width=True,
                     type="primary" if st.session_state.main_view == "tearsheet" else "secondary",
                     key="nav_btn_tearsheet"):
            st.session_state.main_view = "tearsheet"; st.rerun()


def render_footer():
    """Render bottom app footer."""
    st.markdown(f"""
    <div class="app-footer">
      <div class="footer-brand">Momn <span>Screener</span> v14</div>
      <div>NSE Momentum Strategy · Equal-Weight Monthly Rebalancing</div>
      <div>prayan2702 · {datetime.date.today().year}</div>
    </div>
    """, unsafe_allow_html=True)


def render_sidebar():
    """Render sidebar with settings only (no step navigation)."""
    with st.sidebar:
        st.markdown('<p class="sidebar-section-label">⚙️ Screener Settings</p>', unsafe_allow_html=True)

        rm_display = st.selectbox(
            "Ranking Method",
            list(RANKING_MAP.keys()),
            index=0,
            help="Momentum ranking formula"
        )
        st.session_state.ranking_method = RANKING_MAP[rm_display]

        st.session_state.data_source = st.selectbox(
            "Data Source",
            API_OPTIONS,
            index=0
        )
        st.session_state.lookback_date = st.date_input(
            "Lookback Date",
            value=st.session_state.lookback_date,
            max_value=datetime.date.today()
        )
        st.session_state.top_n_rank = st.number_input(
            "Top-N Rank",
            min_value=20, max_value=200, value=100, step=10
        )

        # ── API Authentication ──────────────────────────────
        if st.session_state.data_source == "Upstox":
            st.divider()
            if _UPSTOX_AVAILABLE:
                get_upstox_access_token(sidebar=True)
            else:
                st.warning("⚠️ `pyotp` install nahi hai. YFinance fallback use hoga.")

        elif st.session_state.data_source == "Angel One":
            st.divider()
            if _ANGEL_AVAILABLE:
                get_angelone_client(sidebar=True)
            else:
                st.warning(
                    "⚠️ Angel One ke liye `smartapi-python` + `pyotp` "
                    "`requirements.txt` mein add karo."
                )

        st.divider()

        # ── Step navigation (compact, in sidebar) ──────────
        st.markdown('<p class="sidebar-section-label">📍 Workflow Steps</p>', unsafe_allow_html=True)
        step_labels = {1: "🌐 Universe", 2: "📊 Screener", 3: "⚖️ Rebalance", 4: "💾 Export"}
        for s, lbl in step_labels.items():
            is_active = (st.session_state.current_step == s)
            is_done   = (s == 1 and st.session_state.symbols is not None) or \
                        (s == 2 and st.session_state.screener_done) or \
                        (s == 3 and st.session_state.rebalance_done)
            btn_label = f"{'✅' if is_done else ('▶' if is_active else '○')} {lbl}"
            if st.button(btn_label, key=f"sb_nav_{s}", use_container_width=True,
                         type="primary" if is_active else "secondary"):
                st.session_state.current_step = s
                st.session_state.main_view = "screener"
                st.rerun()

        st.divider()

        # ── Quick Links ─────────────────────────────────────
        st.markdown('<p class="sidebar-section-label">🔗 Quick Links</p>', unsafe_allow_html=True)
        st.markdown(f"""
        <div style="font-size:12px;line-height:2.6;">
        <a href="https://www.nseindia.com/static/market-data/securities-available-for-trading"
           target="_blank" style="color:var(--blue);text-decoration:none;font-weight:500;">
           📥 NSE EQUITY_L.csv</a><br>
        <a href="{APPS_SCRIPT_URL}" target="_blank"
           style="color:var(--blue);text-decoration:none;font-weight:500;">
           ⚖️ Portfolio Rebalancer</a><br>
        <a href="https://prayan2702.github.io/momn-dashboard/" target="_blank"
           style="color:var(--blue);text-decoration:none;font-weight:500;">
           📈 Portfolio Dashboard</a>
        </div>
        """, unsafe_allow_html=True)

        st.divider()

        if st.button("🚪 Logout", use_container_width=True):
            for k in list(st.session_state.keys()):
                del st.session_state[k]
            st.rerun()


# ═══════════════════════════════════════════════════════════════
# ▸ LOGIN PAGE
# ═══════════════════════════════════════════════════════════════
def login_page():
    render_app_header()
    _, mid, _ = st.columns([1, 1.2, 1])
    with mid:
        st.markdown("""
        <div style="margin-top:40px;">
          <div class="login-card">
            <div class="login-icon">📈</div>
            <div class="login-title">Welcome back</div>
            <div class="login-sub">NSE Momentum Screener &nbsp;·&nbsp; v14</div>
          </div>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("<div style='height:1px;'></div>", unsafe_allow_html=True)
        with st.form(key="login_form", clear_on_submit=True):
            u = st.text_input("Username", placeholder="Enter username")
            p = st.text_input("Password", type="password", placeholder="Enter password")
            if st.form_submit_button("Sign In →", use_container_width=True, type="primary"):
                if u == USERNAME and p == PASSWORD:
                    st.session_state.logged_in = True
                    st.rerun()
                else:
                    st.error("Invalid username or password")


if not st.session_state.logged_in:
    login_page()
    st.stop()


# ═══════════════════════════════════════════════════════════════
# ▸ AUTHENTICATED LAYOUT
# ═══════════════════════════════════════════════════════════════
render_app_header()
render_top_nav()
render_sidebar()


# ═══════════════════════════════════════════════════════════════
# ▸ TEARSHEET VIEW
# ═══════════════════════════════════════════════════════════════
if st.session_state.main_view == "tearsheet":
    st.markdown("""
    <div class="tearsheet-header">
      <div class="tearsheet-title">📋 Strategy Tearsheet</div>
      <div class="tearsheet-sub">
        NAV-based QuantStats performance report vs Nifty 50 benchmark
      </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Integrate strategy-tearsheet.py logic ────────────────
    try:
        import quantstats as qs

        @st.cache_data(ttl=3600, show_spinner=False)
        def _load_tearsheet_data(url):
            return pd.read_csv(url)

        @st.cache_data(ttl=3600, show_spinner=False)
        def _build_tearsheet(csv_url):
            data = pd.read_csv(csv_url)
            # Remove unwanted rows
            rows_to_delete = data[data['Date'].isin(
                ['Portfolio Value', 'Absolute Gain', 'Nifty50', 'Day Change']
            )].index
            data.drop(rows_to_delete, inplace=True)
            data = data.dropna(subset=['NAV'])
            data['Date'] = pd.to_datetime(data['Date'], format='%d-%b-%y')
            data = data.sort_values(by='Date')
            data['Date'] = data['Date'].apply(lambda x: x.replace(tzinfo=None))
            data.set_index('Date', inplace=True)
            data['NAV'] = pd.to_numeric(data['NAV'])
            data['Nifty50 Change %'] = data['Nifty50 Change %'].str.rstrip('%').astype('float') / 100
            data['Nifty50 NAV'] = (1 + data['Nifty50 Change %']).cumprod()
            returns = data['NAV'].pct_change().dropna()
            nifty50 = data['Nifty50 Change %'].dropna()
            # Align date ranges
            start = max(returns.index[0], nifty50.index[0])
            end   = min(returns.index[-1], nifty50.index[-1])
            returns = returns[start:end]
            nifty50 = nifty50[start:end]
            returns = returns.replace([np.inf, -np.inf], 0).fillna(0)
            nifty50 = nifty50.replace([np.inf, -np.inf], 0).fillna(0)
            return returns, nifty50

        col_ts, col_reload = st.columns([4, 1])
        with col_reload:
            reload_ts = st.button("🔄 Refresh Data", type="secondary")
        if reload_ts:
            st.cache_data.clear()

        with st.spinner("Loading portfolio NAV data…"):
            returns, nifty50 = _build_tearsheet(TEARSHEET_CSV_URL)

        # ── Quick KPI strip ──────────────────────────────────
        total_ret = (1 + returns).prod() - 1
        ann_ret   = (1 + total_ret) ** (252 / max(len(returns), 1)) - 1
        volatility = returns.std() * (252 ** 0.5)
        sharpe     = (returns.mean() / returns.std()) * (252 ** 0.5) if returns.std() > 0 else 0
        max_dd     = (returns + 1).cumprod().div((returns + 1).cumprod().cummax()).sub(1).min()

        st.markdown(f"""<div class="metric-row">
            {metric_card("Total Return",  f"{total_ret*100:.1f}%",  "blue" if total_ret >= 0 else "red")}
            {metric_card("Ann. Return",   f"{ann_ret*100:.1f}%",   "green" if ann_ret >= 0 else "red")}
            {metric_card("Volatility",    f"{volatility*100:.1f}%","amber")}
            {metric_card("Sharpe Ratio",  f"{sharpe:.2f}",         "violet")}
            {metric_card("Max Drawdown",  f"{max_dd*100:.1f}%",    "red")}
        </div>""", unsafe_allow_html=True)

        # ── Full QuantStats HTML report ──────────────────────
        st.markdown('<div class="section-hdr">Full QuantStats Report</div>', unsafe_allow_html=True)
        with st.spinner("Generating tearsheet…"):
            qs.reports.html(returns, nifty50, output="tearsheet_report.html")
            with open("tearsheet_report.html", "r", encoding="utf-8") as f:
                report_html = f.read()
        st.components.v1.html(report_html, height=900, scrolling=True)

    except ImportError:
        st.error(
            "❌ `quantstats` package not found. "
            "`requirements.txt` mein `quantstats` add karo aur redeploy karo."
        )
    except Exception as e:
        st.error(f"Tearsheet error: {e}")

    render_footer()
    st.stop()


# ═══════════════════════════════════════════════════════════════
# ▸ SCREENER WORKFLOW VIEW
# ═══════════════════════════════════════════════════════════════

# Step progress bar
st.markdown(render_step_bar(st.session_state.current_step), unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════
# STEP 1 — UNIVERSE SETUP
# ═══════════════════════════════════════════════════════════════
if st.session_state.current_step == 1:
    st.markdown('<div class="section-hdr">🌐 Universe Setup</div>', unsafe_allow_html=True)

    c1, c2 = st.columns([1, 2])
    with c1:
        chosen_u = st.selectbox(
            "Select Universe",
            UNIVERSES,
            index=UNIVERSES.index(st.session_state.universe),
            help="AllNSE = NSE ki sabhi EQ stocks. Baaki = Nifty index lists (GitHub se auto-load)"
        )
        st.session_state.universe = chosen_u
    with c2:
        _u_meta = {
            "Nifty50":  ("50 stocks", "🔵", "Large Cap — India ke top 50"),
            "Nifty100": ("100 stocks","🟢", "Large Cap — top 100"),
            "Nifty200": ("200 stocks","🟡", "Large + Mid Cap"),
            "Nifty250": ("250 stocks","🟠", "Mid Cap focused"),
            "Nifty500": ("500 stocks","🔴", "Large + Mid + Small"),
            "N750":     ("750 stocks","🟣", "Total Market index"),
            "AllNSE":   ("2000+ stocks","⚪","Sabhi NSE EQ stocks"),
        }
        _m = _u_meta.get(chosen_u, ("—","⚪",""))
        st.markdown(f"""
        <div class="universe-card">
          <div class="u-icon">{_m[1]}</div>
          <div>
            <div class="u-name">{chosen_u}</div>
            <div class="u-meta">
              <span class="badge badge-blue">{_m[0]}</span>
              &nbsp; {_m[2]}
            </div>
          </div>
        </div>
        """, unsafe_allow_html=True)

    # ── AllNSE: Load Symbol List ──────────────────────────────
    if chosen_u == "AllNSE":
        _NSE_EQUITY_URL = "https://archives.nseindia.com/content/equities/EQUITY_L.csv"
        _NSE_HEADERS = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            ),
            "Accept": "text/html,application/xhtml+xml,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.5",
            "Referer": "https://www.nseindia.com/",
        }

        def _fetch_nse_equity_csv():
            import io as _io
            session = requests.Session()
            session.get("https://www.nseindia.com", headers=_NSE_HEADERS, timeout=15)
            time.sleep(1)
            resp = session.get(_NSE_EQUITY_URL, headers=_NSE_HEADERS, timeout=30)
            resp.raise_for_status()
            df = pd.read_csv(_io.StringIO(resp.text), skipinitialspace=True)
            df.columns = [c.strip() for c in df.columns]
            if 'SERIES' in df.columns:
                df = df[df['SERIES'].str.strip() == 'EQ'].copy()
            df['SYMBOL'] = df['SYMBOL'].str.strip().str.upper()
            return df.reset_index(drop=True)

        if st.session_state.symbols and st.session_state.universe == "AllNSE":
            n = len(st.session_state.symbols)
            st.markdown(f"""<div class="metric-row">
                {metric_card("Loaded Symbols", f"{n:,}", "green")}
                {metric_card("Source", st.session_state.get("allnse_source","NSE"), "blue")}
            </div>""", unsafe_allow_html=True)
            if st.button("🔄 Reload Symbol List", type="secondary"):
                st.session_state.symbols = None
                st.session_state.eq_df   = None
                st.rerun()
        else:
            st.markdown("""
            <div class="nse-link-box">
              <div style="font-size:20px;">📥</div>
              <div>
                <div style="font-weight:600;font-size:13px;">NSE — Securities Available for Trading</div>
                <div class="hint">Button dabao → NSE website se auto-fetch hoga &nbsp;|&nbsp; Ya manually CSV browse karo</div>
              </div>
            </div>
            """, unsafe_allow_html=True)

            btn_col, _ = st.columns([1, 2])
            with btn_col:
                load_btn = st.button("📡 Load from NSE (Auto)", type="primary",
                                     use_container_width=True)

            if load_btn:
                with st.spinner("🌐 NSE website se EQUITY_L.csv fetch ho raha hai..."):
                    try:
                        eq_df = _fetch_nse_equity_csv()
                        syms_ns = [s + ".NS" for s in eq_df["SYMBOL"].tolist()]
                        syms_ns = add_extra_symbols(syms_ns)
                        st.session_state.eq_df  = eq_df
                        st.session_state.symbols = syms_ns
                        st.session_state.universe_label = f"AllNSE (NSE — {len(syms_ns):,} stocks)"
                        st.session_state["allnse_source"] = "NSE Live"
                        st.success(
                            f"✅ NSE se {len(syms_ns):,} EQ stocks loaded "
                            f"(incl. GOLDBEES & SILVERBEES)"
                        )
                        st.rerun()
                    except Exception as e:
                        st.warning(
                            f"⚠️ NSE auto-fetch failed: `{e}`\n\n"
                            "Manually EQUITY_L.csv download karein aur neeche browse karein."
                        )
                        st.session_state["_nse_fetch_failed"] = True

            show_uploader = st.session_state.get("_nse_fetch_failed", False)
            if not load_btn:
                show_uploader = True

            if show_uploader:
                st.markdown(
                    f'<a href="https://www.nseindia.com/static/market-data/securities-available-for-trading" '
                    f'target="_blank" style="font-size:12px;color:var(--blue);">📥 NSE se manually EQUITY_L.csv download karein</a>',
                    unsafe_allow_html=True
                )
                uploaded = st.file_uploader(
                    "📂 EQUITY_L.csv browse karein (manual fallback)",
                    type=["csv"], key="equity_csv"
                )
                if uploaded:
                    try:
                        eq_df = parse_equity_csv(uploaded)
                        syms_ns = [s + ".NS" for s in eq_df["SYMBOL"].tolist()]
                        syms_ns = add_extra_symbols(syms_ns)
                        st.session_state.eq_df  = eq_df
                        st.session_state.symbols = syms_ns
                        st.session_state.universe_label = f"AllNSE (CSV — {len(syms_ns):,} stocks)"
                        st.session_state["allnse_source"] = "CSV Upload"
                        st.success(f"✅ CSV loaded: **{len(syms_ns):,}** EQ stocks (incl. GOLDBEES & SILVERBEES)")
                        st.rerun()
                    except Exception as e:
                        st.error(f"CSV parse error: {e}")

            if not st.session_state.symbols:
                st.info("💡 Symbol list load nahi hua — GitHub fallback (NSE_EQ_ALL.csv) screener run pe use hoga.")
                st.markdown("""
                <div style="background:var(--amber-bg);border:1px solid var(--amber-bdr);border-radius:var(--r-md);
                            padding:10px 16px;font-size:12.5px;color:#92400e;margin-top:6px;">
                ➕ <b>Auto-included:</b> &nbsp;
                <span class="badge badge-amber">🥇 GOLDBEES</span>
                &nbsp;
                <span class="badge badge-amber">🥈 SILVERBEES</span>
                &nbsp; — har universe ke saath automatically add honge
                </div>
                """, unsafe_allow_html=True)

    # ── Other universes ───────────────────────────────────────
    else:
        st.info(f"📡 **{chosen_u}** ki symbol list screener run pe GitHub se auto-load hogi.")
        st.markdown("""
        <div style="background:var(--amber-bg);border:1px solid var(--amber-bdr);border-radius:var(--r-md);
                    padding:10px 16px;font-size:12.5px;color:#92400e;margin-top:6px;">
        ➕ <b>Auto-included:</b> &nbsp;
        <span class="badge badge-amber">🥇 GOLDBEES</span>
        &nbsp;
        <span class="badge badge-amber">🥈 SILVERBEES</span>
        &nbsp; — har universe ke saath automatically add honge
        </div>
        """, unsafe_allow_html=True)
        if st.button("✅ Load Symbol List", type="primary"):
            with st.spinner(f"Loading {chosen_u} from GitHub…"):
                try:
                    syms_ns = load_symbols_from_github(chosen_u)
                    syms_ns = add_extra_symbols(syms_ns)
                    st.session_state.symbols = syms_ns
                    st.session_state.universe_label = f"{chosen_u} ({len(syms_ns)} stocks)"
                    st.success(f"✅ {chosen_u}: **{len(syms_ns)}** symbols loaded (incl. GOLDBEES & SILVERBEES)")
                except Exception as e:
                    st.error(f"Symbol load failed: {e}")

        if st.session_state.symbols and st.session_state.universe == chosen_u:
            n = len(st.session_state.symbols)
            st.markdown(f"""<div class="metric-row">
                {metric_card("Loaded Symbols", f"{n:,}", "green")}
                {metric_card("Universe", chosen_u, "blue")}
            </div>""", unsafe_allow_html=True)

    st.divider()

    if _CACHE_AVAILABLE:
        st.markdown(get_cache_status_html(), unsafe_allow_html=True)
    if _CACHE_UPSTOX_AVAILABLE:
        st.markdown(get_cache_status_html_upstox(), unsafe_allow_html=True)

    if st.session_state.symbols or chosen_u != "AllNSE":
        if st.button("▶ Next: Run Screener →", type="primary"):
            st.session_state.current_step = 2; st.rerun()
    elif chosen_u == "AllNSE" and not st.session_state.symbols:
        if st.button("▶ Next: Run Screener → (GitHub fallback)", type="secondary"):
            st.session_state.current_step = 2; st.rerun()

    render_footer()


# ═══════════════════════════════════════════════════════════════
# STEP 2 — RUN SCREENER
# ═══════════════════════════════════════════════════════════════
elif st.session_state.current_step == 2:
    st.markdown('<div class="section-hdr">📊 Run Momentum Screener</div>', unsafe_allow_html=True)

    if not _CALCS_AVAILABLE:
        st.error("❌ `calculations.py` not found. Project folder mein rakh kar dobara run karo.")
        st.stop()
    if not _DS_AVAILABLE:
        st.warning(
            f"⚠️ `data_service.py` import failed (`{_DS_IMPORT_ERR[:100]}`). "
            "**YFinance** inline fallback use hoga."
        )

    _downloading = st.session_state.get("_downloading", False)

    with st.expander("🔧 Filter Settings", expanded=not _downloading):
        if _downloading:
            st.info("⏳ Data download chal raha hai — filters locked hain.")
            _fp = st.session_state.get("_last_filter_params", {})
            st.markdown(
                f"Close > 200-DMA: **{_fp.get('use_dma200',True)}** &nbsp;|&nbsp; "
                f"12M ROC > 5.5%: **{_fp.get('use_roc12',True)}** &nbsp;|&nbsp; "
                f"Avg Vol: **{_fp.get('volm_cr_min',1.0)} Cr** &nbsp;|&nbsp; "
                f"Min CMP: **₹{_fp.get('close_min',30.0)}**"
            )
            filter_params = _fp
        else:
            f_left, f_right = st.columns([1, 1])
            with f_left:
                st.markdown("**✅ Filters**")
                use_dma200 = st.checkbox("Close > 200-day DMA",    value=True)
                use_roc12  = st.checkbox("12M ROC > 5.5%",         value=True)
                use_roc_cap= st.checkbox("12M return < 1000x",     value=True)
                use_ath    = st.checkbox("Within 25% of ATH",      value=True)
            with f_right:
                st.markdown("**📊 Thresholds**")
                volm_min    = st.slider("Avg Vol (Cr) >",   0.0, 10.0, 1.0, 0.1)
                close_min   = st.slider("Min CMP ₹",        0.0, 500.0, 30.0, 5.0)
                circuit_max = st.slider("Circuit hits/yr <", 1, 100, 20, 1)
                circuit5    = st.slider("5% circuit 3M ≤",  0, 30, 10, 1)

            filter_params = {
                "use_dma200": use_dma200, "use_roc12": use_roc12, "use_roc_cap": use_roc_cap,
                "volm_cr_min": volm_min, "circuit_max": circuit_max, "circuit5_max": circuit5,
                "use_away_ath": use_ath, "close_min": close_min,
            }
            st.session_state["_last_filter_params"] = filter_params

    U          = st.session_state.universe
    api_source = st.session_state.data_source
    end_date   = st.session_state.lookback_date

    col_run, col_info = st.columns([1, 2])
    with col_run:
        _running = st.session_state.get("_run_download", False)
        run_clicked = st.button(
            "⏳ Downloading..." if _running else "▶ Start Data Download",
            type="primary",
            use_container_width=True,
            disabled=_running,
        )
    with col_info:
        n_loaded = len(st.session_state.symbols) if st.session_state.symbols else "—"
        st.markdown(f"""
        <div class="run-info-strip">
        🌐 Universe: <b>{U}</b> &nbsp;|&nbsp;
        📋 Symbols: <b style="color:var(--blue)">{n_loaded}</b> &nbsp;|&nbsp;
        📅 End: <b>{end_date.strftime('%d-%m-%Y')}</b><br>
        📐 Method: <b style="color:var(--violet)">{st.session_state.ranking_method}</b> &nbsp;|&nbsp;
        📡 Source: <b style="color:var(--teal)">{api_source}</b>
        </div>""", unsafe_allow_html=True)

    if run_clicked:
        st.session_state["_run_download"]      = True
        st.session_state["_run_filter_params"] = filter_params
        st.session_state["_run_api_source"]    = api_source
        st.session_state["_run_end_date"]      = end_date
        st.session_state["_run_u"]             = U
        st.rerun()

    if st.session_state.get("_run_download", False):
        _filter_params = st.session_state.get("_run_filter_params", filter_params)
        _api_source    = st.session_state.get("_run_api_source",    api_source)
        _end_date      = st.session_state.get("_run_end_date",      end_date)
        _U             = st.session_state.get("_run_u",             U)

        dates     = build_dates(_end_date)
        prog_bar  = st.progress(0)
        status_tx = st.empty()
        _download_ok = False
        try:

            # ═══════════════════════════════════════════════════
            # BRANCH A — Pre-cached (Instant load from GitHub)
            # ═══════════════════════════════════════════════════
            if _api_source in ("📦 Pre-cached YFinance", "📦 Pre-cached Upstox"):
                _is_upstox_cache = (_api_source == "📦 Pre-cached Upstox")

                if _is_upstox_cache:
                    if not _CACHE_UPSTOX_AVAILABLE:
                        st.error("❌ cache_loader_upstox.py nahi mila. Repo mein add karo.")
                        st.stop()
                    _loader    = load_cache_upstox
                    _meta_fn   = get_cache_meta_upstox
                    _age_fn    = get_cache_age_days_upstox
                    _cache_lbl = "Upstox"
                else:
                    if not _CACHE_AVAILABLE:
                        st.error("❌ cache_loader.py nahi mila. cache_loader.py repo mein add karo.")
                        st.stop()
                    _loader    = load_cache
                    _meta_fn   = get_cache_meta
                    _age_fn    = get_cache_age_days
                    _cache_lbl = "YFinance"

                status_tx.markdown(f"⚡ **{_cache_lbl} pre-cached data GitHub se load ho raha hai...**")
                prog_bar.progress(0.1)
                try:
                    close, high, volume = _loader()
                    prog_bar.progress(0.85)
                    status_tx.markdown("✅ **Cache loaded!** Calculations shuru ho rahi hain...")

                    meta = _meta_fn()
                    failed_from_meta = meta.get("failed_symbols", [])
                    failed_from_meta = [s.replace(".NS", "") for s in failed_from_meta]

                    volume12M_check = volume.loc[dates['date12M']:].copy() if not volume.empty else pd.DataFrame()
                    median_volume   = volume12M_check.median() if not volume12M_check.empty else pd.Series()
                    vol_blank       = median_volume[median_volume.isna()].index.tolist()
                    vol_blank       = [t.replace('.NS', '') for t in vol_blank]

                    failed_blank = list(dict.fromkeys(failed_from_meta + vol_blank))
                    st.session_state.failed_blank = failed_blank

                    age = _age_fn()
                    if age > 3:
                        st.warning(
                            f"⚠️ {_cache_lbl} Cache {int(age)} din purana hai "
                            f"(build: {meta.get('build_date','?')}). "
                            "Data slightly stale ho sakta hai."
                        )
                    else:
                        st.success(
                            f"✅ {_cache_lbl} Cache loaded! "
                            f"{meta.get('symbols_fetched','?'):,} symbols | "
                            f"Build: {meta.get('build_date','?')} | "
                            f"Age: {int(age)} din"
                        )

                    _universe_syms = st.session_state.symbols or []
                    _cache_syms    = set(close.columns.str.replace('.NS', '', regex=False).str.upper())
                    _missing = [
                        s for s in _universe_syms
                        if s.replace('.NS', '').upper() not in _cache_syms
                    ]
                    if _missing:
                        st.info(
                            f"ℹ️ Cache mein **{len(_missing)} stocks missing** hain "
                            f"(universe: {len(_universe_syms)}, cache: {len(_cache_syms)}). "
                            f"YFinance se fetch karke merge kar sakte hain."
                        )
                        st.session_state["_cache_missing_syms"] = _missing
                        if st.button(
                            f"📡 Fetch {len(_missing)} missing stocks from YFinance & merge",
                            key="fetch_missing_btn", type="secondary"
                        ):
                            with st.spinner(f"YFinance se {len(_missing)} missing stocks fetch ho rahi hain..."):
                                try:
                                    _m_close, _m_high, _m_vol, _m_failed = _fetch_yfinance_inline(
                                        _missing, dates['startDate'], dates['endDate'],
                                        prog_bar, status_tx, chunk_size=15
                                    )
                                    if _m_close is not None and not _m_close.empty:
                                        _all_idx = close.index.union(_m_close.index)
                                        close  = close.reindex(_all_idx).combine_first(_m_close.reindex(_all_idx))
                                        high   = high.reindex(_all_idx).combine_first(_m_high.reindex(_all_idx))
                                        volume = volume.reindex(_all_idx).combine_first(_m_vol.reindex(_all_idx))
                                        st.success(
                                            f"✅ {_m_close.shape[1] - len(_m_failed)} missing stocks merged! "
                                            f"Total: {close.shape[1]:,} symbols"
                                        )
                                    if _m_failed:
                                        _fb_extra = [t.replace('.NS', '') for t in _m_failed]
                                        failed_blank = list(dict.fromkeys(failed_blank + _fb_extra))
                                        st.session_state.failed_blank = failed_blank
                                except Exception as _me:
                                    st.warning(f"Missing stocks fetch failed: {_me}")

                except Exception as e:
                    st.error(f"❌ Cache load failed: {e}. YFinance select karke retry karo.")
                    st.stop()

            # ═══════════════════════════════════════════════════
            # BRANCH B — Live fetch
            # ═══════════════════════════════════════════════════
            else:
                if st.session_state.symbols is None or st.session_state.universe != U:
                    with st.spinner(f"Loading {_U} symbols…"):
                        try:
                            if _U == "AllNSE":
                                url = f"{GITHUB_BASE}/NSE_EQ_ALL.csv"
                                df_sym = pd.read_csv(url)
                                df_sym['Yahoo_Symbol'] = df_sym['Symbol'].astype(str).str.strip() + '.NS'
                                syms_ns = df_sym['Yahoo_Symbol'].tolist()
                            else:
                                syms_ns = load_symbols_from_github(_U)
                            syms_ns = add_extra_symbols(syms_ns)
                            st.session_state.symbols = syms_ns
                        except Exception as e:
                            st.error(f"Symbol list load failed: {e}"); st.stop()

                symbols = st.session_state.symbols
                CHUNK   = 50 if _api_source == "Upstox" else (15 if _U == "AllNSE" else 50)
                st.markdown(f"""
                <div style="display:flex;gap:8px;flex-wrap:wrap;margin:8px 0 12px 0;">
                  <span class="badge badge-blue">📦 Chunk: {CHUNK}</span>
                  <span class="badge badge-green">📋 Symbols: {len(symbols):,}</span>
                  <span class="badge badge-violet">📡 Source: {_api_source}</span>
                  <span class="badge badge-amber">🥇 GOLDBEES &amp; 🥈 SILVERBEES included</span>
                </div>
                """, unsafe_allow_html=True)

                _use_ds = _DS_AVAILABLE and _api_source in ("Upstox", "Angel One")
                try:
                    if _use_ds:
                        close, high, volume, failed_symbols = fetch_data(
                            api_source   = _api_source,
                            symbols      = symbols,
                            start_date   = dates['startDate'],
                            end_date     = dates['endDate'],
                            chunk_size   = CHUNK,
                            progress_bar = prog_bar,
                            status_text  = status_tx,
                        )
                    else:
                        close, high, volume, failed_symbols = _fetch_yfinance_inline(
                            symbols, dates['startDate'], dates['endDate'],
                            prog_bar, status_tx, chunk_size=CHUNK
                        )
                except Exception as e:
                    st.error(f"Data fetch error: {e}"); st.stop()

                if close is None or close.empty:
                    st.session_state["_downloading"] = False
                    st.error("❌ Data fetch hua nahi. Internet / token check karo."); st.stop()

                api_failed = [t.replace('.NS','') for t in (failed_symbols or [])]
                volume12M_check = volume.loc[dates['date12M']:].copy() if not volume.empty else pd.DataFrame()
                median_volume   = volume12M_check.median() if not volume12M_check.empty else pd.Series()
                vol_blank       = median_volume[median_volume.isna()].index.tolist()
                vol_blank       = [t.replace('.NS','') for t in vol_blank]
                failed_blank = list(dict.fromkeys(api_failed + vol_blank))
                st.session_state.failed_blank = failed_blank

            # ── Calculate metrics ─────────────────────────────
            status_tx.markdown("⏳ **Calculating momentum metrics...**")
            prog_bar.progress(0.92)
            try:
                dfStats   = build_dfStats(close, high, volume, dates, st.session_state.ranking_method)
                dfFiltered = apply_filters(dfStats.copy(), _filter_params)
                st.session_state.dfStats    = dfStats
                st.session_state.dfFiltered = dfFiltered
                st.session_state.screener_done = True
                st.session_state["_downloading"] = False
                prog_bar.progress(1.0)
                status_tx.markdown("✅ **Screener complete!**")
                _download_ok = True

            except Exception as e:
                _cls = type(e).__name__
                if "StopException" not in _cls and "RerunException" not in _cls:
                    st.error(f"Download/Calculation error: {e}")
                raise

        finally:
            st.session_state["_run_download"] = False

        if _download_ok:
            st.rerun()

    # ── Display results ───────────────────────────────────────
    if st.session_state.screener_done and st.session_state.dfFiltered is not None:
        dfF = st.session_state.dfFiltered
        dfU = st.session_state.dfStats
        n_f = len(dfF); n_u = len(dfU) if dfU is not None else 0
        top_n = st.session_state.top_n_rank
        rank_col = st.session_state.ranking_method

        st.markdown(f"""<div class="metric-row">
            {metric_card("Total Screened", f"{n_u:,}")}
            {metric_card("Passed Filters", f"{n_f:,}", "green")}
            {metric_card("Top-N Universe", f"Top {top_n}", "blue")}
            {metric_card("End Date", st.session_state.lookback_date.strftime('%d %b %Y'), "amber")}
        </div>""", unsafe_allow_html=True)

        _fb = st.session_state.failed_blank or []
        if _fb:
            with st.expander(f"⚠️ {len(_fb)} stocks failed to download — click to view", expanded=False):
                st.dataframe(
                    pd.DataFrame({'S.No.': range(1, len(_fb)+1),
                                  'Failed Stock': _fb}).set_index('S.No.'),
                    use_container_width=False
                )
        else:
            st.success("✅ All stocks downloaded successfully!")

        tab1, tab2 = st.tabs(["✅ Filtered (Top Ranked)", "📊 All Stocks (Unfiltered)"])
        with tab1:
            top_view = dfF.head(top_n).reset_index()
            dcols = ["Rank","Ticker","Close",rank_col,"roc12M","roc6M","roc3M","vol12M","volm_cr","AWAY_ATH","circuit","dma200d"]
            dcols = [c for c in dcols if c in top_view.columns]
            st.dataframe(top_view[dcols].style.format(precision=2),
                         use_container_width=True, height=440)
        with tab2:
            if dfU is not None:
                st.dataframe(dfU.reset_index().head(300).style.format(precision=2),
                             use_container_width=True, height=440)

        st.divider()
        if st.button("▶ Next: Plan Rebalance →", type="primary"):
            st.session_state.current_step = 3; st.rerun()

    elif not st.session_state.screener_done:
        st.markdown("""
        <div class="card" style="text-align:center;padding:40px 20px;">
          <div style="font-size:36px;margin-bottom:12px;">📊</div>
          <div style="font-size:15px;font-weight:600;color:var(--ink-mid);margin-bottom:6px;">
            Ready to screen
          </div>
          <div style="font-size:13px;color:var(--muted);">
            Settings configure karo aur "Start Data Download" dabao
          </div>
        </div>
        """, unsafe_allow_html=True)

    render_footer()


# ═══════════════════════════════════════════════════════════════
# STEP 3 — PLAN REBALANCE
# ═══════════════════════════════════════════════════════════════
elif st.session_state.current_step == 3:
    st.markdown('<div class="section-hdr">⚖️ Plan Rebalance</div>', unsafe_allow_html=True)

    if not st.session_state.screener_done:
        st.warning("⚠️ Pehle Step 2 mein screener run karo.")
        if st.button("← Step 2 par jao"): st.session_state.current_step = 2; st.rerun()
        st.stop()

    port_source = st.radio("Portfolio data source",
                           ["📊 Google Sheet (auto)", "📂 CSV manually upload"],
                           horizontal=True)

    if "📊" in port_source:
        col_load, _ = st.columns([1, 2])
        with col_load:
            if st.button("🔄 Fetch from Google Sheet", type="primary"):
                with st.spinner("Fetching portfolio..."):
                    try:
                        pdf = pd.read_csv(PORTFOLIO_CSV_URL)
                        if 'Current Portfolio' in pdf.columns:
                            st.session_state.reb_portfolio = [
                                str(x).strip().upper() for x in pdf['Current Portfolio'].dropna()
                                if str(x).strip() and str(x).strip().lower() not in ('nan','current portfolio','')
                            ]
                            st.success(f"✅ Portfolio loaded: **{len(st.session_state.reb_portfolio)}** stocks")
                        else:
                            st.error("Column 'Current Portfolio' not found in sheet.")
                    except Exception as e:
                        st.error(f"Sheet fetch failed: {e}")
    else:
        up_reb = st.file_uploader("📂 Upload Portfolio CSV", type=["csv"], key="reb_csv")
        if up_reb:
            try:
                df_reb = pd.read_csv(up_reb)
                df_reb.columns = [c.strip() for c in df_reb.columns]
                col_b = df_reb.columns[1] if len(df_reb.columns) > 1 else df_reb.columns[0]
                st.session_state.reb_portfolio = [
                    str(x).strip().upper() for x in df_reb[col_b].dropna()
                    if str(x).strip() and len(str(x).strip()) > 1
                ]
                st.success(f"✅ Portfolio loaded: {len(st.session_state.reb_portfolio)} stocks")
            except Exception as e:
                st.error(f"CSV parse error: {e}")

    with st.expander("✏️ Manual Edit (comma-separated)", expanded=False):
        port_text = st.text_area("Current Portfolio",
                                  value=", ".join(st.session_state.reb_portfolio or []), height=100)
        if st.button("Apply Manual Edit"):
            st.session_state.reb_portfolio = [
                s.strip().upper() for s in port_text.split(",") if s.strip()
            ]
            st.success("Updated!")

    portfolio = st.session_state.reb_portfolio or []

    # ── Compute rebalance ─────────────────────────────────────
    if portfolio and st.session_state.dfFiltered is not None:
        dfFiltered      = st.session_state.dfFiltered
        dfStats         = st.session_state.dfStats
        top_n           = st.session_state.top_n_rank
        rank_threshold  = top_n

        top_rank_tickers = dfFiltered.reset_index()
        top_rank_tickers = top_rank_tickers[top_rank_tickers['Rank'] <= rank_threshold]['Ticker']

        current_portfolio_tickers = pd.Series(portfolio)
        entry_stocks = top_rank_tickers[~top_rank_tickers.isin(current_portfolio_tickers)]
        exit_stocks  = current_portfolio_tickers[~current_portfolio_tickers.isin(top_rank_tickers)]
        hold_stocks  = current_portfolio_tickers[current_portfolio_tickers.isin(top_rank_tickers)]

        num_sells = len(exit_stocks)
        entry_stocks = entry_stocks.head(num_sells)

        if len(entry_stocks) < num_sells:
            entry_stocks = pd.concat([
                entry_stocks,
                pd.Series([None] * (num_sells - len(entry_stocks)))
            ])

        # ── Reasons for exit (v10 logic unchanged) ────────────
        reasons_for_exit = []
        for ticker in exit_stocks:
            if pd.isna(ticker) or ticker == "":
                reasons_for_exit.append(""); continue
            reasons    = []
            stock_data = dfStats[dfStats['Ticker'] == ticker] if dfStats is not None else pd.DataFrame()
            if len(stock_data) > 0:
                if stock_data.index[0] > rank_threshold:          reasons.append(f"Rank > {rank_threshold}")
                if stock_data['volm_cr'].values[0] <= 1:           reasons.append("Volume ≤ 1 Cr")
                if stock_data['Close'].values[0] <= stock_data['dma200d'].values[0]:
                                                                   reasons.append("Close ≤ 200-DMA")
                if stock_data['roc12M'].values[0] <= 5.5:          reasons.append("12M ROC ≤ 5.5%")
                if stock_data['circuit'].values[0] >= 20:          reasons.append("Circuit ≥ 20")
                if stock_data['AWAY_ATH'].values[0] <= -25:        reasons.append("Away ATH ≤ -25%")
                if stock_data['roc12M'].values[0] >= 1000:         reasons.append("12M ROC ≥ 1000%")
                if stock_data['Close'].values[0] <= 30:            reasons.append("Close ≤ ₹30")
                if stock_data['circuit5'].values[0] > 10:          reasons.append("5% Circuit > 10")
            else:
                reasons.append("Not in selected universe")
            reasons_for_exit.append(", ".join(reasons) if reasons else "Rank dropped")

        reasons_for_exit.extend([""] * (len(entry_stocks) - len(reasons_for_exit)))

        rebalance_table = pd.DataFrame({
            'S.No.':           range(1, num_sells + 1),
            'Sell Stocks':     exit_stocks.tolist(),
            'Buy Stocks':      entry_stocks.tolist(),
            'Reason for Exit': reasons_for_exit,
        })
        rebalance_table = rebalance_table[
            ~(rebalance_table['Sell Stocks'].isna() & rebalance_table['Buy Stocks'].isna())
        ]
        rebalance_table.set_index('S.No.', inplace=True)
        st.session_state.sell_list = exit_stocks.dropna().tolist()
        st.session_state.buy_list  = entry_stocks.dropna().tolist()
        st.session_state.rebalance_table = rebalance_table
        st.session_state.rebalance_done  = True

        # ── Summary strip ──────────────────────────────────────
        st.markdown(f"""<div class="reb-strip">
          <div class="reb-stat"><div class="label">Portfolio</div><div class="val b">{len(portfolio)}</div></div>
          <div class="reb-stat"><div class="label">Top-{rank_threshold} Screener</div><div class="val b">{len(top_rank_tickers)}</div></div>
          <div class="reb-stat"><div class="label">SELL (Exit)</div><div class="val r">{len(exit_stocks)}</div></div>
          <div class="reb-stat"><div class="label">BUY (Entry)</div><div class="val g">{len(entry_stocks.dropna())}</div></div>
          <div class="reb-stat"><div class="label">HOLD</div><div class="val p">{len(hold_stocks)}</div></div>
        </div>""", unsafe_allow_html=True)

        # ── Sell / Buy / Hold columns ──────────────────────────
        col_sell, col_buy, col_hold = st.columns(3)
        with col_sell:
            st.markdown('<div class="section-hdr" style="border-left-color:var(--red)">🔴 SELL List</div>', unsafe_allow_html=True)
            sell_list = exit_stocks.dropna().tolist()
            if sell_list:
                chips = " ".join([f'<span class="chip chip-sell">{s}</span>' for s in sell_list])
                st.markdown(chips, unsafe_allow_html=True)
                cmp_map = {}
                if dfStats is not None:
                    cmp_map = dict(zip(dfStats['Ticker'], dfStats['Close']))
                if dfFiltered is not None:
                    for t, c in zip(dfFiltered.reset_index()['Ticker'], dfFiltered.reset_index()['Close']):
                        if t not in cmp_map:
                            cmp_map[t] = c
                sell_df = pd.DataFrame({
                    "Stock": sell_list,
                    "CMP ₹": [
                        round(cmp_map[s], 2) if s in cmp_map and cmp_map[s] > 0
                        else "N/A *"
                        for s in sell_list
                    ],
                    "Reason": reasons_for_exit[:len(sell_list)]
                })
                st.dataframe(sell_df, hide_index=True, use_container_width=True)
                missing_cmp = [s for s in sell_list if s not in cmp_map or cmp_map.get(s, 0) == 0]
                if missing_cmp:
                    st.caption(
                        f"* {', '.join(missing_cmp)} — CMP unavailable "
                        f"(stock selected universe mein nahi hai). "
                        "Broker app se manually CMP check karo."
                    )
            else:
                st.success("Koi sell nahi hai!")

        with col_buy:
            st.markdown('<div class="section-hdr" style="border-left-color:var(--green)">🟢 BUY List (New Entry)</div>', unsafe_allow_html=True)
            buy_list = entry_stocks.dropna().tolist()
            if buy_list:
                chips = " ".join([f'<span class="chip chip-buy">{s}</span>' for s in buy_list])
                st.markdown(chips, unsafe_allow_html=True)
                cmp_map2 = {}
                if dfStats is not None:
                    cmp_map2 = dict(zip(dfStats['Ticker'], dfStats['Close']))
                buy_df = pd.DataFrame({
                    "Stock": buy_list,
                    "CMP ₹": [round(cmp_map2[s], 2) if s in cmp_map2 else "N/A" for s in buy_list],
                })
                st.dataframe(buy_df, hide_index=True, use_container_width=True)
            else:
                st.info("Koi buy nahi hai.")

        with col_hold:
            st.markdown('<div class="section-hdr" style="border-left-color:var(--violet)">🔵 HOLD (Retain)</div>', unsafe_allow_html=True)
            if not hold_stocks.empty:
                chips = " ".join([f'<span class="chip chip-hold">{s}</span>' for s in hold_stocks.tolist()])
                st.markdown(chips, unsafe_allow_html=True)

        st.markdown('<div class="section-hdr">📋 Rebalance Table (Sell → Buy mapping)</div>', unsafe_allow_html=True)
        if not rebalance_table.empty:
            st.dataframe(rebalance_table, use_container_width=True)

        st.divider()

        # ── Rebalancer Workflow Panel ──────────────────────────────
        st.markdown('<div class="section-hdr">🔄 Rebalancer Workflow</div>', unsafe_allow_html=True)

        sell_list_local = exit_stocks.dropna().tolist()
        buy_list_local  = entry_stocks.dropna().tolist()

        cmp_map3 = {}
        if dfStats is not None:
            cmp_map3 = dict(zip(dfStats['Ticker'], dfStats['Close']))

        _df_sorted = dfFiltered.reset_index()
        if 'Rank' in _df_sorted.columns:
            _df_sorted = _df_sorted.sort_values('Rank', ascending=True)
            _top_filtered = _df_sorted[_df_sorted['Rank'] <= st.session_state.top_n_rank]
        else:
            _top_filtered = _df_sorted.head(st.session_state.top_n_rank)
        top_n_tickers = _top_filtered["Ticker"].tolist()
        worst_rank_text = "\n".join(top_n_tickers) if top_n_tickers else "(no data)"

        st.markdown("""
        <div class="workflow-box">
        <b>📋 Workflow Steps:</b><br>
        <span class="step-tag">1</span> Neeche <b>Top-N Screener list</b> copy karo → Google Sheet ke <b>"Worst Rank Held"</b> column mein paste karo
          <span style="color:var(--muted);font-size:12px;">(ye list rebalancer ko batati hai ki kaun good rank mein hai)</span><br>
        <span class="step-tag">2</span> <b>"Open Portfolio Rebalancer"</b> button dabao → Sell stocks select karo → actual sell value note karo<br>
        <span class="step-tag">3</span> <i>(Optional)</i> Neeche <b>"Buy/Sell order calculate karna chahte hain?"</b> checkbox enable karo → Sell Value enter karo → Buy orders auto-calculate honge
        </div>
        """, unsafe_allow_html=True)

        wa1, wa2 = st.columns([1, 1])
        with wa1:
            n_top = len(top_n_tickers)
            st.markdown(f"**📋 Top-{st.session_state.top_n_rank} Screener List — Google Sheet 'Worst Rank Held' column mein paste karo:**")
            st.caption(f"✅ {n_top} filtered & ranked stocks | Rank 1 se Rank {n_top} tak")
            st.text_area(
                "Top-N list — Google Sheet Worst Rank Held column mein paste karo",
                value=worst_rank_text,
                height=min(160, max(80, len(top_n_tickers) * 6 + 60)),
                key="sell_copy_area",
                label_visibility="collapsed",
                help="Yeh Top-N screener stocks Google Sheet ke Worst Rank Held column mein paste karo"
            )
            import streamlit.components.v1 as _components
            _safe_text = worst_rank_text.replace("`", "'").replace("\\", "/")
            _copy_html = f"""
            <textarea id="cpytxt" style="position:absolute;left:-9999px;">{_safe_text}</textarea>
            <button id="cpybtn"
              onclick="
                var t=document.getElementById('cpytxt');
                t.select(); t.setSelectionRange(0,99999);
                var ok=false;
                try{{ok=document.execCommand('copy');}}catch(e){{}}
                if(!ok && navigator.clipboard){{
                  navigator.clipboard.writeText(t.value).then(function(){{
                    document.getElementById('cpybtn').innerHTML='✅ Copied!';
                    document.getElementById('cpybtn').style.background='#16a34a';
                  }});
                }} else if(ok) {{
                  document.getElementById('cpybtn').innerHTML='✅ Copied!';
                  document.getElementById('cpybtn').style.background='#16a34a';
                }} else {{
                  alert('Manually select text above aur Ctrl+C / Cmd+C dabao');
                }}
              "
              style="background:var(--blue,#2563eb);color:white;border:none;padding:9px 22px;
                     border-radius:8px;font-weight:700;cursor:pointer;font-size:13px;
                     margin-top:6px;letter-spacing:.2px;">
              📋 Copy to Clipboard
            </button>
            """
            _components.html(_copy_html, height=50)

        with wa2:
            st.markdown("**⚖️ Portfolio Rebalancer:**")
            st.markdown(f"""
            <a href="{APPS_SCRIPT_URL}" target="_blank" class="qlink-btn qlink-rebalancer">
              ⚖️ Open Portfolio Rebalancer
            </a>
            <div style="font-size:11.5px;color:var(--muted);margin-top:8px;line-height:1.7;
                        padding:8px 10px;background:var(--canvas);border-radius:6px;border:1px solid var(--border);">
              📌 Wahan se sell karke <b>actual sell value</b> note karo.<br>
              ↩️ Phir neeche woh value enter karo.
            </div>
            """, unsafe_allow_html=True)

        st.divider()

        # ── Order Calculator (optional) ──────────────────────────
        show_order_calc = st.checkbox(
            "⚡ Buy/Sell order calculate karna chahte hain?",
            value=False,
            key="show_order_calc",
        )

        if show_order_calc:
            st.markdown('<div class="section-hdr">⚡ Order Calculator</div>', unsafe_allow_html=True)

            qr1, qr2, qr3, qr4 = st.columns(4)
            with qr1:
                capital_add = st.number_input(
                    "💰 Capital Addition ₹", min_value=0, value=0, step=5000, key="qr_cap"
                )
            with qr2:
                brokerage = st.number_input(
                    "🏦 Brokerage/Stock ₹", min_value=0, value=0, step=10, key="qr_brk"
                )
            with qr3:
                sell_val_input = st.number_input(
                    "💸 Sell Value ₹ (Rebalancer se enter karo)",
                    min_value=0, value=0, step=1000, key="qr_sell",
                )

            sell_brk  = len(sell_list_local) * brokerage
            buy_brk   = len(buy_list_local)  * brokerage
            net_pool  = sell_val_input + capital_add - sell_brk - buy_brk
            per_stock = net_pool / len(buy_list_local) if buy_list_local else 0

            with qr4:
                st.markdown(f"""<div class="metric-card green">
                  <div class="metric-label">Net Pool / Stock</div>
                  <div class="metric-value green">{fmt_inr(per_stock)}</div>
                </div>""", unsafe_allow_html=True)

            st.markdown(f"""<div class="reb-strip">
              <div class="reb-stat"><div class="label">Sell Value</div><div class="val b">₹{sell_val_input:,.0f}</div></div>
              <div class="reb-stat"><div class="label">+ Capital</div><div class="val g">₹{capital_add:,.0f}</div></div>
              <div class="reb-stat"><div class="label">- Sell Brok</div><div class="val r">₹{sell_brk:,.0f}</div></div>
              <div class="reb-stat"><div class="label">- Buy Brok</div><div class="val r">₹{buy_brk:,.0f}</div></div>
              <div class="reb-stat"><div class="label">Net Pool</div><div class="val g">₹{net_pool:,.0f}</div></div>
              <div class="reb-stat"><div class="label">Per Stock</div><div class="val g">{fmt_inr(per_stock)}</div></div>
            </div>""", unsafe_allow_html=True)

            if sell_val_input == 0 and not capital_add:
                st.info("💡 Sell Value enter karo (Portfolio Rebalancer se) → Buy orders auto-calculate honge.")

            if buy_list_local and per_stock > 0:
                st.markdown('<div class="section-hdr">📋 Buy Orders (Estimated)</div>', unsafe_allow_html=True)
                orders = []
                total_invested = 0
                for i, stock in enumerate(buy_list_local, 1):
                    cmp = cmp_map3.get(stock, 0)
                    if cmp > 0:
                        qty = int(per_stock / cmp)
                        val = qty * cmp
                        total_invested += val
                        orders.append({
                            "#":           i,
                            "Stock":       stock,
                            "CMP ₹":       round(cmp, 2),
                            "Gross Alloc": round(per_stock + brokerage),
                            "Brok ₹":      brokerage,
                            "Net Alloc":   round(per_stock),
                            "Qty":         qty,
                            "Value ₹":     round(val),
                        })

                if orders:
                    orders_df = pd.DataFrame(orders)
                    st.dataframe(
                        orders_df.style.format({
                            "CMP ₹": "{:.2f}",
                            "Gross Alloc": "{:,.0f}", "Net Alloc": "{:,.0f}", "Value ₹": "{:,.0f}"
                        }),
                        use_container_width=True, hide_index=True, height=300
                    )
                    leftover = net_pool - total_invested
                    st.markdown(f"""<div class="reb-strip">
                      <div class="reb-stat"><div class="label">Total Invested</div><div class="val g">₹{total_invested:,.0f}</div></div>
                      <div class="reb-stat"><div class="label">Leftover</div><div class="val p">₹{leftover:,.0f}</div></div>
                      <div class="reb-stat"><div class="label">Buy Orders</div><div class="val b">{len(orders)}</div></div>
                    </div>""", unsafe_allow_html=True)

    elif not portfolio:
        st.markdown("""
        <div class="card" style="text-align:center;padding:32px 20px;">
          <div style="font-size:32px;margin-bottom:10px;">⚖️</div>
          <div style="font-size:14px;font-weight:600;color:var(--ink-mid);margin-bottom:6px;">
            Portfolio load nahi hua
          </div>
          <div style="font-size:12.5px;color:var(--muted);">
            Upar se portfolio data load karo (Google Sheet ya CSV)
          </div>
        </div>
        """, unsafe_allow_html=True)

    st.divider()
    if st.button("▶ Next: Apply & Export →", type="primary"):
        st.session_state.current_step = 4; st.rerun()

    render_footer()


# ═══════════════════════════════════════════════════════════════
# STEP 4 — APPLY & EXPORT
# ═══════════════════════════════════════════════════════════════
elif st.session_state.current_step == 4:
    st.markdown('<div class="section-hdr">💾 Apply & Export</div>', unsafe_allow_html=True)

    sell        = st.session_state.sell_list or []
    buy         = st.session_state.buy_list  or []
    portfolio   = st.session_state.reb_portfolio or []
    dfStats     = st.session_state.dfStats
    dfFiltered  = st.session_state.dfFiltered
    reb_table   = st.session_state.rebalance_table
    failed_blank= st.session_state.failed_blank or []
    U           = st.session_state.universe
    top_n       = st.session_state.top_n_rank
    rank_method = st.session_state.ranking_method
    api_source  = st.session_state.data_source
    end_date    = st.session_state.lookback_date

    # ── Summary strip ─────────────────────────────────────────
    st.markdown(f"""<div class="reb-strip">
      <div class="reb-stat"><div class="label">Exits (SELL)</div><div class="val r">{len(sell)}</div></div>
      <div class="reb-stat"><div class="label">New Entries (BUY)</div><div class="val g">{len(buy)}</div></div>
      <div class="reb-stat"><div class="label">Retained (HOLD)</div><div class="val p">{len(portfolio) - len(sell)}</div></div>
      <div class="reb-stat"><div class="label">New Portfolio Size</div><div class="val b">{len(portfolio) - len(sell) + len(buy)}</div></div>
    </div>""", unsafe_allow_html=True)

    if dfFiltered is not None and dfStats is not None:
        st.markdown('<div class="section-hdr">💾 Excel Export (v10 Format — 4 Sheets)</div>', unsafe_allow_html=True)

        if failed_blank:
            df_failed = pd.DataFrame({
                'S.No.':        range(1, len(failed_blank)+1),
                'Failed Stock': failed_blank
            }).set_index('S.No.')
        else:
            df_failed = pd.DataFrame(columns=['Failed Stock'])
            df_failed.index.name = 'S.No.'

        if reb_table is None or reb_table.empty:
            reb_table = pd.DataFrame(columns=['Sell Stocks','Buy Stocks','Reason for Exit'])
            reb_table.index.name = 'S.No.'

        filtered = dfFiltered.copy()
        excel_file = f"{end_date.strftime('%Y-%m-%d')}_{U}_{rank_method}_{api_source}_lookback.xlsx"

        with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
            dfStats.to_excel(   writer, sheet_name="Unfiltered Stocks",     index=True)
            filtered.to_excel(  writer, sheet_name="Filtered Stocks",       index=True)
            df_failed.to_excel( writer, sheet_name="Failed Downloads",      index=True)
            reb_table.to_excel( writer, sheet_name="Portfolio Rebalancing", index=True)

        try:
            format_excel_unfiltered(excel_file, U, top_n)
            format_excel_filtered(excel_file, U, top_n)
            format_simple_sheet(excel_file, "Failed Downloads")
            format_simple_sheet(excel_file, "Portfolio Rebalancing")
        except Exception as e:
            st.warning(f"Excel formatting partial error (file still usable): {e}")

        with open(excel_file, "rb") as f:
            st.download_button(
                label     = "📥 Download Excel (4 Sheets)",
                data      = f.read(),
                file_name = excel_file,
                mime      = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type      = "primary",
            )

        st.success(f"✅ Excel ready: `{excel_file}`")
        st.markdown(f"""
        <div style="background:var(--green-bg);border:1px solid var(--green-bdr);
                    border-radius:var(--r-md);padding:10px 16px;font-size:12px;
                    color:#15803d;margin:4px 0 12px 0;">
        📄 <b>4 Sheets:</b> &nbsp;
        <span class="badge badge-green">Unfiltered Stocks</span>
        &nbsp;<span class="badge badge-green">Filtered Stocks</span>
        &nbsp;<span class="badge badge-green">Failed Downloads</span>
        &nbsp;<span class="badge badge-green">Portfolio Rebalancing</span>
        </div>
        """, unsafe_allow_html=True)

    # ── Quick Links ───────────────────────────────────────────
    st.markdown('<div class="section-hdr">📊 Apps Script Workflow — Quick Links</div>', unsafe_allow_html=True)
    col_links = st.columns(2)
    with col_links[0]:
        st.markdown(f"""
        <a href="{APPS_SCRIPT_URL}" target="_blank" class="qlink-btn qlink-rebalancer">
        ⚖️ Portfolio Rebalancer
        </a>
        """, unsafe_allow_html=True)
    with col_links[1]:
        st.markdown("""
        <a href="https://prayan2702.github.io/momn-dashboard/" target="_blank"
           class="qlink-btn qlink-dashboard">
        📈 Portfolio Dashboard
        </a>
        """, unsafe_allow_html=True)

    if reb_table is not None and not reb_table.empty:
        st.markdown('<div class="section-hdr">📋 Portfolio Rebalancing</div>', unsafe_allow_html=True)
        st.dataframe(reb_table, use_container_width=True)

    st.divider()
    col_r1, col_r2 = st.columns(2)
    with col_r1:
        if st.button("🔄 New Month — Restart from Step 1", use_container_width=True):
            for k in list(st.session_state.keys()):
                del st.session_state[k]
            st.rerun()
    with col_r2:
        if st.button("← Step 3 — Edit Rebalance"):
            st.session_state.current_step = 3; st.rerun()

    render_footer()
