"""
momn_streamlit_app_v14.py  (SOP v2026.08 — Phase 1 update)
=========================
Momentum Screener + Portfolio Rebalancer — v14

Changes vs v13:
  • _verify_pin() bug fix — str() wrap added (TOML integer type mismatch fix)
  • _pin_secret_exists() helper — debug ke liye
  • GitHub Actions + ATH push panel: smarter error messages
    - Secret missing → alag message (TRIGGER_PIN set nahi)
    - Wrong PIN → alag message
  • Logic/calculations untouched

SOP v2026.06 — Multi-Asset Regime Changes (app-level):
  • Allocation updated: Score3=80/15/5, Score2=65/20/15, Score1=45/25/30, Score0=25/30/45
  • VIX Overlay panel: VIX>30→+5%Gold, VIX 20-30→+3%Gold (Liquid→Gold, Equity untouched)
  • Gold drift band: ±7% of PF (was ±5%)
  • Transaction guardrail: ₹15K minimum per GOLDBEES/Liquid transaction
  • Drawdown Protocol: DD≥15% warn, DD≥20% override, DD≥30% emergency
  • Equiweight Maintenance: Proceeds allocation panel — exit-funded regime shift
    Per-stock target = Equity Budget÷30, drift band ±₹20K
  • _alloc_start: updated to new allocations
  • Regime Tab: VIX overlay display on allocation tiles

  NOTE: calculations.py get_regime_score() mein bhi allocation constants update karo:
    ALLOC = {3:(0.80,0.15,0.05), 2:(0.65,0.20,0.15), 1:(0.45,0.25,0.30), 0:(0.25,0.30,0.45)}
    GOLD_CAP = 0.30  # hard max
    GOLD_DRIFT_BAND = 0.07  # ±7% of total portfolio
"""

import io
import os
import json
import base64
import time
import datetime
import warnings

import numpy as np
import pandas as pd
import streamlit as st
import requests

# ── GitHub API helpers (PIN-protected) ───────────────────────
_GH_OWNER = "prayan2702"
_GH_REPO  = "Streamlit_Momn_v13_Cached_DB"

# Workflow file names in .github/workflows/
_WF_YFINANCE = "daily_cache.yml"
_WF_UPSTOX   = "daily_cache_upstox.yml"
_WF_ANGEL    = "daily_cache_angelone.yml"
_WF_FII      = "fetch_fii_data.yml"


def _get_secret(key: str, default: str = "") -> str:
    """st.secrets se safely read karo."""
    try:
        return st.secrets.get(key, default)
    except Exception:
        return default


def _verify_pin(entered: str) -> bool:
    """
    Entered PIN ko TRIGGER_PIN secret se match karo.
    Secret set nahi hai → always False (safe default).
    Note: str() wrap kiya — TOML mein quotes na hone par integer aa sakta hai.
    """
    correct = _get_secret("TRIGGER_PIN", "")
    if not correct:
        return False
    return str(entered).strip() == str(correct).strip()


def _pin_secret_exists() -> bool:
    """TRIGGER_PIN secret set hai ya nahi — debug ke liye."""
    return bool(_get_secret("TRIGGER_PIN", ""))


def _gh_headers() -> dict:
    """GitHub API headers with PAT token."""
    token = _get_secret("GITHUB_PAT", "")
    return {
        "Authorization": f"Bearer {token}",
        "Accept":        "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
    }


def _trigger_workflow(workflow_file: str) -> tuple[bool, str]:
    """
    GitHub Actions workflow_dispatch trigger karo.
    Returns: (success: bool, message: str)
    """
    url = (
        f"https://api.github.com/repos/{_GH_OWNER}/{_GH_REPO}"
        f"/actions/workflows/{workflow_file}/dispatches"
    )
    try:
        r = requests.post(
            url,
            headers=_gh_headers(),
            json={"ref": "main"},
            timeout=15,
        )
        if r.status_code == 204:
            return True, f"✅ Workflow `{workflow_file}` triggered!"
        elif r.status_code == 401:
            return False, "❌ GitHub PAT invalid ya expired. Secrets check karo."
        elif r.status_code == 404:
            return False, f"❌ Workflow `{workflow_file}` nahi mila. Repo/name check karo."
        elif r.status_code == 422:
            return False, "❌ Branch `main` nahi mili. Repo settings check karo."
        else:
            return False, f"❌ HTTP {r.status_code}: {r.text[:200]}"
    except Exception as e:
        return False, f"❌ Network error: {e}"


def _push_json_to_github(path: str, content_dict: dict, commit_msg: str) -> tuple[bool, str]:
    """
    JSON dict ko GitHub repo mein push karo (create or update).
    path = repo root se relative, e.g. "ath_memory.json"
    Returns: (success: bool, message: str)
    """
    url = f"https://api.github.com/repos/{_GH_OWNER}/{_GH_REPO}/contents/{path}"
    try:
        # Pehle current SHA fetch karo (update ke liye zaroori)
        r_get = requests.get(url, headers=_gh_headers(), timeout=10)
        sha = r_get.json().get("sha") if r_get.status_code == 200 else None

        content_b64 = base64.b64encode(
            json.dumps(content_dict, indent=2, ensure_ascii=False).encode("utf-8")
        ).decode("ascii")

        payload = {
            "message": commit_msg,
            "content": content_b64,
            "branch":  "main",
        }
        if sha:
            payload["sha"] = sha

        r_put = requests.put(url, headers=_gh_headers(), json=payload, timeout=20)
        if r_put.status_code in (200, 201):
            action = "updated" if sha else "created"
            return True, f"✅ `{path}` GitHub pe {action}!"
        elif r_put.status_code == 401:
            return False, "❌ GitHub PAT invalid ya expired. Secrets check karo."
        else:
            return False, f"❌ HTTP {r_put.status_code}: {r_put.text[:200]}"
    except Exception as e:
        return False, f"❌ Network error: {e}"

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side

warnings.filterwarnings("ignore")

# ── Local modules ──────────────────────────────────────────────
try:
    from calculations import build_dfStats, apply_filters
    _CALCS_AVAILABLE = True
except ImportError:
    _CALCS_AVAILABLE = False

import yfinance as yf  # always available (in requirements.txt)

# ── data_service (may fail if SmartApi/pyotp not installed) ───
_DS_AVAILABLE   = False
_DS_IMPORT_ERR  = ""
try:
    from data_service import fetch_data
    _DS_AVAILABLE = True
except Exception as _e:
    _DS_IMPORT_ERR = str(_e)

# ── upstox_auth ───────────────────────────────────────────────
_UPSTOX_AVAILABLE = False
try:
    from upstox_auth import get_upstox_access_token
    _UPSTOX_AVAILABLE = True
except Exception:
    pass

# ── angelone_auth ─────────────────────────────────────────────
_ANGEL_AVAILABLE = False
try:
    from angelone_auth import get_angelone_client
    _ANGEL_AVAILABLE = True
except Exception:
    pass

# ── cache_loader (pre-built Parquet cache — YFinance) ────────
_CACHE_AVAILABLE = False
try:
    from cache_loader import load_cache, get_cache_meta, get_cache_age_days, get_cache_status_html
    _CACHE_AVAILABLE = True
except ImportError:
    pass

# ── cache_loader_upstox (pre-built Parquet cache — Upstox) ───
_CACHE_UPSTOX_AVAILABLE = False
try:
    from cache_loader_upstox import (
        load_cache          as load_cache_upstox,
        get_cache_meta      as get_cache_meta_upstox,
        get_cache_age_days  as get_cache_age_days_upstox,
        get_cache_status_html as get_cache_status_html_upstox,
    )
    _CACHE_UPSTOX_AVAILABLE = True
except ImportError:
    pass

# ── cache_loader_angelone (pre-built Parquet cache — Angel One) ──
_CACHE_ANGEL_AVAILABLE = False
try:
    from cache_loader_angelone import (
        load_cache            as load_cache_angel,
        get_cache_meta        as get_cache_meta_angel,
        get_cache_age_days    as get_cache_age_days_angel,
        get_cache_status_html as get_cache_status_html_angel,
    )
    _CACHE_ANGEL_AVAILABLE = True
except ImportError:
    pass

# ── cache_loader_fyers (pre-built Parquet cache — Fyers) ──────
_CACHE_FYERS_AVAILABLE = False
try:
    from cache_loader_fyers import (
        load_cache            as load_cache_fyers,
        get_cache_meta        as get_cache_meta_fyers,
        get_cache_age_days    as get_cache_age_days_fyers,
        get_cache_status_html as get_cache_status_html_fyers,
    )
    _CACHE_FYERS_AVAILABLE = True
except ImportError:
    pass

# ── fyers_auth (live Fyers fetch) ─────────────────────────────
_FYERS_AVAILABLE = False
try:
    from fyers_auth import get_fyers_client
    _FYERS_AVAILABLE = True
except Exception:
    pass

# ── Inline YFinance fetcher (fallback when data_service fails) ─
def _fetch_yfinance_inline(symbols_ns, start_date, end_date,
                            progress_bar, status_text, chunk_size=15):
    """Pure yfinance fetch — no data_service dependency.

    yfinance end is EXCLUSIVE — end=today gives only yesterday.
    Fix: always use end = end_date + 1 day to include today's close.
    """
    import datetime as _dtmod
    # yfinance end is EXCLUSIVE: add 1 day so today's data is included
    if hasattr(end_date, 'date'):
        _yf_end = end_date + _dtmod.timedelta(days=1)
    else:
        _yf_end = end_date + _dtmod.timedelta(days=1) if hasattr(end_date, 'day') else end_date
    close_chunks, high_chunks, vol_chunks = [], [], []
    failed = []
    total  = len(symbols_ns)
    for k in range(0, total, chunk_size):
        chunk = symbols_ns[k:k + chunk_size]
        pct   = min((k + chunk_size) / total, 1.0)
        status_text.markdown(f"⏳ **Fetching {k+1}–{min(k+chunk_size, total)} / {total}**")
        progress_bar.progress(pct * 0.88)
        try:
            raw = yf.download(chunk, start=start_date, end=_yf_end,
                              progress=False, auto_adjust=True, threads=True,
                              multi_level_index=False)
            if not raw.empty:
                close_chunks.append(raw["Close"])
                high_chunks.append(raw["High"])
                vol_val = raw["Close"].multiply(raw.get("Volume", 1))
                vol_chunks.append(vol_val)
        except Exception as e:
            failed.extend(chunk)
        time.sleep(0.5)

    if not close_chunks:
        return None, None, None, failed

    close  = pd.concat(close_chunks,  axis=1)
    high   = pd.concat(high_chunks,   axis=1)
    volume = pd.concat(vol_chunks,    axis=1)
    close  = close.loc[:,  ~close.columns.duplicated()]
    high   = high.loc[:,   ~high.columns.duplicated()]
    volume = volume.loc[:, ~volume.columns.duplicated()]
    return close, high, volume, failed

# ═══════════════════════════════════════════════════════════════
# PAGE CONFIG
# ═══════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="Momn Screener v13",
    page_icon="📈",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── CSS ────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Outfit:wght@400;500;600;700;800&family=DM+Mono:wght@400;500&display=swap');

/* ── Design Tokens ── */
:root {
    --navy:       #0d1b2a;
    --navy-mid:   #1a2e45;
    --navy-light: #1e3a5f;
    --slate:      #334155;
    --muted:      #64748b;
    --border:     #e2e8f0;
    --bg:         #f8fafc;
    --bg-white:   #ffffff;
    --teal:       #0ea5e9;
    --teal-dark:  #0284c7;
    --green:      #16a34a;
    --green-bg:   #dcfce7;
    --green-bdr:  #86efac;
    --red:        #dc2626;
    --red-bg:     #fee2e2;
    --red-bdr:    #fca5a5;
    --blue:       #2563eb;
    --blue-bg:    #dbeafe;
    --blue-bdr:   #93c5fd;
    --violet:     #7c3aed;
    --violet-bg:  #ede9fe;
    --amber:      #d97706;
    --amber-bg:   #fef3c7;
    --text-main:  #0f172a;
    --text-sub:   #475569;
    --radius-sm:  6px;
    --radius-md:  10px;
    --radius-lg:  14px;
    --shadow-sm:  0 1px 3px rgba(0,0,0,.08), 0 1px 2px rgba(0,0,0,.06);
    --shadow-md:  0 4px 12px rgba(0,0,0,.10), 0 2px 4px rgba(0,0,0,.06);
}

/* ════════════════════════════
   APP HEADER
   ════════════════════════════ */
.app-header {
    background: linear-gradient(135deg, var(--navy) 0%, var(--navy-light) 55%, #0c2340 100%);
    border-bottom: 1px solid rgba(56,189,248,.25);
    padding: 14px 24px;
    margin: -1rem -1rem 1.5rem -1rem;
    display: flex;
    align-items: center;
    justify-content: space-between;
    position: relative;
    overflow: hidden;
}
.app-header::before {
    content: '';
    position: absolute;
    inset: 0;
    background: radial-gradient(ellipse 60% 80% at 70% 50%, rgba(14,165,233,.12) 0%, transparent 70%);
    pointer-events: none;
}
.app-title {
    color: #f1f5f9;
    font-size: 21px;
    font-weight: 800;
    letter-spacing: -.3px;
}
.app-title span { color: var(--teal); }
.app-subtitle {
    color: #94a3b8;
    font-size: 11.5px;
    margin-top: 3px;
    letter-spacing: .2px;
}
.app-badge {
    background: rgba(14,165,233,.15);
    border: 1px solid rgba(14,165,233,.3);
    color: var(--teal);
    font-size: 10px;
    font-weight: 700;
    padding: 2px 8px;
    border-radius: 20px;
    letter-spacing: .5px;
    text-transform: uppercase;
    margin-left: 10px;
}
.app-header-right {
    text-align: right;
    color: #94a3b8;
    font-size: 11.5px;
    line-height: 1.6;
}
.app-header-right .user-tag {
    display: inline-block;
    background: rgba(255,255,255,.07);
    border: 1px solid rgba(255,255,255,.12);
    padding: 2px 10px;
    border-radius: 20px;
    font-size: 11px;
    color: #cbd5e1;
}

/* ════════════════════════════
   STEP PROGRESS BAR
   ════════════════════════════ */
.step-bar {
    display: flex;
    align-items: center;
    gap: 0;
    background: var(--bg-white);
    border: 1px solid var(--border);
    border-radius: var(--radius-lg);
    padding: 12px 20px;
    margin-bottom: 1.4rem;
    overflow-x: auto;
    box-shadow: var(--shadow-sm);
}
.step-item {
    display: flex;
    align-items: center;
    gap: 9px;
    padding: 8px 16px;
    border-radius: var(--radius-md);
    font-size: 13px;
    font-weight: 600;
    white-space: nowrap;
    transition: all .2s;
}
.step-item.done {
    background: var(--green-bg);
    color: #15803d;
    border: 1px solid var(--green-bdr);
}
.step-item.active {
    background: linear-gradient(135deg, #dbeafe 0%, #ede9fe 100%);
    color: var(--blue);
    border: 1.5px solid var(--blue-bdr);
    box-shadow: 0 0 0 3px rgba(37,99,235,.12);
}
.step-item.pending {
    color: #94a3b8;
    border: 1px solid transparent;
}
.step-circle {
    width: 26px;
    height: 26px;
    border-radius: 50%;
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 11px;
    font-weight: 800;
    flex-shrink: 0;
}
.done   .step-circle { background: var(--green); color: #fff; }
.active .step-circle { background: var(--blue);  color: #fff; box-shadow: 0 0 0 3px rgba(37,99,235,.2); }
.pending .step-circle { background: #e2e8f0; color: #94a3b8; }
.step-connector {
    width: 36px;
    height: 2px;
    background: var(--border);
    flex-shrink: 0;
    border-radius: 2px;
    margin: 0 2px;
}
.step-connector.done-line {
    background: linear-gradient(90deg, var(--green) 0%, #22c55e 100%);
}

/* ════════════════════════════
   METRIC CARDS
   ════════════════════════════ */
.metric-row  { display: flex; gap: 12px; flex-wrap: wrap; margin: 14px 0; }
.metric-card {
    background: var(--bg-white);
    border: 1px solid var(--border);
    border-left: 4px solid var(--slate);
    border-radius: var(--radius-md);
    padding: 12px 18px;
    min-width: 140px;
    box-shadow: var(--shadow-sm);
    transition: box-shadow .2s, transform .15s;
}
.metric-card:hover {
    box-shadow: var(--shadow-md);
    transform: translateY(-1px);
}
.metric-card.green  { border-left-color: var(--green); }
.metric-card.red    { border-left-color: var(--red); }
.metric-card.blue   { border-left-color: var(--blue); }
.metric-card.violet { border-left-color: var(--violet); }
.metric-card.amber  { border-left-color: var(--amber); }
.metric-label {
    font-size: 10px;
    color: var(--muted);
    text-transform: uppercase;
    letter-spacing: .6px;
    font-weight: 600;
}
.metric-value {
    font-size: 22px;
    font-weight: 800;
    color: var(--text-color, var(--text-main));
    margin-top: 4px;
    letter-spacing: -.3px;
}
.metric-value.green  { color: var(--green); }
.metric-value.red    { color: var(--red); }
.metric-value.blue   { color: var(--blue); }
.metric-value.violet { color: var(--violet); }
.metric-value.amber  { color: var(--amber); }

/* ════════════════════════════
   SECTION HEADERS
   ════════════════════════════ */
.section-hdr {
    font-size: 14px;
    font-weight: 700;
    color: var(--text-color);
    border-left: 4px solid var(--teal);
    padding: 6px 0 6px 12px;
    margin: 1.4rem 0 .9rem;
    background: linear-gradient(90deg, rgba(14,165,233,.06) 0%, transparent 60%);
    border-radius: 0 var(--radius-sm) var(--radius-sm) 0;
}

/* ════════════════════════════
   NSE LINK BOX
   ════════════════════════════ */
.nse-link-box {
    background: linear-gradient(135deg, #eff6ff 0%, #f0fdf4 100%);
    border: 1px solid var(--blue-bdr);
    border-radius: var(--radius-md);
    padding: 14px 18px;
    display: flex;
    align-items: center;
    gap: 12px;
    margin: 12px 0;
    color: var(--text-color, #1e293b);
}
.nse-link-box b { color: inherit; }
.nse-link-box a { color: var(--blue); font-weight: 700; font-size: 13px; text-decoration: none; }
.nse-link-box a:hover { text-decoration: underline; }
.nse-link-box .hint { font-size: 11px; color: var(--text-color, #64748b); opacity: 0.75; margin-top: 3px; }

/* ════════════════════════════
   CHIPS (SELL / BUY / HOLD)
   ════════════════════════════ */
.chip {
    display: inline-block;
    padding: 4px 10px;
    border-radius: 20px;
    font-size: 11px;
    font-weight: 700;
    margin: 3px 2px;
    letter-spacing: .2px;
    border: 1px solid transparent;
}
.chip-sell {
    background: var(--red-bg);
    color: var(--red);
    border-color: var(--red-bdr);
}
.chip-buy {
    background: var(--green-bg);
    color: var(--green);
    border-color: var(--green-bdr);
}
.chip-hold {
    background: #f1f5f9;
    color: var(--text-sub);
    border-color: #cbd5e1;
}

/* ════════════════════════════
   REBALANCE STRIP
   ════════════════════════════ */
.reb-strip {
    display: flex;
    gap: 0;
    flex-wrap: wrap;
    background: var(--bg-white);
    border: 1px solid var(--border);
    border-radius: var(--radius-md);
    overflow: hidden;
    margin: 12px 0;
    box-shadow: var(--shadow-sm);
}
.reb-stat {
    flex: 1;
    min-width: 100px;
    padding: 12px 16px;
    border-right: 1px solid var(--border);
    text-align: center;
}
.reb-stat:last-child { border-right: none; }
.reb-stat .label {
    font-size: 9.5px;
    color: var(--muted);
    text-transform: uppercase;
    letter-spacing: .6px;
    font-weight: 600;
    margin-bottom: 4px;
}
.reb-stat .val {
    font-size: 18px;
    font-weight: 800;
    letter-spacing: -.2px;
}
.reb-stat .val.r { color: var(--red); }
.reb-stat .val.g { color: var(--green); }
.reb-stat .val.b { color: var(--blue); }
.reb-stat .val.p { color: var(--violet); }

/* ════════════════════════════
   WORKFLOW INFO BOX
   ════════════════════════════ */
.workflow-box {
    background: linear-gradient(135deg, #f0f9ff 0%, #f0fdf4 100%);
    border: 1px solid var(--blue-bdr);
    border-left: 4px solid var(--blue);
    border-radius: var(--radius-md);
    padding: 14px 18px;
    margin: 10px 0;
    font-size: 13px;
    line-height: 2;
    color: var(--text-main);
}
.workflow-box .step-tag {
    display: inline-block;
    background: var(--blue);
    color: #fff;
    border-radius: 12px;
    padding: 1px 8px;
    font-size: 10.5px;
    font-weight: 700;
    margin-right: 6px;
}

/* ════════════════════════════
   LOGIN PAGE
   ════════════════════════════ */
.login-wrap {
    display: flex;
    justify-content: center;
    padding: 40px 16px;
}
.login-card {
    background: var(--bg-white);
    border: 1px solid var(--border);
    border-radius: 18px;
    box-shadow: 0 8px 32px rgba(0,0,0,.10);
    padding: 40px 44px;
    width: 100%;
    text-align: center;
    margin-bottom: 0;
}
.login-logo {
    text-align: center;
    margin-bottom: 6px;
    font-size: 38px;
}
.login-title {
    text-align: center;
    font-size: 22px;
    font-weight: 800;
    color: var(--text-main);
    margin-bottom: 4px;
}
.login-sub {
    text-align: center;
    font-size: 12.5px;
    color: var(--muted);
    margin-bottom: 28px;
}

/* ════════════════════════════
   QUICK LINK BUTTONS (Step 4)
   ════════════════════════════ */
.qlink-btn {
    display: block;
    text-align: center;
    font-weight: 700;
    font-size: 13px;
    padding: 12px 20px;
    border-radius: var(--radius-md);
    text-decoration: none !important;
    margin: 6px 0;
    transition: opacity .2s, transform .15s;
    border: none;
    color: #ffffff !important;
}
.qlink-btn:hover { opacity: .9; transform: translateY(-1px); color: #ffffff !important; }
.qlink-rebalancer {
    background: linear-gradient(135deg, #1a237e 0%, #283593 100%);
    border: 1px solid rgba(255,255,255,.1);
    box-shadow: 0 4px 12px rgba(26,35,126,.25);
}
.qlink-dashboard {
    background: linear-gradient(135deg, #5b21b6 0%, #7c3aed 100%);
    border: 1px solid rgba(255,255,255,.1);
    box-shadow: 0 4px 12px rgba(124,58,237,.25);
}

/* ════════════════════════════
   SIDEBAR STEP NAV TWEAKS
   ════════════════════════════ */
section[data-testid="stSidebar"] .block-container {
    padding-top: 1rem;
}

/* ════════════════════════════
   DARK MODE OVERRIDES
   Streamlit dark mode sets --text-color to a light value automatically.
   We use that variable + explicit dark selector for full coverage.
   ════════════════════════════ */
@media (prefers-color-scheme: dark) {
    :root {
        --text-main: #f1f5f9;
        --text-sub:  #cbd5e1;
        --muted:     #94a3b8;
        --border:    #1e293b;
        --bg:        #0f172a;
        --bg-white:  #1e293b;
    }
}

/* Streamlit injects data-theme="dark" on <html> when dark mode is active */
[data-theme="dark"] :root,
[data-theme="dark"] {
    --text-main: #f1f5f9;
    --text-sub:  #cbd5e1;
    --muted:     #94a3b8;
    --border:    #1e293b;
    --bg:        #0f172a;
    --bg-white:  #1e293b;
}

/* Explicit dark-mode rules targeting Streamlit's stApp class */
[data-theme="dark"] .section-hdr {
    color: #f1f5f9 !important;
    background: linear-gradient(90deg, rgba(14,165,233,.10) 0%, transparent 60%);
}

@media (prefers-color-scheme: dark) {
    .section-hdr {
        color: #f1f5f9 !important;
        background: linear-gradient(90deg, rgba(14,165,233,.10) 0%, transparent 60%);
    }
}

[data-theme="dark"] .step-item.pending,
[data-theme="dark"] .metric-label,
[data-theme="dark"] .reb-stat .label {
    color: #94a3b8;
}

[data-theme="dark"] .metric-value:not([class*=" "]):not(.green):not(.red):not(.blue):not(.violet):not(.amber),
[data-theme="dark"] .metric-value {
    color: var(--text-color, #f1f5f9);
}

[data-theme="dark"] .step-bar,
[data-theme="dark"] .metric-card,
[data-theme="dark"] .reb-strip {
    background: #1e293b;
    border-color: #334155;
}

[data-theme="dark"] .workflow-box {
    background: linear-gradient(135deg, #0c1e35 0%, #0a1f14 100%);
    border-color: #1e40af;
    color: #e2e8f0;
}

[data-theme="dark"] .nse-link-box {
    background: linear-gradient(135deg, #0c1e35 0%, #0a1f14 100%);
    border-color: #1e40af;
}

[data-theme="dark"] .chip-hold {
    background: #1e293b;
    color: #94a3b8;
    border-color: #334155;
}

[data-theme="dark"] .login-card {
    background: #1e293b;
    border-color: #334155;
}

[data-theme="dark"] .login-title,
[data-theme="dark"] .login-sub {
    color: #e2e8f0;
}
[data-theme="dark"] .nse-link-box {
    background: linear-gradient(135deg, #0c1e35 0%, #0a1f14 100%);
    border-color: #1e40af;
    color: #e2e8f0;
}
[data-theme="dark"] .nse-link-box b    { color: #f1f5f9; }
[data-theme="dark"] .nse-link-box .hint{ color: #94a3b8; opacity: 1; }
[data-theme="dark"] .nse-link-box a    { color: #60a5fa; }
@media (prefers-color-scheme: dark) {
    .nse-link-box {
        background: linear-gradient(135deg, #0c1e35 0%, #0a1f14 100%) !important;
        border-color: #1e40af !important; color: #e2e8f0 !important;
    }
    .nse-link-box b    { color: #f1f5f9 !important; }
    .nse-link-box .hint{ color: #94a3b8 !important; opacity: 1 !important; }
    .nse-link-box a    { color: #60a5fa !important; }
}
</style>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════
# CONSTANTS
# ═══════════════════════════════════════════════════════════════
UNIVERSES    = ['Nifty50','Nifty100','Nifty200','Nifty250','Nifty500','N750','AllNSE']
API_OPTIONS  = [
    "📦 Pre-cached YFinance",
    "📦 Pre-cached Upstox",
    "📦 Pre-cached Angel One",
    "📦 Pre-cached Fyers",
    "YFinance",
    "Upstox",
    "Angel One",
    "Fyers",
]
RANKING_MAP  = {
    "AvgZScore 12M/6M/3M":    "avgZScore12_6_3",
    "AvgZScore 12M/9M/6M/3M": "avgZScore12_9_6_3",
    "AvgSharpe 12M/6M/3M":    "avgSharpe12_6_3",
    "AvgSharpe 9M/6M/3M":     "avgSharpe9_6_3",
    "AvgSharpe 12M/9M/6M/3M": "avg_All",
    "Sharpe12M":               "sharpe12M",
    "Sharpe3M":                "sharpe3M",
}
PORTFOLIO_CSV_URL = (
    "https://docs.google.com/spreadsheets/d/e/"
    "2PACX-1vS4HDgiell4n1kd08OnlzOQobfPzeDtVyWJ8gETFlYbz27qhOmfqKZOoIXZItRQEq5ANATYIcZJm0gk"
    "/pub?output=csv"
)
APPS_SCRIPT_URL = (
    "https://script.google.com/macros/s/"
    "AKfycbwUNaPd82fIyQXBrPguLBZBv4tLA94Y_Uw4g-8_W77qRvmpQgJvK6_huvWcjVy0XRkc/exec"
)
GITHUB_BASE = "https://raw.githubusercontent.com/prayan2702/Streamlit_Momn_v13_Cached_DB/refs/heads/main"
DASHBOARD_API_URL = (
    "https://script.google.com/macros/s/"
    "AKfycbwecEMb7rhlvdkcYTq_ext1MfMtRwIv5givDI2h-Ke39icmHzqfzLCLnxpzYBx5bm5c9A/exec"
)

# ── Credentials: st.secrets se lo (NEVER hardcode in source code) ──────────
# Streamlit Cloud → App Settings → Secrets mein add karo:
#   [auth]
#   username = "your_username"
#   password = "your_password"
_auth = st.secrets.get("auth", {})
USERNAME = _auth.get("username", "")
PASSWORD = _auth.get("password", "")

# ═══════════════════════════════════════════════════════════════
# SESSION STATE INIT
# ═══════════════════════════════════════════════════════════════
_defaults = {
    "logged_in":      False,
    "current_step":   1,
    "universe":       "AllNSE",
    "symbols":        None,       # list[str] — .NS suffixed
    "eq_df":          None,       # only for AllNSE CSV upload
    "dfStats":        None,
    "dfFiltered":     None,
    "failed_blank":   [],
    "reb_portfolio":  None,
    "sell_list":      None,
    "buy_list":       None,
    "rebalance_table": None,
    "lookback_date":  datetime.date.today(),
    "ranking_method": "avgZScore12_6_3",
    "data_source":    "YFinance",
    "top_n_rank":     100,
    "screener_done":  False,
    "rebalance_done": False,
    # ── Pending merge flow (pre-cached + missing stocks fix) ──
    "_pending_merge":   False,   # True = waiting for user decision
    "_pending_close":   None,
    "_pending_high":    None,
    "_pending_volume":  None,
    "_pending_fp":      None,
    "_pending_dates":   None,
    # ── Cross-source review flow (Angel vs Upstox ATH/Close diff) ──
    "_cross_review_done":      False,   # True = review completed/skipped
    "_cross_review_overrides": {},      # {ticker: "primary"|"secondary"}
    "_cross_sec_close":        None,    # secondary cache close df (top 400 slice)
    "_cross_sec_high":         None,    # secondary cache high df (top 400 slice)
    "_cross_diff_df":          None,    # DataFrame of stocks with significant diffs
    "_cross_primary_label":    "",      # "Angel One" or "Upstox"
    "_cross_secondary_label":  "",      # the other one
    "_cross_filter_params":    None,    # filter params to re-apply after override
    "_cross_top_n":            400,     # how many top-ranked stocks to compare
    "_cross_error":            None,    # error string if comparison failed
    "_cross_error_detail":     None,    # full traceback
    # ── ATH Override Memory (persisted in ath_memory.json) ──────
    "_ath_memory":             {},     # loaded from file on first cross-review run
    "_ath_memory_loaded":      False,  # sentinel to load from file exactly once
    # ── Phase 1: Regime state (7-signal weighted, QFSM) ─────────
    "_regime_state":           None,   # RegimeState object
    "_regime_prev_state":      None,   # previous RegimeState (for confirmation tracking)
    "_rt_nifty_close":         None,   # fetched Nifty close
    "_rt_nifty_dma200":        None,   # fetched Nifty 200DMA
    "_rt_rank_history":        [],     # list of weekly top-50 dicts
    "_rb_memory_loaded":       False,  # rebalance_memory.json loaded flag
    "_rb_memory":              {},     # last rebalance memory dict
}
for k, v in _defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ═══════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════
def fmt_inr(v):
    if pd.isna(v): return "—"
    v = int(round(v))
    if abs(v) >= 10_000_000: return f"₹{v/10_000_000:.1f}Cr"
    if abs(v) >= 100_000:    return f"₹{v/100_000:.1f}L"
    return f"₹{v:,}"

# ── ATH Override Memory (persists across sessions via JSON file) ──
_ATH_MEMORY_FILE = "ath_memory.json"

def _load_ath_memory() -> dict:
    """Load ATH override memory from local JSON file.
    Returns empty dict if file not found or unreadable."""
    try:
        if os.path.exists(_ATH_MEMORY_FILE):
            with open(_ATH_MEMORY_FILE, "r") as f:
                data = json.load(f)
            if isinstance(data, dict):
                return data
    except Exception:
        pass
    return {}

def _save_ath_memory(mem: dict) -> bool:
    """Save ATH override memory to local JSON file.
    Returns True on success, False on error."""
    try:
        with open(_ATH_MEMORY_FILE, "w") as f:
            json.dump(mem, f, indent=2, ensure_ascii=False)
        return True
    except Exception:
        return False

# ── Rebalance Memory helpers ─────────────────────────────────────
_RB_MEMORY_FILE = "rebalance_memory.json"

def _load_rb_memory() -> dict:
    """Load rebalance memory from local JSON, fallback GitHub."""
    try:
        if os.path.exists(_RB_MEMORY_FILE):
            with open(_RB_MEMORY_FILE, "r") as f:
                return json.load(f)
    except Exception:
        pass
    try:
        r = requests.get(
            f"https://raw.githubusercontent.com/{_GH_OWNER}/{_GH_REPO}/main/rebalance_memory.json",
            timeout=8,
        )
        if r.status_code == 200:
            return r.json()
    except Exception:
        pass
    return {}

def _save_rb_memory(mem: dict) -> bool:
    try:
        with open(_RB_MEMORY_FILE, "w") as f:
            json.dump(mem, f, indent=2, ensure_ascii=False)
        return True
    except Exception:
        return False

def step_html(current):
    steps = [(1,"Universe Setup"),(2,"Run Screener"),(3,"Plan Rebalance"),(4,"Apply & Export")]
    html = '<div class="step-bar">'
    for i,(n,label) in enumerate(steps):
        cls = "done" if n < current else ("active" if n == current else "pending")
        sym = "✓"    if n < current else str(n)
        html += f'<div class="step-item {cls}"><div class="step-circle">{sym}</div>{label}</div>'
        if i < len(steps)-1:
            lc = "done-line" if n < current else ""
            html += f'<div class="step-connector {lc}"></div>'
    return html + '</div>'

def metric_card(label, value, color=""):
    cls = f"metric-card {color}" if color else "metric-card"
    val_cls = f"metric-value {color}" if color else "metric-value"
    return f'<div class="{cls}"><div class="metric-label">{label}</div><div class="{val_cls}">{value}</div></div>'

def parse_equity_csv(f) -> pd.DataFrame:
    df = pd.read_csv(f, skipinitialspace=True)
    df.columns = [c.strip() for c in df.columns]
    if 'SERIES' in df.columns:
        df = df[df['SERIES'].str.strip() == 'EQ'].copy()
    df['SYMBOL'] = df['SYMBOL'].str.strip().str.upper()
    return df.reset_index(drop=True)

def load_symbols_from_github(universe: str) -> list:
    """Returns list of .NS symbols for the chosen universe (not AllNSE)."""
    if universe == 'N750':
        url = f"{GITHUB_BASE}/ind_niftytotalmarket_list.csv"
    else:
        url = f"{GITHUB_BASE}/ind_{universe.lower()}list.csv"
    df = pd.read_csv(url)
    df['Yahoo_Symbol'] = df['Symbol'].astype(str).str.strip() + '.NS'
    return df['Yahoo_Symbol'].tolist()

# ── Always-include symbols (added to every universe) ──────────
EXTRA_SYMBOLS = ["GOLDBEES.NS", "SILVERBEES.NS"]

def add_extra_symbols(syms: list) -> list:
    """Append GOLDBEES & SILVERBEES if not already present."""
    result = list(syms)
    for s in EXTRA_SYMBOLS:
        if s not in result:
            result.append(s)
    return result

def build_dates(end_date: datetime.date) -> dict:
    from dateutil.relativedelta import relativedelta
    end = datetime.datetime.combine(end_date, datetime.time())
    return {
        'startDate': datetime.datetime(2000, 1, 1),
        'endDate':   end,
        'date12M':   end - relativedelta(months=12),
        'date9M':    end - relativedelta(months=9),
        'date6M':    end - relativedelta(months=6),
        'date3M':    end - relativedelta(months=3),
        'date1M':    end - relativedelta(months=1),
    }

# ── v10-identical Excel formatting ────────────────────────────
def format_excel_unfiltered(file_name, universe, top_n):
    """Format 'Unfiltered Stocks' sheet — exact v10 logic."""
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"),  bottom=Side(style="thin"))
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border    = thin
            cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = 'A2'
    hdr_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")
    for col in range(1, ws.max_column + 1):
        c = ws.cell(row=1, column=col)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center")
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    no_fill   = PatternFill(start_color="d6b4fc", end_color="d6b4fc", fill_type="solid")
    bold_font = Font(bold=True)
    headers   = [c.value for c in ws[1]]

    def ci(name): return headers.index(name) + 1 if name in headers else None

    rank_threshold = top_n
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    idx = {k: ci(k) for k in ['volm_cr','Close','dma200d','AWAY_ATH','roc12M',
                                'circuit','roc1M','circuit5','Ticker','Rank']}

    for row in range(2, ws.max_row + 1):
        failed = False
        def v(col): return ws.cell(row=row, column=col).value if col else None
        def mark(col):
            nonlocal failed
            ws.cell(row=row, column=col).fill = no_fill
            ws.cell(row=row, column=col).font = bold_font
            failed = True
        if (vol := v(idx['volm_cr']))  is not None and vol < 1:           mark(idx['volm_cr'])
        cl = v(idx['Close']); dm = v(idx['dma200d'])
        if cl is not None and dm is not None and cl <= dm:                 mark(idx['Close'])
        if (aa := v(idx['AWAY_ATH']))  is not None and aa <= -25:         mark(idx['AWAY_ATH'])
        roc = v(idx['roc12M'])
        if roc is not None and roc <= 5.5:                                 mark(idx['roc12M'])
        if (ci_ := v(idx['circuit']))  is not None and ci_ >= 20:         mark(idx['circuit'])
        if cl is not None and cl <= 30:                                    mark(idx['Close'])
        if (c5 := v(idx['circuit5']))  is not None and c5 > 10:           mark(idx['circuit5'])
        if roc is not None and roc > 1000:                                 mark(idx['roc12M'])
        if failed and idx['Ticker']:
            ws.cell(row=row, column=idx['Ticker']).fill = no_fill
        if idx['Rank'] and (rk := v(idx['Rank'])) is not None and rk <= rank_threshold:
            ws.cell(row=row, column=idx['Rank']).fill = green_fill

    # ATH round
    ath_col = ci('ATH')
    if ath_col:
        for r in range(2, ws.max_row + 1):
            c = ws.cell(row=r, column=ath_col)
            if isinstance(c.value, (int, float)):
                c.value = round(c.value)
    wb.save(file_name)


def format_excel_filtered(file_name, universe, top_n):
    """Format 'Filtered Stocks' sheet — exact v10 logic."""
    wb = openpyxl.load_workbook(file_name)
    ws = wb["Filtered Stocks"]
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"),  bottom=Side(style="thin"))
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border    = thin
            cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = 'A2'
    hdr_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")
    for col in range(1, ws.max_column + 1):
        c = ws.cell(row=1, column=col)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center")
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    # ATH round
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == "ATH":
            for r in range(2, ws.max_row + 1):
                c = ws.cell(row=r, column=col)
                if isinstance(c.value, (int, float)):
                    c.value = round(c.value)
            break
    # AWAY_ATH % suffix
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == "AWAY_ATH":
            for r in range(2, ws.max_row + 1):
                c = ws.cell(row=r, column=col)
                if isinstance(c.value, (int, float)):
                    c.value = f"{c.value}%"
            break

    # Rank highlight + summary
    rank_threshold = top_n
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == "Rank":
            rank_75_count = 0
            for r in range(2, ws.max_row + 1):
                c = ws.cell(row=r, column=col)
                if isinstance(c.value, (int, float)) and c.value <= rank_threshold:
                    c.fill = green_fill
                    rank_75_count += 1
            total_filtered = ws.max_row - 1
            ws.append([])
            ws.append(["Summary"])
            summary_start = ws.max_row
            ws.append([f"Total Filtered Stocks: {total_filtered}"])
            ws.append([f"Number of Stocks within {rank_threshold} Rank: {rank_75_count}"])
            for r in ws.iter_rows(min_row=summary_start, max_row=ws.max_row, min_col=1, max_col=1):
                for cell in r:
                    cell.font = Font(bold=True)
            break
    wb.save(file_name)


def format_simple_sheet(file_name, sheet_name):
    """Format 'Failed Downloads' and 'Portfolio Rebalancing' sheets — exact v10 logic."""
    wb = openpyxl.load_workbook(file_name)
    if sheet_name not in wb.sheetnames:
        wb.save(file_name); return
    ws = wb[sheet_name]
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"),  bottom=Side(style="thin"))
    hdr_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border    = thin
            cell.alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, ws.max_column + 1):
        c = ws.cell(row=1, column=col)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center")
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
    ws.freeze_panes = 'A2'

    if sheet_name == "Portfolio Rebalancing":
        headers  = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        sell_col = headers.index('Sell Stocks') + 1 if 'Sell Stocks' in headers else None
        buy_col  = headers.index('Buy Stocks')  + 1 if 'Buy Stocks'  in headers else None
        sell_fill = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")
        buy_fill  = PatternFill(start_color="D7FFD7", end_color="D7FFD7", fill_type="solid")
        for r in range(2, ws.max_row + 1):
            if sell_col: ws.cell(row=r, column=sell_col).fill = sell_fill
            if buy_col:  ws.cell(row=r, column=buy_col).fill  = buy_fill

    if sheet_name == "Failed Downloads":
        headers   = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        stock_col = headers.index('Failed Stock') + 1 if 'Failed Stock' in headers else None
        org_fill  = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
        if stock_col:
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=stock_col).fill = org_fill
    wb.save(file_name)


# ═══════════════════════════════════════════════════════════════
# LOGIN
# ═══════════════════════════════════════════════════════════════
def login_page():
    _, mid, _ = st.columns([1, 1.4, 1])
    with mid:
        st.markdown("""
        <div class="login-card">
          <div class="login-logo">📈</div>
          <div class="login-title">Momn Screener</div>
          <div class="login-sub">NSE Momentum Strategy &nbsp;·&nbsp; v13</div>
        </div>
        <div style="height:12px;"></div>
        """, unsafe_allow_html=True)
        with st.form(key="login_form", clear_on_submit=True):
            u = st.text_input("👤 Username", placeholder="Enter username")
            p = st.text_input("🔒 Password", type="password", placeholder="Enter password")
            if st.form_submit_button(label="Sign In →", use_container_width=True, type="primary"):
                if u == USERNAME and p == PASSWORD:
                    st.session_state.logged_in = True
                    st.rerun()
                else:
                    st.error("❌ Invalid username or password")

if not st.session_state.logged_in:
    login_page()
    st.stop()

# ═══════════════════════════════════════════════════════════════
# HEADER
# ═══════════════════════════════════════════════════════════════
_today_str = datetime.date.today().strftime("%d %b %Y")
st.markdown(f"""
<div class="app-header">
  <div>
    <div class="app-title">📈 <span>Momn</span> Screener + Rebalancer <span class="app-badge">v13</span></div>
    <div class="app-subtitle">NSE Momentum Strategy &nbsp;·&nbsp; Equal-Weight Monthly Rebalancing</div>
  </div>
  <div class="app-header-right">
    <div class="user-tag">👤 prayan2702</div><br>
    <span style="font-size:10.5px;">📅 {_today_str}</span>
  </div>
</div>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════
# SIDEBAR
# ═══════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("### ⚙️ Workflow Steps")
    step_labels = {1:"Universe Setup", 2:"Run Screener",
                   3:"Plan Rebalance", 4:"Apply & Export"}
    step_icons  = {1:"🌐", 2:"📊", 3:"⚖️", 4:"💾"}
    for s, lbl in step_labels.items():
        is_active = (st.session_state.current_step == s)
        is_done   = (s == 1 and st.session_state.symbols is not None) or \
                    (s == 2 and st.session_state.screener_done) or \
                    (s == 3 and st.session_state.rebalance_done)
        dot = "✓" if is_done else ("→" if is_active else "○")
        label_text = f"{dot} {step_icons[s]} {s}. {lbl}"
        if st.button(label_text, key=f"nav_{s}", use_container_width=True,
                     type="primary" if is_active else "secondary"):
            st.session_state.current_step = s; st.rerun()

    st.divider()
    st.markdown("### 🔧 Screener Settings")
    rm_display = st.selectbox("📐 Ranking Method", list(RANKING_MAP.keys()), index=0)
    st.session_state.ranking_method = RANKING_MAP[rm_display]

    st.session_state.data_source = st.selectbox("📡 Data Source", API_OPTIONS, index=0)

    # ── 5-Day Rolling Cache Date Selector ──────────────────
    _ds = st.session_state.data_source
    _avail_dates = []
    try:
        if "Upstox" in _ds and _CACHE_UPSTOX_AVAILABLE:
            from cache_loader_upstox import list_available_dates as _ld_dates
            _avail_dates = _ld_dates()
        elif "Angel" in _ds and _CACHE_ANGEL_AVAILABLE:
            from cache_loader_angelone import list_available_dates as _ld_dates
            _avail_dates = _ld_dates()
        elif _CACHE_AVAILABLE:
            from cache_loader import list_available_dates as _ld_dates
            _avail_dates = _ld_dates()
    except Exception:
        _avail_dates = []

    if _avail_dates and "Pre-cached" in _ds:
        _date_opts   = list(reversed(_avail_dates))  # latest first
        _prev_sel    = st.session_state.get("cache_selected_date", _date_opts[0])
        _default_idx = _date_opts.index(_prev_sel) if _prev_sel in _date_opts else 0
        _sel_date    = st.selectbox(
            "📅 Cache Date",
            options=_date_opts,
            index=_default_idx,
            help=f"5 cached dates available. Latest = {_date_opts[0]}",
            key="cache_date_selectbox",
        )
        st.session_state["cache_selected_date"] = _sel_date
        if _sel_date == _date_opts[0]:
            st.caption("✅ Latest cache loaded")
        else:
            st.caption(f"⚠️ Historical cache: {_sel_date}")
    else:
        st.session_state["cache_selected_date"] = None

    st.session_state.lookback_date = st.date_input(
        "📅 Lookback Date", value=st.session_state.lookback_date,
        max_value=datetime.date.today()
    )
    st.session_state.top_n_rank = st.number_input(
        "🏆 Top-N Rank", min_value=20, max_value=200, value=100, step=10
    )

    # ── API Authentication ──────────────────────────────────
    if st.session_state.data_source == "Upstox":
        st.divider()
        if _UPSTOX_AVAILABLE:
            get_upstox_access_token(sidebar=True)
        else:
            st.warning("⚠️ `pyotp` install nahi hai. YFinance fallback use hoga.")

    elif st.session_state.data_source == "Angel One":
        st.divider()
        if _ANGEL_AVAILABLE:
            _ao_client = get_angelone_client(sidebar=True)
            if _ao_client:
                st.session_state["angelone_client"] = _ao_client
        else:
            st.sidebar.markdown("""
            <div style="background:#fef3c7;border:1px solid #fcd34d;border-left:4px solid #d97706;
                        border-radius:10px;padding:10px 14px;font-size:12px;color:#92400e;margin:6px 0;">
              ⚠️ <b>Angel One unavailable</b><br>
              <span style="font-size:11px;">
              <code>requirements.txt</code> mein add karo: <code>smartapi-python</code> + <code>pyotp</code><br>
              Abhi <b>YFinance</b> fallback use hoga.
              </span>
            </div>""", unsafe_allow_html=True)

    elif st.session_state.data_source == "Fyers":
        st.divider()
        if _FYERS_AVAILABLE:
            get_fyers_client(sidebar=True)
        else:
            st.sidebar.markdown("""
            <div style="background:#fef3c7;border:1px solid #fcd34d;border-left:4px solid #d97706;
                        border-radius:10px;padding:10px 14px;font-size:12px;color:#92400e;margin:6px 0;">
              ⚠️ <b>Fyers unavailable</b><br>
              <span style="font-size:11px;">
              <code>requirements.txt</code> mein add karo: <code>fyers-apiv3</code> + <code>pyotp</code><br>
              Abhi <b>YFinance</b> fallback use hoga.
              </span>
            </div>""", unsafe_allow_html=True)

    st.divider()
    st.markdown("### 🔗 Quick Links")
    st.markdown(f"""
    <div style="font-size:12.5px;line-height:2.6;color:var(--text-sub);">
    <a href="https://www.nseindia.com/static/market-data/securities-available-for-trading" target="_blank"
       style="color:var(--blue);font-weight:600;text-decoration:none;">📥 NSE EQUITY_L.csv</a><br>
    <a href="{APPS_SCRIPT_URL}" target="_blank"
       style="color:var(--teal);font-weight:600;text-decoration:none;">⚖️ Portfolio Rebalancer</a><br>
    <a href="https://prayan2702.github.io/momn-dashboard/" target="_blank"
       style="color:var(--violet);font-weight:600;text-decoration:none;">📈 Portfolio Dashboard</a>
    </div>
    """, unsafe_allow_html=True)

    # ── ATH Memory Panel ─────────────────────────────────────────
    st.divider()
    _ath_mem_sidebar = st.session_state.get("_ath_memory") or {}
    _n_mem = len(_ath_mem_sidebar)
    st.markdown(
        f"### 📚 ATH Memory &nbsp;"
        f"<span style='font-size:11px;color:var(--muted);font-weight:400;'>({_n_mem} stocks)</span>",
        unsafe_allow_html=True,
    )
    if _n_mem > 0:
        _mem_bytes = json.dumps(_ath_mem_sidebar, indent=2, ensure_ascii=False).encode("utf-8")

        # ── Download button ──────────────────────────────────
        st.download_button(
            label="💾 Download ath_memory.json",
            data=_mem_bytes,
            file_name="ath_memory.json",
            mime="application/json",
            use_container_width=True,
        )

        # ── Push to GitHub (PIN-protected) ───────────────────
        with st.expander("☁️ GitHub pe push karo", expanded=False):
            _ath_pin_inp = st.text_input(
                "🔑 PIN", type="password",
                key="ath_push_pin", placeholder="Streamlit secret PIN"
            )
            if st.button("📤 ath_memory.json → GitHub", use_container_width=True, key="ath_gh_push"):
                if not _verify_pin(_ath_pin_inp):
                    if not _pin_secret_exists():
                        st.error("❌ TRIGGER_PIN secret set nahi hai")
                        st.caption("ℹ️ Streamlit Secrets mein `TRIGGER_PIN = \"your_pin\"` add karo")
                    else:
                        st.error("❌ Wrong PIN")
                elif not _get_secret("GITHUB_PAT"):
                    st.error("❌ GITHUB_PAT secret set nahi hai")
                else:
                    with st.spinner("☁️ GitHub pe push ho raha hai..."):
                        _ok, _msg = _push_json_to_github(
                            path="ath_memory.json",
                            content_dict=_ath_mem_sidebar,
                            commit_msg=f"🤖 ATH memory update ({_n_mem} stocks) — {datetime.date.today()}",
                        )
                    if _ok:
                        st.success(_msg)
                    else:
                        st.error(_msg)

        # ── Clear + Preview ──────────────────────────────────
        if st.button("🗑 Memory Clear karo", use_container_width=True, key="sb_clear_mem"):
            st.session_state["_ath_memory"] = {}
            st.session_state["_ath_memory_loaded"] = True
            _save_ath_memory({})
            st.rerun()

        with st.expander(f"📋 Memory preview ({_n_mem} stocks)"):
            for _t, _e in list(_ath_mem_sidebar.items())[:20]:
                st.markdown(
                    f"<div style='font-size:11px;line-height:1.8;'>"
                    f"<b style='color:var(--text-main);'>{_t}</b> &nbsp;"
                    f"<span style='color:var(--teal);'>{_e.get('chosen_lbl','?')}</span> &nbsp;"
                    f"<span style='color:var(--muted);'>ATH {_e.get('chosen_ath',0):,.0f}</span> &nbsp;"
                    f"<span style='color:var(--muted);font-size:10px;'>{_e.get('reviewed_date','')}</span>"
                    f"</div>",
                    unsafe_allow_html=True,
                )
            if _n_mem > 20:
                st.caption(f"... aur {_n_mem - 20} stocks (download karo full list)")
    else:
        st.caption("Abhi koi ATH review save nahi hua. Cross-source review karke 'Apply' karo.")

    # ── Rebalance Memory Panel ───────────────────────────────────────
    st.divider()
    st.markdown("### 💾 Rebalance Memory", unsafe_allow_html=True)
    with st.expander("Save / View Rebalance Memory", expanded=False):
        st.caption("Monthly rebalance ke baad state save karo (PIN-protected — same PIN as ATH).")
        _rb_mem_sb = st.session_state.get("_rb_memory", {})
        if _rb_mem_sb:
            _rb_al = _rb_mem_sb.get("allocation", {})
            st.markdown(
                f"📅 **{_rb_mem_sb.get('last_rebalance_date','—')}** &nbsp;|&nbsp; "
                f"Band **{_rb_mem_sb.get('regime_band','—')}** &nbsp;|&nbsp; "
                f"Score **{_rb_mem_sb.get('weighted_score','—')}**\n\n"
                f"Eq **{_rb_al.get('equity_pct','—')}%** / Gold **{_rb_al.get('gold_pct','—')}%** / Cash **{_rb_al.get('cash_pct','—')}%**"
            )
        else:
            st.info("No rebalance memory saved yet.")

        _rs_sb = st.session_state.get("_regime_state")
        if _rs_sb is not None:
            _rb_notes_inp = st.text_area("Notes (optional):", key="rb_notes_sb", height=100)
            _rb_pin_inp   = st.text_input("🔑 PIN", type="password", max_chars=6, key="rb_pin_sb")
            if st.button("💾 Save Rebalance Memory", use_container_width=True, key="rb_save_sb"):
                if not _verify_pin(_rb_pin_inp):
                    if not _pin_secret_exists():
                        st.error("❌ TRIGGER_PIN secret set nahi hai")
                    else:
                        st.error("❌ Wrong PIN")
                else:
                    _rb_data = {
                        "last_rebalance_date": str(datetime.date.today()),
                        "regime_band":         _rs_sb.effective_band,
                        "weighted_score":      round(_rs_sb.raw_score, 3),
                        "qfsm_mode":           _rs_sb.qfsm_mode,
                        "status":              _rs_sb.status,
                        "allocation": {
                            "equity_pct":  round(_rs_sb.equity * 100, 1),
                            "gold_pct":    round(_rs_sb.gold   * 100, 1),
                            "cash_pct":    round(_rs_sb.cash   * 100, 1),
                        },
                        "signals":           _rs_sb.signals,
                        "vix_overlay_pct":   _rs_sb.vix_overlay_pct,
                        "notes":             _rb_notes_inp,
                    }
                    _save_rb_memory(_rb_data)
                    st.session_state["_rb_memory"] = _rb_data
                    if _get_secret("GITHUB_PAT", ""):
                        with st.spinner("☁️ GitHub pe push ho raha hai..."):
                            _ok2, _msg2 = _push_json_to_github(
                                path="rebalance_memory.json",
                                content_dict=_rb_data,
                                commit_msg=f"auto: rebalance memory {_rb_data['last_rebalance_date']}",
                            )
                        st.success(_msg2 if _ok2 else f"Local saved. GitHub: {_msg2}")
                    else:
                        st.success("✅ Saved locally. (Set GITHUB_PAT to also push to GitHub.)")
        else:
            st.caption("Market Regime tab mein refresh karo → tab state load hogi → phir save karo.")

    # ── GitHub Actions Trigger Panel ─────────────────────────────
    st.divider()
    st.markdown("### ⚡ GitHub Actions", unsafe_allow_html=True)
    with st.expander("🔄 Cache rebuild trigger karo", expanded=False):
        st.caption("PIN se protected — workflows manually trigger honge")
        _ga_pin = st.text_input(
            "🔑 PIN", type="password",
            key="ga_trigger_pin", placeholder="Streamlit secret PIN"
        )
        _ga_pin_ok = _verify_pin(_ga_pin) if _ga_pin else False

        _wf_map = {
            "📦 YFinance":   _WF_YFINANCE,
            "📡 Upstox":     _WF_UPSTOX,
            "🤖 Angel One":  _WF_ANGEL,
            "📊 FII / DII":  _WF_FII,
        }
        for _lbl, _wf in _wf_map.items():
            if st.button(
                f"🔄 Trigger {_lbl}",
                use_container_width=True,
                key=f"trigger_{_wf}",
                disabled=not _ga_pin_ok,
            ):
                if not _get_secret("GITHUB_PAT"):
                    st.error("❌ GITHUB_PAT secret set nahi hai")
                else:
                    with st.spinner(f"GitHub se {_lbl} workflow trigger ho raha hai..."):
                        _ok, _msg = _trigger_workflow(_wf)
                    if _ok:
                        st.success(_msg)
                        st.caption("⏱ Build ~20-90 min lagega. GitHub Actions tab mein progress dekho.")
                    else:
                        st.error(_msg)

        if not _ga_pin_ok and _ga_pin:
            _pin_exists = _pin_secret_exists()
            if not _pin_exists:
                st.error("❌ TRIGGER_PIN secret Streamlit mein set nahi hai")
                st.caption("ℹ️ Settings → Secrets mein `TRIGGER_PIN = \"your_pin\"` add karo (quotes zaroori hain)")
            else:
                st.error("❌ Wrong PIN")
        elif not _ga_pin:
            st.caption("PIN daalo → buttons active honge")

    if st.button("🚪 Logout", use_container_width=True):
        for k in list(st.session_state.keys()):
            del st.session_state[k]
        st.rerun()

    st.markdown(f"""
    <div style="margin-top:12px;padding:10px;background:var(--bg);border:1px solid var(--border);
                border-radius:var(--radius-md);text-align:center;font-size:10.5px;color:var(--muted);">
    📅 {datetime.date.today().strftime('%d %b %Y')}<br>
    <span style="color:var(--teal);font-weight:700;">Momn Screener v13</span>
    </div>
    """, unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════
# TOP-LEVEL TABS
# ══════════════════════════════════════════════════════════════════
import math as _math_rt, datetime as _dt_rt

# ── Fast NAV + VIX helpers ─────────────────────────────────────
@st.cache_data(ttl=300, show_spinner=False)
def _fetch_vix_yf():
    """India VIX from yfinance — tries fast_info first, then download fallback."""
    try:
        import yfinance as _yf
        # Attempt 1: fast_info (instantaneous, no download)
        try:
            _tk = _yf.Ticker("^INDIAVIX")
            _info = _tk.fast_info
            _vix = getattr(_info, "lastPrice", None) or getattr(_info, "last_price", None)
            if _vix and float(_vix) > 0:
                return round(float(_vix), 2)
        except Exception:
            pass
        # Attempt 2: download last 5 days
        for _period in ["5d", "1mo"]:
            try:
                _df = _yf.download(
                    "^INDIAVIX", period=_period, progress=False,
                    auto_adjust=True, multi_level_index=False
                )
                if not _df.empty:
                    _last = _df["Close"].dropna()
                    if len(_last) > 0:
                        return round(float(_last.iloc[-1]), 2)
            except Exception:
                continue
    except Exception:
        pass
    return None

@st.cache_data(ttl=300, show_spinner=False)
def _fetch_nav_from_sheet(sheet_csv_url):
    """NAV series from Google Sheet public CSV — column G (index 6), skip 2 header rows."""
    try:
        import requests as _rq, io as _io, pandas as _pd_nav
        r = _rq.get(sheet_csv_url, timeout=12, headers={"User-Agent": "Mozilla/5.0"})
        if r.status_code == 200:
            df = _pd_nav.read_csv(_io.StringIO(r.text), header=None, skiprows=2)
            # Column G = index 6 has NAV values
            nav_col = _pd_nav.to_numeric(df[6], errors="coerce").dropna()
            nav_series = nav_col[(nav_col > 0) & (nav_col < 100000)].tolist()
            if len(nav_series) > 5:
                return nav_series
    except Exception:
        pass
    return []

# Google Sheet NAV CSV URL — update gid to your NAV sheet tab
# Format: https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=csv&gid={GID}
# Or use pub CSV URL from File > Share > Publish to web
_NAV_SHEET_CSV = "https://docs.google.com/spreadsheets/d/e/2PACX-1vSAf1ZhBXcMmi_QSxHnU6LI9PLKGbgEu16-FNKBjELhTk2x6qwLuuG7jAFx-mIn_ezNn45uL1G00MD1/pub?gid=612197947&single=true&output=csv"

def _build_mmi_gauge(sc, fc, lbl, em, src_lbl="", date_str=""):
    """MMI-style semicircle speedometer gauge — properly centered."""
    import math as _gm
    # Geometry — center at (155,155), semicircle arc spans 180° (left→top→right)
    cx, cy   = 155, 155   # needle pivot / arc center
    r_arc    = 108        # radius of arc midline
    r_inner  = 88         # inner radius of track
    r_outer  = 128        # outer radius of track (track width ≈ 40px)
    r_needle = 98         # needle tip length

    def _pt(angle_deg, radius):
        """Math angle (0=right,90=up,180=left) → SVG (x,y)."""
        rad = _gm.radians(angle_deg)
        return (cx + radius * _gm.cos(rad), cy - radius * _gm.sin(rad))

    # Segment boundary angles: Bear 180→135, Neutral 135→90, MildBull 90→45, StrongBull 45→0
    seg_defs = [
        (180, 135, "#ef4444"),
        (135,  90, "#f59e0b"),
        ( 90,  45, "#38bdf8"),
        ( 45,   0, "#10b981"),
    ]
    seg_mid_angles   = [157.5, 112.5, 67.5, 22.5]
    active_idx       = min(sc, 3)

    # ── Background track (grey ring)
    bx0, by0 = _pt(180, r_arc)
    bx1, by1 = _pt(  0, r_arc)
    bg_svg = (f'<path d="M {bx0:.1f} {by0:.1f} A {r_arc} {r_arc} 0 0 1 {bx1:.1f} {by1:.1f}" '
              f'fill="none" stroke="#1e2736" stroke-width="40" stroke-linecap="butt"/>')

    # ── Color segments
    segs_svg = ""
    for i, (sa, ea, col) in enumerate(seg_defs):
        sx, sy = _pt(sa, r_arc)
        ex, ey = _pt(ea, r_arc)
        is_active = (i == active_idx)
        sw  = "40" if is_active else "28"
        op  = "1"  if is_active else "0.50"
        segs_svg += (f'<path d="M {sx:.1f} {sy:.1f} A {r_arc} {r_arc} 0 0 1 {ex:.1f} {ey:.1f}" '
                     f'fill="none" stroke="{col}" stroke-width="{sw}" stroke-linecap="butt" opacity="{op}"/>')

    # ── Divider ticks between segments
    ticks_svg = ""
    for ta in [135, 90, 45]:
        t0x, t0y = _pt(ta, r_inner - 4)
        t1x, t1y = _pt(ta, r_outer + 4)
        ticks_svg += (f'<line x1="{t0x:.1f}" y1="{t0y:.1f}" '
                      f'x2="{t1x:.1f}" y2="{t1y:.1f}" stroke="#0d1520" stroke-width="4"/>')

    # ── Labels just outside the arc (r_outer + 16)
    r_lbl = r_outer + 18
    lbl_defs = [
        (175, "Bear",     "#ef4444", "end"),
        (128, "Neutral",  "#f59e0b", "end"),
        ( 52, "Mild Bull","#38bdf8", "start"),
        (  5, "Strong",   "#10b981", "start"),
    ]
    lbl_svg = ""
    for i, (la, lt, lc, anc) in enumerate(lbl_defs):
        lx, ly = _pt(la, r_lbl)
        fw = "800" if i == active_idx else "600"
        lbl_svg += (f'<text x="{lx:.1f}" y="{ly:.1f}" text-anchor="{anc}" '
                    f'dominant-baseline="central" font-family="Segoe UI,sans-serif" '
                    f'font-size="11" font-weight="{fw}" fill="{lc}">{lt}</text>')

    # ── Score number in center
    score_svg = (f'<text x="{cx}" y="{cy+28}" text-anchor="middle" dominant-baseline="central" '
                 f'font-family="Segoe UI,sans-serif" font-size="30" font-weight="900" '
                 f'fill="white" opacity="0.50">{sc}</text>')

    # ── Needle
    ntx, nty = _pt(seg_mid_angles[active_idx], r_needle)
    cnx, cny = _pt(seg_mid_angles[active_idx] + 180, 16)

    src_d = (f'<div style="font-size:10px;color:#94a3b8;margin-top:2px;">'
             f'{src_lbl}</div>') if src_lbl else ""
    dt_d  = (f'<div style="font-size:10px;color:#64748b;margin-top:3px;">'
             f'{date_str} &nbsp;·&nbsp; Score {sc}/3</div>') if date_str else ""

    # viewBox: left margin so Bear label fits, right margin so Strong label fits
    # Bear at 175°: x = 155 + 146*cos(175°) ≈ 155 - 145.4 = 9.6  → left edge ~0
    # Strong at 5°: x = 155 + 146*cos(5°)   ≈ 155 + 145.4 = 300.4 → right edge ~310
    # Arc top at 90°: y = 155 - 108 = 47; label top at 155-146 = 9 → viewBox y start = 5
    return (
        '<style>'
        '*{box-sizing:border-box;margin:0;padding:0;}'
        'body{background:transparent;font-family:"Segoe UI",system-ui,sans-serif;}'
        '.gcard{background:#111827;border:1px solid #1e293b;border-radius:16px;'
        'padding:26px 32px 26px;text-align:center;display:block;width:fit-content;margin:0 auto;}'
        '.gtitle{font-size:9px;color:#b3bbc7;text-transform:uppercase;letter-spacing:1.5px;font-weight:700;margin-bottom:4px;}'
        f'.gname{{font-size:22px;font-weight:800;color:{fc};margin-top:4px;}}'
        '.leg{display:flex;gap:8px;justify-content:center;flex-wrap:wrap;margin-top:10px;}'
        '.li{display:flex;align-items:center;gap:4px;font-size:10px;color:#94a3b8;}'
        '.ld{width:7px;height:7px;border-radius:50%;flex-shrink:0;}'
        'svg{display:block;margin:0 auto;overflow:visible;}'
        '</style>'
        '<div class="gcard">'
        '<div class="gtitle">MARKET REGIME</div>'
        '<svg width="350" height="260" viewBox="0 5 320 175">'
        f'{bg_svg}'
        f'{segs_svg}'
        f'{ticks_svg}'
        f'{lbl_svg}'
        f'{score_svg}'
        f'<line x1="{cnx:.1f}" y1="{cny:.1f}" x2="{ntx:.1f}" y2="{nty:.1f}" '
        f'stroke="white" stroke-width="3.5" stroke-linecap="round"/>'
        f'<circle cx="{cx}" cy="{cy}" r="10" fill="{fc}" stroke="#111827" stroke-width="3"/>'
        f'<circle cx="{cx}" cy="{cy}" r="4"  fill="white"/>'
        '</svg>'
        f'<div class="gname">{em} {lbl}</div>'
        f'{dt_d}{src_d}'
        '<div class="leg">'
        '<div class="li"><div class="ld" style="background:#ef4444"></div>Bear</div>'
        '<div class="li"><div class="ld" style="background:#f59e0b"></div>Neutral (1)</div>'
        '<div class="li"><div class="ld" style="background:#38bdf8"></div>Mild Bull (2)</div>'
        '<div class="li"><div class="ld" style="background:#10b981"></div>Strong Bull (3)</div>'
        '</div></div>'
    )

_tab_screener, _tab_regime, _tab_sim = st.tabs([
    "📊 Screener & Rebalancer",
    "🌡️ Market Regime",
    "🧪 Multi-Asset Simulator",
])

# ════════════════════════════════════════════════════
# REGIME TAB — v2026.08  (7-signal weighted, QFSM)
# ════════════════════════════════════════════════════
with _tab_regime:
    st.session_state["_curr_tab"] = "regime"
    try:
        from calculations import (
            get_full_regime_result, get_next_rebalance_dates,
            RegimeState, score_to_band,
        )
        import streamlit.components.v1 as _stc_rt
        _rt_calcs_ok = True
    except ImportError as _imp_err:
        st.warning(f"⚠️ calculations.py import failed: {_imp_err}")
        _rt_calcs_ok = False

    if _rt_calcs_ok:

        @st.cache_data(ttl=3600, show_spinner=False)
        def _load_best_cache_regime():
            try:
                if _CACHE_UPSTOX_AVAILABLE:
                    from cache_loader_upstox import load_cache as _lc
                    c,h,v = _lc(); return c,h,v,"Upstox"
            except Exception: pass
            try:
                if _CACHE_ANGEL_AVAILABLE:
                    from cache_loader_angelone import load_cache as _lc2
                    c,h,v = _lc2(); return c,h,v,"Angel One"
            except Exception: pass
            try:
                if _CACHE_AVAILABLE:
                    from cache_loader import load_cache as _lc3
                    c,h,v = _lc3(); return c,h,v,"YFinance"
            except Exception: pass
            return None,None,None,"None"

        @st.cache_data(ttl=1800, show_spinner=False)
        def _fetch_nifty_rt():
            """^NSEI close + 200DMA from yfinance (cached 30 min)."""
            try:
                import yfinance as _yf2
                _nf = _yf2.download("^NSEI", period="300d", progress=False)["Close"].dropna()
                if len(_nf) < 5: return None, None
                return (round(float(_nf.iloc[-1]),2),
                        round(float(_nf.rolling(200, min_periods=150).mean().iloc[-1]),2))
            except Exception:
                return None, None

        # ── Refresh ────────────────────────────────────────────────────────────
        _rt_hc1, _rt_hc2 = st.columns([5, 1])
        with _rt_hc2:
            if st.button("🔄 Refresh", key="rt_refresh_btn"):
                _load_best_cache_regime.clear()
                _fetch_nifty_rt.clear()
                for _k in ["_rt_dfS","_rt_src","_rt_navs","_rt_vx","_rt_wret",
                           "_rt_nifty_close","_rt_nifty_dma200","_regime_state"]:
                    st.session_state.pop(_k, None)
                st.rerun()

        # ── Load dfStats (once per session) ────────────────────────────────────
        if st.session_state.get("_rt_dfS") is None:
            with st.spinner("⚡ Best cache loading (Upstox → Angel One → YFinance)..."):
                _rt_cl, _rt_hi, _rt_vo, _rt_src_l = _load_best_cache_regime()
            if _rt_cl is not None:
                try:
                    from calculations import build_dfStats as _bds_rt
                    from dateutil.relativedelta import relativedelta as _rdelta
                    import datetime as _dtb2
                    _rt_end_d = (_rt_cl.index[-1].date()
                                 if hasattr(_rt_cl.index[-1],'date')
                                 else _dt_rt.date.today())
                    _rt_ed_dt = _dtb2.datetime.combine(_rt_end_d, _dtb2.time())
                    _rt_dts   = {
                        'startDate': _dtb2.datetime(2000,1,1), 'endDate': _rt_ed_dt,
                        'date12M': _rt_ed_dt-_rdelta(months=12),
                        'date9M':  _rt_ed_dt-_rdelta(months=9),
                        'date6M':  _rt_ed_dt-_rdelta(months=6),
                        'date3M':  _rt_ed_dt-_rdelta(months=3),
                        'date1M':  _rt_ed_dt-_rdelta(months=1),
                    }
                    _rt_dfS_obj = _bds_rt(_rt_cl, _rt_hi, _rt_vo, _rt_dts, "avgZScore12_6_3")
                    st.session_state["_rt_dfS"] = _rt_dfS_obj
                    st.session_state["_rt_src"] = _rt_src_l
                    st.success(f"✅ {_rt_src_l} cache loaded · {len(_rt_dfS_obj):,} stocks · {_rt_end_d}")
                except Exception as _e_rt2:
                    st.error(f"Calculation error: {_e_rt2}")
            else:
                st.warning("⚠️ Koi cache available nahi. Screener tab mein pehle run karo.")

        # ── Load live signals (VIX, NAV, Nifty) ───────────────────────────────
        if st.session_state.get("_rt_navs") is None:
            _rtv = _fetch_vix_yf()
            _rtn = _fetch_nav_from_sheet(_NAV_SHEET_CSV)
            _rtw = round((_rtn[-1]/_rtn[-6]-1)*100,2) if len(_rtn)>=6 else None
            st.session_state["_rt_navs"] = _rtn
            st.session_state["_rt_vx"]   = _rtv
            st.session_state["_rt_wret"] = _rtw

        if st.session_state.get("_rt_nifty_close") is None:
            _nc_rt, _nd_rt = _fetch_nifty_rt()
            st.session_state["_rt_nifty_close"]  = _nc_rt
            st.session_state["_rt_nifty_dma200"] = _nd_rt

        # ── Pull cached values ─────────────────────────────────────────────────
        _rt_dfS  = st.session_state.get("_rt_dfS")
        _rt_navs = st.session_state.get("_rt_navs", [])
        _rt_vx   = st.session_state.get("_rt_vx")
        _rt_wret = st.session_state.get("_rt_wret")
        _rt_src  = st.session_state.get("_rt_src", "—")
        _rt_nc   = st.session_state.get("_rt_nifty_close")
        _rt_nd   = st.session_state.get("_rt_nifty_dma200")
        _rt_rk_h = st.session_state.get("_rt_rank_history", [])
        _prev_rs = st.session_state.get("_regime_prev_state")

        if _rt_dfS is not None:

            # ── Compute full regime result ──────────────────────────────────────
            _rs = get_full_regime_result(
                dfStats=_rt_dfS,
                equity_nav_series=_rt_navs or None,
                vix_value=_rt_vx,
                nifty_close=_rt_nc,
                nifty_dma200=_rt_nd,
                rank_history=_rt_rk_h or None,
                fii_score=0.5,
                prev_state=_prev_rs,
                total_capital=0.0,
                dd_pct=0.0,
            )
            st.session_state["_regime_state"] = _rs

            _rt_dts2   = get_next_rebalance_dates()
            _rt_date_s = _dt_rt.date.today().strftime("%d %b %Y")
            _rt_sc     = _rs.effective_band
            _rt_lbl    = _rs.label()
            _rt_fc, _rt_em = {
                3: ("#00d09e", "🟢"), 2: ("#38bdf8", "🔵"),
                1: ("#f59e0b", "🟡"), 0: ("#f87171", "🔴"),
            }[_rt_sc]
            _rt_nf = _rt_dts2["next_friday"]
            _rt_nr = _rt_dts2["next_monthly_rb"]
            _rt_df = (_rt_nf - _dt_rt.date.today()).days

            # Status + date banner
            _st_c  = {"STABLE":"#15803d","PENDING":"#d97706","CONFIRMED":"#1d4ed8",
                      "DD_OVERRIDE":"#dc2626"}.get(_rs.status, "#6b7280")
            _st_bg = {"STABLE":"#dcfce7","PENDING":"#fef3c7","CONFIRMED":"#dbeafe",
                      "DD_OVERRIDE":"#fee2e2"}.get(_rs.status, "#f1f5f9")
            _vix_d = ""
            if _rt_vx:
                _vc0 = "#dc2626" if _rt_vx > 20 else "#15803d"
                _vix_d = (f'<div style="background:#fef3c7;border:1px solid #fcd34d;'
                           f'border-radius:8px;padding:6px 12px;font-size:12px;'
                           f'color:{_vc0};font-family:DM Mono,monospace;">'
                           f'VIX: <b>{round(_rt_vx,1)}</b>{"  🔴" if _rt_vx>20 else ""}</div>')

            st.markdown(
                f'<div style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:12px;">' +
                f'<div style="background:#dbeafe;border:1px solid #93c5fd;border-radius:8px;padding:6px 12px;font-size:12px;color:#1d4ed8;font-family:DM Mono,monospace;">📅 <b>Next Friday:</b> {_rt_nf.strftime("%d %b %Y")} ({_rt_df}d)</div>' +
                f'<div style="background:#dcfce7;border:1px solid #86efac;border-radius:8px;padding:6px 12px;font-size:12px;color:#15803d;font-family:DM Mono,monospace;">📆 <b>Monthly RB:</b> {_rt_nr.strftime("%d %b %Y")}</div>' +
                f'<div style="background:{_st_bg};border:1px solid;border-radius:8px;padding:6px 12px;font-size:12px;font-weight:700;color:{_st_c};font-family:DM Mono,monospace;">⚡ {_rs.status} ({_rs.confirmation_count}/2)</div>' +
                _vix_d +
                f'<div style="background:#f1f5f9;border:1px solid #cbd5e1;border-radius:8px;padding:6px 12px;font-size:11px;color:#475569;">📦 {_rt_src}</div>' +
                '</div>',
                unsafe_allow_html=True,
            )

            if _rs.dd_override_active:
                st.error("🚨 DD OVERRIDE — Portfolio DD ≥ 20%. Bear allocation forced (25/30/45). All signals bypassed.")

            # ── Gauge (reuses existing _build_mmi_gauge) ───────────────────────
            _rt_gauge_html = _build_mmi_gauge(_rt_sc, _rt_fc, _rt_lbl, _rt_em, _rt_src, _rt_date_s)

            # ── Signal card helper ─────────────────────────────────────────────
            def _mk_sig(icon, title, sub, val_txt, score_val, weight, ok, partial=False):
                c  = "#f59e0b" if partial else ("#00d09e" if ok else "#f87171")
                bg = "#fef3c7" if partial else ("#e8fdf2" if ok else "#fef2f2")
                bd = "#fcd34d" if partial else ("#86efac" if ok else "#fca5a5")
                return (f'<div class="sig" style="background:{bg};border-color:{bd}">' +
                        f'<div class="si">{icon}</div>' +
                        f'<div class="sb"><div class="st" style="color:{c}">{title}</div>' +
                        f'<div class="sc" style="color:#374151">{sub}</div>' +
                        f'<div class="sv" style="color:{c}">{val_txt}</div></div>' +
                        f'<div class="sbg" style="background:{c};color:white">{score_val:.2g}/{weight}pt</div>' +
                        '</div>')

            _sigs  = _rs.signals
            _smeta = _rs.signal_meta
            _s1v   = _sigs.get("s1_nav",0)
            _s2v   = _sigs.get("s2_breadth",0)
            _s3v   = _sigs.get("s3_roc",0)
            _s4v   = _sigs.get("s4_vix",0)
            _s5v   = _sigs.get("s5_nifty",0)
            _s6v   = _sigs.get("s6_ad",0)
            _s7v   = _sigs.get("s7_rank",0)

            _nav_txt = (f'NAV {_smeta["nav_current"]:.2f} vs DMA {_smeta["nav_dma200"]:.2f} ({_smeta.get("gap_pct",0):+.1f}%)' if _smeta.get("nav_current") else "NAV data loading...")
            _nif_txt = (f'Nifty {_rt_nc:,.0f} vs DMA {_rt_nd:,.0f} (ratio {_smeta.get("nifty_ratio",1):.3f})' if _rt_nc else "Nifty: loading...")

            _rt_sig_html = (
                '<style>body{margin:0;padding:0;background:transparent;font-family:"Segoe UI",sans-serif;}' +
                '.sigs{display:flex;flex-direction:column;gap:8px;}' +
                '.sig{border-radius:10px;padding:11px 14px;border:1.5px solid;display:flex;align-items:center;gap:12px;}' +
                '.si{font-size:20px;flex-shrink:0;}.sb{flex:1;}' +
                '.st{font-size:11px;font-weight:700;letter-spacing:.4px;text-transform:uppercase;margin-bottom:2px;}' +
                '.sc{font-size:11px;margin-bottom:3px;}.sv{font-size:13px;font-weight:800;}' +
                '.sbg{padding:3px 10px;border-radius:20px;font-size:11px;font-weight:700;flex-shrink:0;}' +
                '</style><div class="sigs">' +
                _mk_sig("✅" if _s1v>0 else "❌","S1 — Equity Curve Trend","NAV > 200-Day Moving Average",_nav_txt,_s1v,1.5,_s1v>0) +
                _mk_sig("✅" if _s2v>0 else "❌","S2 — Market Breadth","% Stocks > 200DMA > 50%",f'{_smeta.get("breadth_pct",0):.1f}% above DMA',_s2v,1.0,_s2v>0) +
                _mk_sig("✅" if _s3v>0 else "❌","S3 — Universe Momentum","Median 3M ROC > 0%",f'{_smeta.get("median_roc3m",0):+.1f}% median 3M ROC',_s3v,1.0,_s3v>0) +
                _mk_sig("⚠️" if 0<_s4v<1.5 else ("✅" if _s4v>0 else "❌"),"S4 — India VIX ★","VIX ≤ 20=PASS | 20-25=Partial | >25=FAIL",f'VIX {round(_rt_vx,1) if _rt_vx else "loading..."}',_s4v,1.5,_s4v>0,0<_s4v<1.5) +
                _mk_sig("⚠️" if 0<_s5v<1.5 else ("✅" if _s5v>0 else "❌"),"S5 — Nifty 200DMA ★","^NSEI Close > 200-Day Moving Average",_nif_txt,_s5v,1.5,_s5v>0,0<_s5v<1.5) +
                _mk_sig("✅" if _s6v>0 else "❌","S6 — A-D Ratio ★","Advances/Total > 45% (from 1M ROC)",f'A-D: {_smeta.get("ad_ratio",0.5)*100:.0f}%',_s6v,1.0,_s6v>0) +
                _mk_sig("✅" if _s7v>0 else "❌","S7 — Rank Stability ★","Top-50 overlap > 60% vs 4 weeks ago",f'Overlap: {_smeta.get("rank_overlap_pct",65):.0f}%',_s7v,0.5,_s7v>0) +
                '</div>'
            )

            _g_col, _s_col = st.columns([1, 1.8])
            with _g_col:
                _stc_rt.html(_rt_gauge_html, height=360)
                _scbg = {3:"#dcfce7",2:"#dbeafe",1:"#fef3c7",0:"#fee2e2"}[_rt_sc]
                _scfc = {3:"#15803d",2:"#1d4ed8",1:"#d97706",0:"#dc2626"}[_rt_sc]
                st.markdown(
                    f'<div style="background:{_scbg};border-radius:8px;padding:10px;text-align:center;margin-top:4px;">' +
                    f'<div style="font-size:11px;color:{_scfc};font-weight:700;text-transform:uppercase;">Weighted Score</div>' +
                    f'<div style="font-size:36px;font-weight:900;color:{_scfc};line-height:1.1;">{_rs.raw_score:.2f}</div>' +
                    f'<div style="font-size:11px;color:{_scfc};">/ 8.5 pts · {_rt_lbl}</div>' +
                    '</div>', unsafe_allow_html=True)
            with _s_col:
                _stc_rt.html(_rt_sig_html, height=580)

            st.markdown("---")

            # ── QFSM Allocation tiles ──────────────────────────────────────────
            _qc  = "#7c3aed" if _rs.qfsm_mode=="BLEND" else "#15803d"
            _qbg = "#ede9fe" if _rs.qfsm_mode=="BLEND" else "#dcfce7"
            _qlbl = (f"⚛ QFSM BLEND — Score {_rs.raw_score:.2f} in transition zone"
                     if _rs.qfsm_mode=="BLEND" else f"✅ Standard Band {_rs.effective_band} allocation")
            st.markdown(f'<div style="background:{_qbg};border:1px solid {_qc};border-radius:6px;padding:6px 14px;font-size:12px;color:{_qc};font-weight:700;margin-bottom:8px;">{_qlbl}</div>', unsafe_allow_html=True)

            _rta1, _rta2, _rta3 = st.columns(3)
            for _rtcol, (lbl, pct, fc, bg) in zip(
                [_rta1, _rta2, _rta3],
                [("📈 Equity",_rs.equity,"#1d4ed8","#dbeafe"),
                 ("🥇 GOLDBEES",_rs.gold,"#b45309","#fef3c7"),
                 ("💵 Liquid",_rs.cash,"#374151","#f1f5f9")]
            ):
                with _rtcol:
                    st.markdown(
                        f'<div style="background:{bg};border:1px solid {fc};border-radius:8px;padding:14px;text-align:center;">' +
                        f'<div style="font-size:11px;color:{fc};margin-bottom:6px">{lbl}</div>' +
                        f'<div style="font-size:30px;font-weight:800;color:{fc}">{pct*100:.1f}%</div>' +
                        '</div>', unsafe_allow_html=True)

            # ── VIX overlay display ────────────────────────────────────────────
            if _rt_vx is not None and _rs.vix_overlay_pct > 0:
                _vc2 = "#dc2626" if _rt_vx > 30 else "#d97706"
                st.markdown(
                    f'<div style="background:{"#fef2f2" if _rt_vx>30 else "#fef3c7"};border:1.5px solid {_vc2};border-left:4px solid {_vc2};border-radius:8px;padding:10px 14px;font-size:12px;color:{_vc2};margin-top:8px;">' +
                    f'⚡ <b>VIX Overlay Active (VIX {_rt_vx:.1f}):</b> +{_rs.vix_overlay_pct:.1f}pp Gold (Liquid → Gold) | Equity UNTOUCHED' +
                    '</div>', unsafe_allow_html=True)
            elif _rt_vx is not None:
                st.caption(f"VIX {_rt_vx:.1f} ≤ 20 — No overlay. Base allocation applies.")

            # ── Score history ──────────────────────────────────────────────────
            if _rs.history:
                with st.expander("📅 Score History (last 8 weeks)", expanded=False):
                    _hdf = pd.DataFrame(_rs.history)
                    _hdf.columns = [c.replace("_"," ").title() for c in _hdf.columns]
                    st.dataframe(_hdf, use_container_width=True, hide_index=True)

            st.caption(
                f"Score {_rs.raw_score:.2f}/8.5 · {_rt_lbl} · Conf: {_rs.status} · "
                f"QFSM: {_rs.qfsm_mode} · {_rt_src} · {_rt_date_s} | "
                f"Alloc: {_rs.equity*100:.1f}/{_rs.gold*100:.1f}/{_rs.cash*100:.1f} (Eq/Gold/Cash)"
            )

            # ── Last Rebalance Memory display ──────────────────────────────────
            _rb_mem = st.session_state.get("_rb_memory", {})
            if _rb_mem:
                with st.expander("💾 Last Rebalance Memory", expanded=False):
                    _rb_alloc = _rb_mem.get("allocation", {})
                    _rb_txt = (
                        f"**Date:** {_rb_mem.get('last_rebalance_date','—')} | "
                        f"**Band:** {_rb_mem.get('regime_band','—')} | "
                        f"**Score:** {_rb_mem.get('weighted_score','—')} | "
                        f"**QFSM:** {_rb_mem.get('qfsm_mode','—')}\n\n"
                        f"**Alloc:** Eq {_rb_alloc.get('equity_pct','—')}% / "
                        f"Gold {_rb_alloc.get('gold_pct','—')}% / "
                        f"Cash {_rb_alloc.get('cash_pct','—')}%\n\n"
                        f"**Notes:** {_rb_mem.get('notes','—')}"
                    )
                    st.markdown(_rb_txt)


# MULTI-ASSET SIMULATOR TAB — full HTML widget (matches rebalancepanel)
# ════════════════════════════════════════════════════════════════════
with _tab_sim:
    import streamlit.components.v1 as _sim_stc

    _SIM_HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<style>
:root{
  --bg:#f8fafc;--surface:#fff;--surface2:#f1f5f9;--border:#e2e8f0;
  --text:#1e293b;--text2:#475569;--text3:#94a3b8;
  --navy:#1a237e;--navy2:#283593;
  --buy:#2e7d32;--buy-bg:#e8f5e9;--buy-border:#a5d6a7;
  --sell:#c62828;--sell-bg:#ffebee;--sell-border:#ef9a9a;
  --blue:#1565c0;--blue-bg:#e3f2fd;--blue-border:#90caf9;
  --orange:#e65100;--orange-bg:#fff3e0;--orange-border:#ffb74d;
  --teal:#00695c;--teal-bg:#e0f2f1;--teal-border:#80cbc4;
  --amber:#ff8f00;--amber-bg:#fffde7;--amber-border:#ffe082;
  --purple:#6a1b9a;--purple-bg:#f3e5f5;--purple-border:#ce93d8;
  --shadow:0 2px 8px rgba(0,0,0,.07);--radius:8px;
}
*{box-sizing:border-box;margin:0;padding:0;}
body{font-family:'Segoe UI',system-ui,sans-serif;font-size:12.5px;background:var(--bg);color:var(--text);}
::-webkit-scrollbar{width:4px;}::-webkit-scrollbar-thumb{background:var(--border);border-radius:2px;}

/* INPUT GRID */
.inp-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:0;border-bottom:2px solid var(--border);background:var(--surface);}
.inp-col{padding:8px 10px;border-right:1px solid var(--border);}
.inp-col:last-child{border-right:none;}
.inp-col-hdr{font-size:9px;font-weight:800;color:var(--navy);text-transform:uppercase;letter-spacing:.6px;border-bottom:1px solid var(--border);padding-bottom:4px;margin-bottom:7px;}
.inp-field{display:flex;flex-direction:column;gap:2px;margin-bottom:5px;}
.inp-lbl{font-size:10px;color:var(--text2);font-weight:600;}
.inp-ctrl{width:100%;padding:4px 7px;border:1.5px solid var(--border);border-radius:5px;font-size:12px;font-weight:600;background:var(--surface);color:var(--text);outline:none;transition:border-color .15s;}
.inp-ctrl[type=number]{text-align:right;}
.inp-ctrl:focus{border-color:var(--navy);}
select.inp-ctrl{text-align:left;}

/* RESULTS */
.results{padding:8px;}
.sim-sec{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius);overflow:hidden;margin-bottom:8px;}
.sim-sec-hdr{padding:7px 12px;font-size:10px;font-weight:800;text-transform:uppercase;letter-spacing:.5px;border-bottom:1px solid var(--border);}
.sim-sec-hdr.navy{background:#e8eaf6;color:var(--navy);}
.sim-sec-hdr.teal{background:var(--teal-bg);color:var(--teal);}
.sim-sec-hdr.green{background:var(--buy-bg);color:var(--buy);}
.sim-sec-hdr.sell{background:var(--sell-bg);color:var(--sell);}
.sim-sec-hdr.amber{background:var(--amber-bg);color:var(--orange);}
.sim-sec-hdr.purple{background:var(--purple-bg);color:var(--purple);}
.sim-sec-body{padding:8px 12px;}

/* SCORE GAUGE */
.score-gauge{text-align:center;padding:10px 12px;border-radius:8px;border:2px solid var(--border);}
.score-gauge.s3{background:#e8f5e9;border-color:#a5d6a7;}
.score-gauge.s2{background:#e3f2fd;border-color:#90caf9;}
.score-gauge.s1{background:#fffde7;border-color:#ffe082;}
.score-gauge.s0{background:#ffebee;border-color:#ef9a9a;}
.score-num{font-size:48px;font-weight:900;line-height:1;}
.score-gauge.s3 .score-num,.score-gauge.s3 .score-lbl{color:#2e7d32;}
.score-gauge.s2 .score-num,.score-gauge.s2 .score-lbl{color:#1565c0;}
.score-gauge.s1 .score-num,.score-gauge.s1 .score-lbl{color:#e65100;}
.score-gauge.s0 .score-num,.score-gauge.s0 .score-lbl{color:#c62828;}
.score-lbl{font-size:13px;font-weight:700;margin-top:3px;}

/* ALLOC CARDS */
.alloc-strip{display:flex;gap:8px;}
.alloc-card{flex:1;border-radius:8px;padding:9px;text-align:center;border:1.5px solid var(--border);}
.alloc-card.ac-eq{background:#e8f5e9;border-color:#a5d6a7;}
.alloc-card.ac-gold{background:#fffde7;border-color:#ffe082;}
.alloc-card.ac-cash{background:#e3f2fd;border-color:#90caf9;}
.ac-label{font-size:9px;font-weight:700;text-transform:uppercase;letter-spacing:.5px;color:var(--text2);margin-bottom:2px;}
.ac-pct{font-size:22px;font-weight:900;}
.alloc-card.ac-eq .ac-pct{color:#2e7d32;}
.alloc-card.ac-gold .ac-pct{color:#e65100;}
.alloc-card.ac-cash .ac-pct{color:#1565c0;}
.ac-rs{font-size:10.5px;font-weight:700;color:var(--text2);margin-top:2px;}

/* THIS WEEK CARD */
.tw-card{border-radius:8px;padding:10px 13px;border:2px solid;}
.tw-card.tw-normal{background:#e0f7fa;border-color:#00838f;}
.tw-card.tw-pause{background:#fff8e1;border-color:#f9a825;}
.tw-card.tw-emerg{background:#fbe9e7;border-color:#d84315;}
.tw-card.tw-complete{background:var(--buy-bg);border-color:var(--buy-border);}
.tw-hdr{font-size:11px;font-weight:800;text-transform:uppercase;letter-spacing:.4px;margin-bottom:7px;}
.tw-card.tw-normal .tw-hdr{color:#00838f;}
.tw-card.tw-pause  .tw-hdr{color:#f9a825;}
.tw-card.tw-emerg  .tw-hdr{color:#d84315;}
.tw-card.tw-complete .tw-hdr{color:var(--buy);}
.tw-row{display:flex;align-items:center;gap:8px;padding:4px 0;border-bottom:1px dashed rgba(0,0,0,.1);font-size:12px;}
.tw-row:last-child{border-bottom:none;}
.tw-lbl{flex:1;font-weight:600;color:var(--text);}
.tw-val{font-size:12.5px;font-weight:800;}
.tw-val.buy{color:var(--buy);}  .tw-val.sell{color:var(--sell);}
.tw-val.hold{color:var(--text2);}  .tw-val.navy{color:var(--navy);}

/* SHIFT TABLE */
.shift-tbl{width:100%;border-collapse:collapse;font-size:11px;}
.shift-tbl th{background:#e8eaf6;padding:5px 7px;color:var(--navy);font-weight:700;text-align:center;font-size:10px;text-transform:uppercase;border-bottom:1.5px solid #c5cae9;}
.shift-tbl td{padding:5px 7px;border-bottom:1px solid var(--border);text-align:center;}
.shift-tbl tr.cur-week td{background:var(--blue-bg);font-weight:700;}
.shift-tbl tr.done td{opacity:.5;text-decoration:line-through;}
.shift-tbl tr.target-row td{background:var(--buy-bg);font-weight:700;}

/* GAP TABLE */
.gap-tbl{width:100%;border-collapse:collapse;font-size:11.5px;}
.gap-tbl th{background:var(--surface2);padding:5px 8px;text-align:left;font-weight:700;color:var(--text2);font-size:10px;text-transform:uppercase;border-bottom:1.5px solid var(--border);}
.gap-tbl td{padding:5px 8px;border-bottom:1px solid var(--border);}
.gap-tbl tr:hover td{background:var(--surface2);}
.cell-buy{color:var(--buy);font-weight:700;} .cell-sell{color:var(--sell);font-weight:700;}
.cell-hold{color:var(--text3);font-style:italic;}

/* PROCEEDS FLOW */
.pf-flow{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius);}
.pf-step{display:flex;align-items:flex-start;gap:10px;padding:7px 12px;border-bottom:1px solid var(--border);font-size:11.5px;}
.pf-step:last-child{border-bottom:none;}
.pf-num{background:var(--navy);color:#fff;border-radius:50%;width:20px;height:20px;display:flex;align-items:center;justify-content:center;font-size:10px;font-weight:800;flex-shrink:0;margin-top:1px;}
.pf-body{flex:1;}
.pf-lbl{font-weight:700;color:var(--text);}
.pf-det{color:var(--text2);font-size:11px;margin-top:1px;line-height:1.5;}
.pf-amt{font-weight:800;font-size:12.5px;white-space:nowrap;}
.pf-amt.pos{color:var(--buy);} .pf-amt.neg{color:var(--sell);} .pf-amt.na{color:var(--text3);}

/* WHATS-IF TABLE */
.wi-tbl{width:100%;border-collapse:collapse;font-size:11px;}
.wi-tbl th{background:#e8eaf6;padding:5px 7px;color:var(--navy);font-weight:700;text-align:center;font-size:10px;text-transform:uppercase;border-bottom:1.5px solid #c5cae9;}
.wi-tbl td{padding:5px 8px;border-bottom:1px solid var(--border);text-align:center;}
.wi-tbl tr.cur td{background:var(--buy-bg);font-weight:700;}

/* BOXES */
.info-box{background:var(--blue-bg);padding:8px 12px;font-size:11.5px;color:var(--blue);border-radius:6px;margin:6px 0;border-left:3px solid var(--blue-border);}
.warn-box{background:var(--sell-bg);padding:8px 12px;font-size:11.5px;color:var(--sell);border-radius:6px;margin:6px 0;border-left:3px solid var(--sell-border);}
.success-box{background:var(--buy-bg);padding:8px 12px;font-size:11.5px;color:var(--buy);border-radius:6px;margin:6px 0;border-left:3px solid var(--buy-border);}
.suggest-box{background:var(--orange-bg);padding:8px 12px;font-size:11.5px;color:var(--orange);border-radius:6px;margin:6px 0;border-left:3px solid var(--orange-border);}
.dd-box{border-radius:8px;padding:8px 12px;border:1.5px solid;margin:6px 0;}
.dd-box.dd-ok{background:var(--buy-bg);border-color:var(--buy-border);}
.dd-box.dd-watch{background:var(--amber-bg);border-color:var(--amber-border);}
.dd-box.dd-override{background:var(--sell-bg);border-color:var(--sell-border);}
.dd-box.dd-emerg{background:#2a0000;border-color:#ef9a9a;}
.dd-box-title{font-size:9px;font-weight:700;text-transform:uppercase;letter-spacing:.5px;}
.dd-box.dd-ok .dd-box-title{color:var(--buy);}
.dd-box.dd-watch .dd-box-title{color:var(--orange);}
.dd-box.dd-override .dd-box-title,.dd-box.dd-emerg .dd-box-title{color:var(--sell);}
.dd-pct{font-size:20px;font-weight:900;margin-top:2px;}
.dd-box.dd-ok .dd-pct{color:var(--buy);}
.dd-box.dd-watch .dd-pct{color:var(--orange);}
.dd-box.dd-override .dd-pct,.dd-box.dd-emerg .dd-pct{color:var(--sell);}
.dd-detail{font-size:11px;color:var(--text2);margin-top:2px;line-height:1.5;}

/* BADGE */
.bdg{display:inline-flex;align-items:center;padding:2px 8px;border-radius:20px;font-size:10.5px;font-weight:700;}
.bdg-buy{background:var(--buy-bg);color:var(--buy);border:1px solid var(--buy-border);}
.bdg-sell{background:var(--sell-bg);color:var(--sell);border:1px solid var(--sell-border);}
.bdg-navy{background:#e8eaf6;color:var(--navy);border:1px solid #c5cae9;}
.empty{padding:24px;text-align:center;color:var(--text3);font-style:italic;}

.sum-strip{display:flex;gap:8px;flex-wrap:wrap;padding:6px 12px;background:var(--surface2);border-top:1px solid var(--border);}
.sum-item{display:flex;flex-direction:column;}
.sum-lbl{font-size:9px;color:var(--text3);text-transform:uppercase;letter-spacing:.3px;}
.sum-val{font-size:12.5px;font-weight:700;}
.sum-val.g{color:var(--buy);} .sum-val.r{color:var(--sell);}
.sum-val.b{color:var(--blue);} .sum-val.o{color:var(--orange);}
</style>
</head>
<body>

<!-- INPUT GRID -->
<div style="background:linear-gradient(90deg,#1a237e,#283593);color:#fff;padding:8px 14px;display:flex;align-items:center;justify-content:space-between;">
  <div>
    <div style="font-size:13px;font-weight:800;">🧪 Multi-Asset Regime Simulator</div>
    <div style="font-size:10px;opacity:.65;">Koi bhi scenario simulate karo — pura deployment plan real-time milega</div>
  </div>
  <button onclick="simReset()" style="background:rgba(255,255,255,.15);color:#fff;border:1.5px solid rgba(255,255,255,.35);border-radius:20px;padding:5px 14px;font-size:12px;font-weight:700;cursor:pointer;">🔄 Reset</button>
</div>

<div class="inp-grid">
  <!-- PORTFOLIO -->
  <div class="inp-col">
    <div class="inp-col-hdr">💼 Portfolio</div>
    <div class="inp-field"><span class="inp-lbl">Total PF ₹</span><input type="number" id="sim-total" value="2000000" class="inp-ctrl" oninput="runSim()"></div>
    <div class="inp-field"><span class="inp-lbl">ATH PF ₹</span><input type="number" id="sim-ath" value="2000000" class="inp-ctrl" oninput="runSim()"></div>
    <div class="inp-field"><span class="inp-lbl">Exit Proceeds this Month ₹</span><input type="number" id="sim-exits" value="200000" class="inp-ctrl" oninput="runSim()"></div>
    <div class="inp-field"><span class="inp-lbl">Capital Addition ₹</span><input type="number" id="sim-capadd" value="0" class="inp-ctrl" oninput="runSim()"></div>
  </div>
  <!-- REGIME -->
  <div class="inp-col">
    <div class="inp-col-hdr">🌡️ Regime</div>
    <div class="inp-field"><span class="inp-lbl">Previous Month Score</span>
      <select id="sim-prev" class="inp-ctrl" onchange="runSim()">
        <option value="-1">— First time —</option>
        <option value="0">0 — 🐻 Bear</option>
        <option value="1">1 — ⚖️ Neutral</option>
        <option value="2" selected>2 — 📈 Mild Bull</option>
        <option value="3">3 — 🐂 Strong Bull</option>
      </select>
    </div>
    <div class="inp-field"><span class="inp-lbl">Current Month Score</span>
      <select id="sim-cur" class="inp-ctrl" onchange="runSim()">
        <option value="0">0 — 🐻 Bear</option>
        <option value="1">1 — ⚖️ Neutral</option>
        <option value="2">2 — 📈 Mild Bull</option>
        <option value="3" selected>3 — 🐂 Strong Bull</option>
      </select>
    </div>
    <div class="inp-field"><span class="inp-lbl">India VIX</span><input type="number" id="sim-vix" value="16" step="0.5" class="inp-ctrl" oninput="runSim()"></div>
  </div>
  <!-- ASSETS -->
  <div class="inp-col">
    <div class="inp-col-hdr">🏦 Current Assets</div>
    <div class="inp-field"><span class="inp-lbl">Equity ₹</span><input type="number" id="sim-eq-cur" value="1300000" class="inp-ctrl" oninput="runSim()"></div>
    <div class="inp-field"><span class="inp-lbl">GOLDBEES ₹</span><input type="number" id="sim-gold-cur" value="400000" class="inp-ctrl" oninput="runSim()"></div>
    <div class="inp-field"><span class="inp-lbl">Liquid Fund ₹</span><input type="number" id="sim-liq-cur" value="300000" class="inp-ctrl" oninput="runSim()"></div>
    <div class="inp-field"><span class="inp-lbl">GOLDBEES CMP ₹/unit</span><input type="number" id="sim-gold-cmp" value="125" step="0.5" class="inp-ctrl" oninput="runSim()"></div>
  </div>
  <!-- WEEK -->
  <div class="inp-col">
    <div class="inp-col-hdr">📆 Deployment Week</div>
    <div class="inp-field"><span class="inp-lbl">Deploying which week?</span>
      <select id="sim-week" class="inp-ctrl" onchange="runSim()">
        <option value="0">Month start / Week 1</option>
        <option value="1">✅ Week 1 done → Now Week 2</option>
        <option value="2">✅ Weeks 1–2 done → Now Week 3</option>
        <option value="3">✅ Weeks 1–3 done → Now Week 4</option>
      </select>
    </div>
    <div class="inp-field"><span class="inp-lbl">Weekly NAV Return % (this Friday)</span><input type="number" id="sim-nav-ret" value="0" step="0.1" class="inp-ctrl" oninput="runSim()"></div>
  </div>
</div>

<!-- RESULTS -->
<div class="results" id="simResults">
  <div class="empty">Values enter karo — results yahan automatically aayenge</div>
</div>

<script>
var ALLOC={3:[80,15,5],2:[65,20,15],1:[45,25,30],0:[25,30,45]};
var LBLS={3:'🐂 Strong Bull',2:'📈 Mild Bull',1:'⚖️ Neutral',0:'🐻 Bear'};
var SCLS={3:'s3',2:'s2',1:'s1',0:'s0'};

function fmt(n){return isNaN(n)?'0':Math.round(n).toLocaleString('en-IN');}
function fmtD(n){return isNaN(n)?'0':parseFloat(n).toLocaleString('en-IN',{minimumFractionDigits:0,maximumFractionDigits:2});}
function gv(id){return parseFloat(document.getElementById(id).value)||0;}
function gi(id){return parseInt(document.getElementById(id).value);}

function runSim(){
  var tot=gv('sim-total'),ath=gv('sim-ath'),exits=gv('sim-exits'),capadd=gv('sim-capadd');
  var prevScore=gi('sim-prev'),curScore=gi('sim-cur');
  var vix=gv('sim-vix'),eqCur=gv('sim-eq-cur'),goldCur=gv('sim-gold-cur');
  var liqCur=gv('sim-liq-cur'),goldCmp=gv('sim-gold-cmp')||125;
  var weekNum=gi('sim-week'),navRet=gv('sim-nav-ret');

  if(tot<=0){document.getElementById('simResults').innerHTML='<div class="empty">Total PF ₹ enter karo</div>';return;}

  // ── DD Check ──
  var ddPct=0,ddCls='dd-ok',ddDetail='Normal.',ddOverride=false,ddEmerg=false,score=curScore;
  if(ath>0){
    ddPct=(tot/ath-1)*100;
    if(ddPct>=-15){ddCls='dd-ok';ddDetail='Normal — weekly check as usual.';}
    else if(ddPct>-20){ddCls='dd-watch';ddDetail='⚠️ DD '+Math.abs(ddPct).toFixed(1)+'% ≥ 15% — Pause capital adds. Weekly mandatory check.';}
    else if(ddPct>-30){ddCls='dd-override';ddOverride=true;score=0;ddDetail='🔴 DD Override: Score forced → 0 (Bear). 4-week defensive shift.';}
    else{ddCls='dd-emerg';ddOverride=true;ddEmerg=true;score=0;ddDetail='🚨 Emergency DD ≥ 30%: SINGLE WEEK → Eq 20% | Gold 30% | Cash 50%.';}
  }

  // ── VIX Overlay ──
  var base=ALLOC[score].slice();
  var vixAdj=vix>30?5:vix>20?3:0;
  var goldPct=ddEmerg?30:Math.min(30,base[1]+vixAdj);
  var eqPct=ddEmerg?20:base[0];
  var cashPct=ddEmerg?50:100-eqPct-goldPct;
  var eqBudget=tot*eqPct/100,goldTgt=tot*goldPct/100,liqTgt=tot*cashPct/100;
  var goldDiff=goldTgt-goldCur,liqDiff=liqTgt-liqCur;
  var isPause=(vix>30&&navRet<-5);
  var scoreChanged=(prevScore>=0&&prevScore!==score);
  var isRec=score>prevScore;
  var totalWeeks=scoreChanged?(Math.abs(score-prevScore)===1?2:Math.abs(score-prevScore)===2?3:4):0;
  var thisWeek=weekNum+1;

  var html='';

  // ════ 1. SCORE + DD ════
  html+='<div class="sim-sec"><div class="sim-sec-hdr navy">🌡️ Regime Score &amp; Allocation</div><div class="sim-sec-body">';
  html+='<div style="display:flex;gap:10px;flex-wrap:wrap;align-items:flex-start;">';
  html+='<div class="score-gauge '+SCLS[score]+'" style="min-width:120px;flex:0 0 auto;padding:10px 12px;">'+
    '<div style="font-size:9px;opacity:.6;font-weight:700;text-transform:uppercase;margin-bottom:2px;">SCORE</div>'+
    '<div class="score-num">'+score+(ddOverride?'*':'')+'</div>'+
    '<div class="score-lbl">'+LBLS[score]+'</div>'+
    (ddOverride?'<div style="font-size:10px;color:var(--sell);font-weight:700;margin-top:3px;">'+(ddEmerg?'Emergency 🚨':'DD Override 🔴')+'</div>':'')+
  '</div>';
  html+='<div style="flex:1;min-width:200px;">';
  html+='<div class="alloc-strip" style="margin-bottom:8px;">'+
    '<div class="alloc-card ac-eq"><div class="ac-label">📈 Equity</div><div class="ac-pct">'+eqPct+'%</div><div class="ac-rs">₹'+fmt(eqBudget)+'</div></div>'+
    '<div class="alloc-card ac-gold"><div class="ac-label">🥇 GOLDBEES</div><div class="ac-pct">'+goldPct+'%</div><div class="ac-rs">₹'+fmt(goldTgt)+'</div></div>'+
    '<div class="alloc-card ac-cash"><div class="ac-label">💵 Liquid</div><div class="ac-pct">'+cashPct+'%</div><div class="ac-rs">₹'+fmt(liqTgt)+'</div></div>'+
  '</div>';
  // DD box
  var ddBanner=ath>0?('<div class="dd-box '+ddCls+'"><div class="dd-box-title">📉 Drawdown from ATH ₹'+fmt(ath)+'</div>'+
    '<div class="dd-pct">'+Math.abs(ddPct).toFixed(1)+'%</div><div class="dd-detail">'+ddDetail+'</div></div>'):'';
  html+=ddBanner;
  // VIX overlay
  if(vixAdj>0){
    html+='<div class="'+(vix>30?'warn-box':'info-box')+'">⚡ <strong>VIX Overlay:</strong> VIX '+fmtD(vix)+' → Gold +'+vixAdj+'% (base '+base[1]+'%→'+goldPct+'%). Funded from Liquid. Equity NAHI badla.</div>';
  }
  if(scoreChanged){
    var sc_dir=isRec?'📈 Recovery':'📉 Defensive';
    var sc_msg=isRec?'Exit proceeds → Equity (Liquid se fund). '+totalWeeks+'-week plan.':'Natural exits → Liquid + Gold. '+totalWeeks+'-week plan.';
    html+='<div class="'+(isRec?'success-box':'warn-box')+'"><strong>'+sc_dir+': Score '+prevScore+' → '+score+'</strong><br><span style="font-size:11px;">'+sc_msg+'</span></div>';
  }
  html+='</div></div></div></div>';

  // ════ 2. CURRENT vs TARGET GAP ════
  html+='<div class="sim-sec"><div class="sim-sec-hdr teal">📊 Current vs Target — Asset Gap</div><div class="sim-sec-body">';
  html+='<table class="gap-tbl"><thead><tr>'+
    '<th>Asset</th><th>Current ₹</th><th>Current %</th>'+
    '<th>Target %</th><th>Target ₹</th><th>Gap ₹</th><th>Drift %</th><th>Action</th></tr></thead><tbody>';

  function assetRow(name,cur,tgt,tpct){
    var gap=tgt-cur,drift=tot>0?Math.abs(cur-tgt)/tot*100:0;
    var acCls=Math.abs(gap)<15000?'cell-hold':gap>0?'cell-buy':'cell-sell';
    var acTxt=Math.abs(gap)<15000?'HOLD &lt;₹15K':gap>0?'🟢 BUY / ADD':'🔴 SELL / REDEEM';
    return '<tr><td><strong>'+name+'</strong></td><td>₹'+fmt(cur)+'</td>'+
      '<td>'+((cur/tot)*100).toFixed(1)+'%</td>'+
      '<td>'+tpct+'%</td><td>₹'+fmt(tgt)+'</td>'+
      '<td class="'+(gap>=0?'cell-buy':'cell-sell')+'">'+fmt(gap)+'</td>'+
      '<td class="'+(drift>7?'cell-sell':drift>3?'cell-buy':'cell-hold')+'">'+drift.toFixed(1)+'%</td>'+
      '<td class="'+acCls+'">'+acTxt+'</td></tr>';
  }
  html+=assetRow('📈 Equity',eqCur,eqBudget,eqPct);
  html+=assetRow('🥇 GOLDBEES',goldCur,goldTgt,goldPct);
  html+=assetRow('💵 Liquid',liqCur,liqTgt,cashPct);
  html+='</tbody></table>';
  if(Math.abs(goldDiff)>=15000&&goldCmp>0){
    var units=Math.floor(Math.abs(goldDiff)/goldCmp);
    html+='<div style="margin-top:6px;font-size:11.5px;color:var(--text2);">🥇 ≈ <strong>'+units+' GOLDBEES units</strong> @ ₹'+fmtD(goldCmp)+'/unit</div>';
  }
  html+='<div class="sum-strip">'+
    '<div class="sum-item"><span class="sum-lbl">Eq Budget</span><span class="sum-val b">₹'+fmt(eqBudget)+'</span></div>'+
    '<div class="sum-item"><span class="sum-lbl">Per Stock (÷30)</span><span class="sum-val g">₹'+fmt(eqBudget/30)+'</span></div>'+
    '<div class="sum-item"><span class="sum-lbl">Gold Gap</span><span class="sum-val '+(goldDiff>0?'g':'r')+'">'+fmt(goldDiff)+'</span></div>'+
    '<div class="sum-item"><span class="sum-lbl">Liquid Gap</span><span class="sum-val '+(liqDiff>0?'g':'r')+'">'+fmt(liqDiff)+'</span></div>'+
  '</div></div></div>';

  // ════ 3. THIS WEEK'S ACTIONS ════
  html+=buildThisWeekSection(prevScore,score,tot,weekNum,navRet,vix,isPause,goldCur,goldCmp,liqCur,eqPct,goldPct,cashPct,eqBudget,ddOverride,ddEmerg,thisWeek,totalWeeks,isRec);

  // ════ 4. FULL SHIFT TABLE ════
  if(scoreChanged&&totalWeeks>0){
    html+=buildShiftSection(prevScore,score,tot,weekNum,navRet,vix,isPause,totalWeeks,isRec);
  }

  // ════ 5. PROCEEDS ALLOCATION ════
  html+=buildProceedsSection(prevScore,score,tot,exits,capadd,eqPct,goldPct,cashPct,goldCur,liqCur,eqBudget,goldDiff,liqDiff,scoreChanged,isRec);

  // ════ 6. WHAT-IF NEXT MONTH ════
  html+=buildWhatIfSection(score,tot,prevScore);

  // ════ 7. GUARDRAILS ════
  html+='<div class="sim-sec"><div class="sim-sec-hdr purple">📐 SOP v2026 Quick Reference</div><div class="sim-sec-body">';
  html+='<div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;font-size:11.5px;line-height:1.8;">';
  html+='<div><div style="font-weight:700;color:var(--navy);margin-bottom:4px;">Allocation Table</div>'+
    '<table style="width:100%;border-collapse:collapse;font-size:11px;">'+
    '<tr style="background:#e8eaf6;"><td style="padding:3px 6px;font-weight:700;color:var(--navy);">Score</td><td style="padding:3px 6px;font-weight:700;color:var(--buy);">Eq%</td><td style="padding:3px 6px;font-weight:700;color:var(--orange);">Gold%</td><td style="padding:3px 6px;font-weight:700;color:var(--blue);">Cash%</td></tr>'+
    '<tr style="'+(score===3?'background:var(--buy-bg);font-weight:700;':'')+'"><td style="padding:3px 6px;">3 🐂 Strong Bull</td><td>80%</td><td>15%</td><td>5%</td></tr>'+
    '<tr style="'+(score===2?'background:var(--blue-bg);font-weight:700;':'')+'"><td style="padding:3px 6px;">2 📈 Mild Bull</td><td>65%</td><td>20%</td><td>15%</td></tr>'+
    '<tr style="'+(score===1?'background:var(--amber-bg);font-weight:700;':'')+'"><td style="padding:3px 6px;">1 ⚖️ Neutral</td><td>45%</td><td>25%</td><td>30%</td></tr>'+
    '<tr style="'+(score===0?'background:var(--sell-bg);font-weight:700;':'')+'"><td style="padding:3px 6px;">0 🐻 Bear</td><td>25%</td><td>30%</td><td>45%</td></tr>'+
    '</table></div>';
  html+='<div><div style="font-weight:700;color:var(--navy);margin-bottom:4px;">Guardrails</div>'+
    '• Min ₹15K per Gold/Liquid transaction<br>'+
    '• Gold drift band ±7% of total PF<br>'+
    '• Equity drift: ±₹20,000 per stock<br>'+
    '• Standalone equity sells: <strong>KABHI NAHI</strong><br>'+
    '• DD ≥20% → Score forced 0 (Bear)<br>'+
    '• DD ≥30% → Emergency single-week move<br>'+
    '• VIX ≤20: No overlay | 20-30: +3% Gold | >30: +5% Gold<br>'+
    '• Pause week: VIX>30 <strong>AND</strong> NAV &lt; -5% both<br>'+
    '• Score ±1 → 2-week plan | ±2 → 3-week | ±3 → 4-week'+
    '</div></div></div></div>';

  document.getElementById('simResults').innerHTML=html;
}

function buildThisWeekSection(prev,score,tot,weekNum,navRet,vix,isPause,goldCur,goldCmp,liqCur,eqPct,goldPct,cashPct,eqBudget,ddOverride,ddEmerg,thisWeek,totalWeeks,isRec){
  var goldTgt=tot*goldPct/100,liqTgt=tot*cashPct/100;
  var goldDiff=goldTgt-goldCur,liqDiff=liqTgt-liqCur;
  var scoreChanged=(prev>=0&&prev!==score);

  var cls='tw-normal',hdrTxt='',rows='';

  if(ddEmerg){
    cls='tw-emerg'; hdrTxt='🚨 EMERGENCY — DD≥30% — SINGLE WEEK MOVE';
    rows=tw('Equity → target 20%','₹'+fmt(tot*0.20),'sell')+
      tw('GOLDBEES → target 30%','₹'+fmt(tot*0.30),'hold')+
      tw('Liquid → target 50%','₹'+fmt(tot*0.50),'buy')+
      tw('⚠️ No phasing — execute ALL in one week','Sabse urgent action','sell');
  } else if(isPause&&scoreChanged){
    cls='tw-pause'; hdrTxt='⏸ WEEK '+thisWeek+' PAUSED — VIX>30 AND Weekly NAV<−5%';
    rows=tw('VIX',fmtD(vix),'sell')+tw('Weekly NAV Return',navRet.toFixed(1)+'%','sell')+
      tw('This week deployment','HOLD — Carry forward to next Friday','hold')+
      tw('Reason','BOTH conditions met: VIX>30 AND NAV<−5%','hold');
  } else if(!scoreChanged||prev<0){
    cls='tw-normal'; hdrTxt='📅 Score Same ('+score+'→'+score+') — Normal Monthly Rebalance';
    var gA=Math.abs(goldDiff)<15000?'HOLD (<₹15K)':goldDiff>0?'BUY ₹'+fmt(goldDiff)+' (≈'+Math.floor(goldDiff/(gv("sim-gold-cmp")||125))+' units)':'SELL ₹'+fmt(Math.abs(goldDiff));
    var lA=Math.abs(liqDiff)<15000?'HOLD (<₹15K)':liqDiff>0?'ADD ₹'+fmt(liqDiff):'REDEEM ₹'+fmt(Math.abs(liqDiff));
    rows=tw('1️⃣ Equity — sell exits, buy new stocks at target','Budget ₹'+fmt(eqBudget)+' | Per stock ₹'+fmt(eqBudget/30),'navy')+
      tw('2️⃣ GOLDBEES (drift ±7% check)',gA,(Math.abs(goldDiff)<15000?'hold':goldDiff>0?'buy':'sell'))+
      tw('3️⃣ Liquid Fund',lA,(Math.abs(liqDiff)<15000?'hold':liqDiff>0?'buy':'sell'))+
      tw('4️⃣ Drift check — any stock ±₹20K from target?','Surplus bachhe to 1–2 correct karo. No surplus → skip.','hold');
  } else if(thisWeek>totalWeeks){
    cls='tw-complete'; hdrTxt='✅ Shift Plan Complete — All '+totalWeeks+' weeks done!';
    rows=tw('Status','Target allocation reached ✅','buy')+
      tw('Next Friday','Fresh score check → Normal monthly mode','hold');
  } else {
    var fa=ALLOC[prev],ta=ALLOC[score];
    var p=thisWeek/totalWeeks,pP=(thisWeek-1)/totalWeeks;
    var eq=Math.round(fa[0]+(ta[0]-fa[0])*p),gold=Math.round(fa[1]+(ta[1]-fa[1])*p),cash=100-eq-gold;
    var pEq=thisWeek===1?fa[0]:Math.round(fa[0]+(ta[0]-fa[0])*pP);
    var pGold=thisWeek===1?fa[1]:Math.round(fa[1]+(ta[1]-fa[1])*pP);
    var pCash=100-pEq-pGold;
    var eqD=Math.round((eq-pEq)*tot/100),goldD=Math.round((gold-pGold)*tot/100),cashD=Math.round((cash-pCash)*tot/100);
    var isFin=thisWeek===totalWeeks;
    cls=isRec?'tw-normal':'tw-pause';
    hdrTxt=(isRec?'📈':'📉')+' Week '+thisWeek+' of '+totalWeeks+(isFin?' — FINAL 🎯':'')+' | Score '+prev+'→'+score;
    rows=tw('🎯 Target this Friday','Eq '+eq+'% | Gold '+gold+'% | Cash '+cash+'%','navy')+
      tw('📈 Equity ₹','₹'+fmt(tot*eq/100),'navy')+
      tw('🥇 GOLDBEES ₹','₹'+fmt(tot*gold/100)+(goldD!==0?' ('+(goldD>0?'+':'')+fmt(goldD)+')'  :''),goldD>0?'buy':goldD<0?'sell':'hold')+
      tw('💵 Liquid ₹','₹'+fmt(tot*cash/100)+(cashD!==0?' ('+(cashD>0?'+':'')+fmt(cashD)+')':''),cashD>0?'buy':cashD<0?'sell':'hold');
    if(isRec){
      rows+=tw('💵 Redeem from Liquid','₹'+fmt(Math.abs(cashD)),'sell')+
        tw('📈 Buy Equity (new entries)','₹'+fmt(Math.abs(eqD))+' | Per stock: ₹'+fmt(tot*eq/100/30),'buy');
      if(goldD<0) rows+=tw('🥇 Trim Gold → Liquid','₹'+fmt(Math.abs(goldD)),'sell');
    } else {
      rows+=tw('🔴 Sell Equity → Liquid','₹'+fmt(Math.abs(eqD)),'sell');
      if(goldD>0) rows+=tw('🥇 Buy GOLDBEES','₹'+fmt(goldD),'buy');
      rows+=tw('💵 Liquid target','₹'+fmt(tot*cash/100),'buy');
    }
    rows+=tw('⚠️ Next Friday: Score re-check',isFin?'Target reached ✅':'If same → Week '+(thisWeek+1)+'. Changed? → Fresh plan.','hold');
  }

  return '<div class="sim-sec"><div class="sim-sec-hdr '+(cls==='tw-emerg'?'sell':cls==='tw-pause'?'amber':'green')+'">'+
    '⚡ Week '+thisWeek+' — This Week\'s Actions</div>'+
    '<div class="sim-sec-body"><div class="tw-card '+cls+'">'+
    '<div class="tw-hdr">'+hdrTxt+'</div>'+rows+
    '</div></div></div>';
}

function tw(lbl,val,cls){
  return '<div class="tw-row"><span class="tw-lbl">'+lbl+'</span><span class="tw-val '+cls+'">'+val+'</span></div>';
}

function buildShiftSection(prev,score,tot,weekNum,navRet,vix,isPause,totalWeeks,isRec){
  var fa=ALLOC[prev],ta=ALLOC[score];
  var h='<div class="sim-sec"><div class="sim-sec-hdr '+(isRec?'green':'sell')+'">📆 Full '+(isRec?'Recovery':'Defensive')+' Shift Plan — Score '+prev+' → '+score+' ('+totalWeeks+'-week plan)</div><div class="sim-sec-body">';
  h+='<div class="'+(isRec?'success-box':'warn-box')+'" style="margin-bottom:8px;">'+
    '<strong>'+(isRec?'📈 Recovery: ':'📉 Defensive: ')+totalWeeks+'-Friday plan.</strong> '+
    (Math.abs(score-prev)>=2?'Score Δ≥2: 2-month phased. Month 2 pe fresh check — changed? Restart plan. ':'Natural exits fund shift. No standalone sells. ')+
    'Pause condition: VIX>30 <strong>AND</strong> Weekly NAV&lt;−5% (both zaruri).</div>';

  h+='<table class="shift-tbl"><thead><tr>'+
    '<th>Week</th><th>Eq%</th><th>Gold%</th><th>Cash%</th>'+
    '<th>Eq ₹</th><th>Gold ₹</th><th>Cash ₹</th>'+
    '<th>Per Stock ₹</th><th>Action This Friday</th></tr></thead><tbody>';

  // Baseline row
  h+='<tr><td style="color:var(--text2);font-weight:700;">Start</td>'+
    '<td>'+fa[0]+'%</td><td>'+fa[1]+'%</td><td>'+fa[2]+'%</td>'+
    '<td style="color:var(--buy)">₹'+fmt(tot*fa[0]/100)+'</td>'+
    '<td style="color:var(--orange)">₹'+fmt(tot*fa[1]/100)+'</td>'+
    '<td style="color:var(--blue)">₹'+fmt(tot*fa[2]/100)+'</td>'+
    '<td style="color:var(--buy)">₹'+fmt(tot*fa[0]/100/30)+'</td>'+
    '<td style="font-size:10px;color:var(--text2);">Baseline (current)</td></tr>';

  for(var w=1;w<=totalWeeks;w++){
    var p=w/totalWeeks,pP=(w-1)/totalWeeks;
    var eq=Math.round(fa[0]+(ta[0]-fa[0])*p),gold=Math.round(fa[1]+(ta[1]-fa[1])*p),cash=100-eq-gold;
    var pEq=w===1?fa[0]:Math.round(fa[0]+(ta[0]-fa[0])*pP);
    var pGold=w===1?fa[1]:Math.round(fa[1]+(ta[1]-fa[1])*pP),pCash=100-pEq-pGold;
    var eqD=Math.round((eq-pEq)*tot/100),goldD=Math.round((gold-pGold)*tot/100),cashD=Math.round((cash-pCash)*tot/100);
    var done=w<=weekNum,isCur=w===weekNum+1,isFin=w===totalWeeks;
    var rowCls=done?'done':isCur?'cur-week':isFin?'target-row':'';
    var isWkPause=isCur&&isPause;
    var act=isWkPause?'<strong style="color:#f9a825;">⏸ PAUSE — VIX>30 AND NAV&lt;−5%</strong>':
      isRec?'Redeem ₹'+fmt(Math.abs(cashD))+' Liquid → ₹'+fmt(Math.abs(eqD))+' Equity ('+eq+'% Eq, '+gold+'% Gold, '+cash+'% Cash)'+(goldD<0?' + Trim Gold ₹'+fmt(Math.abs(goldD)):''):
            'Sell ₹'+fmt(Math.abs(eqD))+' Equity → Liquid'+(goldD>0?' + Buy Gold ₹'+fmt(goldD):'');
    h+='<tr class="'+rowCls+'">'+
      '<td><strong>Fri '+w+(done?' ✅':isCur?' 👈':isFin&&!done?' 🎯':'')+'</strong></td>'+
      '<td><strong>'+eq+'%</strong></td><td><strong>'+gold+'%</strong></td><td><strong>'+cash+'%</strong></td>'+
      '<td style="color:var(--buy)">₹'+fmt(tot*eq/100)+'</td>'+
      '<td style="color:var(--orange)">₹'+fmt(tot*gold/100)+'</td>'+
      '<td style="color:var(--blue)">₹'+fmt(tot*cash/100)+'</td>'+
      '<td style="color:var(--buy)">₹'+fmt(tot*eq/100/30)+'</td>'+
      '<td style="font-size:10.5px;">'+act+'</td></tr>';
  }
  h+='<tr class="target-row"><td><strong>✅ Target</strong></td>'+
    '<td>'+ta[0]+'%</td><td>'+ta[1]+'%</td><td>'+ta[2]+'%</td>'+
    '<td style="color:var(--buy)">₹'+fmt(tot*ta[0]/100)+'</td>'+
    '<td style="color:var(--orange)">₹'+fmt(tot*ta[1]/100)+'</td>'+
    '<td style="color:var(--blue)">₹'+fmt(tot*ta[2]/100)+'</td>'+
    '<td style="color:var(--buy)">₹'+fmt(tot*ta[0]/100/30)+'</td>'+
    '<td style="font-size:10px;color:var(--buy);font-weight:700;">Final allocation reached ✅</td></tr>';
  h+='</tbody></table>';
  h+='<div class="info-box" style="margin-top:6px;font-size:11px;">⚠️ Har Friday: Score re-check karo. Changed? → Plan restart. Same? → Next week ke liye plan continue karo.</div>';
  h+='</div></div>';
  return h;
}

function buildProceedsSection(prev,score,tot,exits,capadd,eqPct,goldPct,cashPct,goldCur,liqCur,eqBudget,goldDiff,liqDiff,scoreChanged,isRec){
  var totalProc=exits+capadd;
  var h='<div class="sim-sec"><div class="sim-sec-hdr green">💰 Proceeds Allocation — ₹'+fmt(totalProc)+' (exits ₹'+fmt(exits)+' + capital ₹'+fmt(capadd)+')</div><div class="sim-sec-body">';
  var perStock=eqBudget/30;
  var isNeutralBear=(score<=1);

  // Step-by-step flow
  h+='<div class="pf-flow">';
  var rem=totalProc,sn=0;
  function step(lbl,det,amt,cls){
    sn++;
    return '<div class="pf-step"><div class="pf-num">'+sn+'</div>'+
      '<div class="pf-body"><div class="pf-lbl">'+lbl+'</div><div class="pf-det">'+det+'</div></div>'+
      '<div class="pf-amt '+cls+'">'+amt+'</div></div>';
  }

  var liqFirst=0;
  if(isNeutralBear&&exits>0){
    liqFirst=Math.round(exits*0.25);
    rem-=liqFirst;
    h+=step('25% Exits → Liquid Reserve','Bear/Neutral: exit proceeds ka 25% pehle liquid reserve mein rakho. Capital add separate hai.','₹'+fmt(liqFirst),'pos');
  }

  var goldFund=0;
  if(scoreChanged&&goldDiff>15000){
    goldFund=Math.min(rem,goldDiff);rem-=goldFund;
    h+=step('🥇 Gold Gap Fund (1st PRIORITY)','Regime shift → Gold target badla. ALWAYS equity se pehle fund karo. Drift >7% OR shift.','₹'+fmt(goldFund),'pos');
  } else if(Math.abs(goldDiff)/tot*100>7){
    goldFund=Math.abs(goldDiff)>15000?Math.min(rem,Math.abs(goldDiff)):0;
    if(goldFund>0){rem-=goldFund;h+=step('🥇 Gold Drift >7% Correction','Monthly RB pe action needed.',goldDiff>0?'₹'+fmt(goldFund):'-₹'+fmt(goldFund),goldDiff>0?'pos':'neg');}
    else{h+=step('🥇 Gold','Drift >7% but diff &lt;₹15K — HOLD.','HOLD','na');}
  } else {
    h+=step('🥇 Gold','Drift ≤7% AND diff &lt;₹15K — HOLD. No action.','HOLD','na');
  }

  var liqFund=0;
  if(scoreChanged&&liqDiff>15000){
    liqFund=Math.min(rem,liqDiff);rem-=liqFund;
    h+=step('💵 Liquid Gap Fund','Regime shift → Liquid target badla. Remaining se fund.','₹'+fmt(liqFund),'pos');
  } else if(Math.abs(liqDiff)>15000){
    h+=step('💵 Liquid','Diff ₹'+fmt(Math.abs(liqDiff))+' (non-shift drift). Adjust if surplus bachhe.','If surplus','na');
  } else {
    h+=step('💵 Liquid','Diff &lt;₹15K — HOLD.','HOLD','na');
  }

  var eqFund=Math.max(0,rem);
  var nEntries=perStock>0?Math.floor(eqFund/perStock):0;
  h+=step('📈 New Equity Entries (target weight ÷30)','Remaining → new momentum stocks at new per-stock target. Weight = Equity Budget ÷ 30.','₹'+fmt(eqFund)+' → '+nEntries+' stocks','pos');

  var surplus=eqFund-nEntries*perStock;
  h+=step('🎯 Drift Band + Surplus','Koi existing stock ±₹20K drift + surplus? → 1–2 correct. No surplus? SKIP. Standalone sells KABHI NAHI.','₹'+fmt(surplus)>500?'₹'+fmt(surplus)+' surplus':'₹0','na');
  h+='</div>';

  // Actual breakdown table
  h+='<div style="margin-top:10px;background:var(--blue-bg);border:1px solid var(--blue-border);border-radius:6px;padding:8px 12px;">';
  h+='<div style="font-weight:700;color:var(--blue);margin-bottom:6px;font-size:12px;">📋 ₹'+fmt(totalProc)+' ka breakdown:</div>';
  if(liqFirst>0) h+='<div style="font-size:11.5px;margin-bottom:2px;">→ Liquid reserve (25% of exits): <strong style="color:var(--blue)">₹'+fmt(liqFirst)+'</strong></div>';
  if(goldFund>0) h+='<div style="font-size:11.5px;margin-bottom:2px;">→ Gold gap: <strong style="color:var(--orange)">₹'+fmt(goldFund)+'</strong></div>';
  if(liqFund>0)  h+='<div style="font-size:11.5px;margin-bottom:2px;">→ Liquid gap: <strong style="color:var(--blue)">₹'+fmt(liqFund)+'</strong></div>';
  h+='<div style="font-size:11.5px;margin-bottom:2px;">→ Equity entries: <strong style="color:var(--buy)">₹'+fmt(eqFund)+'</strong> → '+nEntries+' stocks @ ₹'+fmt(perStock)+'/stock</div>';
  if(surplus>500) h+='<div style="font-size:11.5px;">→ Surplus: <strong style="color:var(--purple)">₹'+fmt(surplus)+'</strong> → Liquid park karo</div>';
  h+='</div>';
  h+='</div></div>';
  return h;
}

function buildWhatIfSection(score,tot,prevScore){
  var h='<div class="sim-sec"><div class="sim-sec-hdr amber">🔮 What If — Next Month Score Changes?</div><div class="sim-sec-body">';
  h+='<table class="wi-tbl"><thead><tr>'+
    '<th>Next Score</th><th>Regime</th><th>Equity Target</th>'+
    '<th>Gold Target</th><th>Cash Target</th><th>Weeks</th><th>Main Action</th></tr></thead><tbody>';
  [0,1,2,3].forEach(function(ns){
    var na=ALLOC[ns],isCur=ns===score;
    var weeks=isCur?0:(ns>score?3:Math.abs(ns-score)===2?3:ns<score?4:0);
    if(Math.abs(ns-score)>=2) weeks=Math.abs(ns-score)===2?3:4;
    var action=isCur?'Normal monthly RB only (no shift)':
      ns>score?'Redeem Liquid → Equity ('+weeks+' Fridays)':
               'Sell Equity → Liquid'+(na[1]>ALLOC[score][1]?' + Buy Gold':'')+'('+weeks+' Fridays)';
    h+='<tr class="'+(isCur?'cur':'')+'">'+
      '<td><strong>'+(isCur?'✅ ':'')+(ns===prevScore?'⬅️ ':'')+''+ns+'</strong></td>'+
      '<td>'+LBLS[ns]+'</td>'+
      '<td style="color:var(--buy)">'+na[0]+'% — ₹'+fmt(tot*na[0]/100)+'</td>'+
      '<td style="color:var(--orange)">'+na[1]+'% — ₹'+fmt(tot*na[1]/100)+'</td>'+
      '<td style="color:var(--blue)">'+na[2]+'% — ₹'+fmt(tot*na[2]/100)+'</td>'+
      '<td>'+(isCur?'—':weeks+' wks')+'</td>'+
      '<td style="font-size:10.5px;text-align:left;">'+action+'</td></tr>';
  });
  h+='</tbody></table>';
  h+='<div class="info-box" style="font-size:11px;margin-top:6px;">💡 ✅ = current score | ⬅️ = prev score. Score change pe plan restart hota hai fresh score se. Incomplete weeks abandon nahi hote — fresh plan shuru hota hai.</div>';
  h+='</div></div>';
  return h;
}

function gv(id){var el=document.getElementById(id);return el?parseFloat(el.value)||0:0;}

function simReset(){
  document.getElementById('sim-total').value=2000000;
  document.getElementById('sim-ath').value=2000000;
  document.getElementById('sim-exits').value=200000;
  document.getElementById('sim-capadd').value=0;
  document.getElementById('sim-prev').value=2;
  document.getElementById('sim-cur').value=3;
  document.getElementById('sim-vix').value=16;
  document.getElementById('sim-eq-cur').value=1300000;
  document.getElementById('sim-gold-cur').value=400000;
  document.getElementById('sim-liq-cur').value=300000;
  document.getElementById('sim-gold-cmp').value=125;
  document.getElementById('sim-week').value=0;
  document.getElementById('sim-nav-ret').value=0;
  runSim();
}
window.onload=function(){runSim();};
</script>
</body>
</html>"""

    _sim_stc.html(_SIM_HTML, height=1100, scrolling=True)

# ════════════════════════════════════════════════════════════════════
# SCREENER TAB — all step content
# ════════════════════════════════════════════════════════════════════
with _tab_screener:
    st.session_state["_curr_tab"] = "screener"
    # ── Step progress bar ──────────────────────────────────────────
    st.markdown(step_html(st.session_state.current_step), unsafe_allow_html=True)
    _cur_step = st.session_state.current_step

# ═══════════════════════════════════════════════════════════════
# STEP 1 — UNIVERSE SETUP  (runs only when screener tab active)
# ═══════════════════════════════════════════════════════════════
    if st.session_state.current_step == 1:
        st.markdown('<div class="section-hdr">🌐 Step 1 — Universe Setup</div>', unsafe_allow_html=True)

        c1, c2 = st.columns([1, 2])
        with c1:
            chosen_u = st.selectbox(
                "🌐 Select Universe",
                UNIVERSES,
                index=UNIVERSES.index(st.session_state.universe),
                help="AllNSE = NSE ki sabhi EQ stocks. Baaki = Nifty index lists (GitHub se auto-load)"
            )
            st.session_state.universe = chosen_u
        with c2:
            _u_meta = {
                "Nifty50":  ("50 stocks", "🔵", "Large Cap — India ke top 50"),
                "Nifty100": ("100 stocks","🟢", "Large Cap — top 100"),
                "Nifty200": ("200 stocks","🟡", "Large + Mid Cap"),
                "Nifty250": ("250 stocks","🟠", "Mid Cap focused"),
                "Nifty500": ("500 stocks","🔴", "Large + Mid + Small"),
                "N750":     ("750 stocks","🟣", "Total Market index"),
                "AllNSE":   ("2000+ stocks","⚪","Sabhi NSE EQ stocks"),
            }
            _m = _u_meta.get(chosen_u, ("—","⚪",""))
            st.markdown(f"""
            <div style="background:var(--bg-white);border:1px solid var(--border);border-left:4px solid var(--teal);
                        border-radius:var(--radius-md);padding:12px 18px;margin-top:4px;
                        box-shadow:var(--shadow-sm);display:flex;align-items:center;gap:14px;">
              <span style="font-size:28px;">{_m[1]}</span>
              <div>
                <div style="font-size:15px;font-weight:800;color:var(--text-main)">{chosen_u}</div>
                <div style="font-size:12px;color:var(--muted);margin-top:2px;">
                  <b style="color:var(--blue)">{_m[0]}</b> &nbsp;·&nbsp; {_m[2]}
                </div>
              </div>
            </div>
            """, unsafe_allow_html=True)

        # ── AllNSE: Load Symbol List ───────────────────────────────
        if chosen_u == "AllNSE":

            # GitHub-cached EQUITY_L.csv (committed by cache_builder GitHub Action)
            _GITHUB_EQ_URL = (
                "https://raw.githubusercontent.com/prayan2702/"
                "Streamlit_Momn_v13_Cached_DB/main/EQUITY_L.csv"
            )
            # NSE direct URLs — blocked on cloud, may work locally
            _NSE_FALLBACK_URLS = [
                "https://nsearchives.nseindia.com/content/equities/EQUITY_L.csv",
                "https://archives.nseindia.com/content/equities/EQUITY_L.csv",
                "https://www1.nseindia.com/content/equities/EQUITY_L.csv",
            ]
            _NSE_HDR = {
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0.0.0 Safari/537.36"
                ),
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language": "en-US,en;q=0.9",
                "Accept-Encoding": "gzip, deflate, br",
                "Connection": "keep-alive",
                "Referer": "https://www.nseindia.com/",
                "sec-ch-ua": '"Chromium";v="124", "Google Chrome";v="124", "Not-A.Brand";v="99"',
                "sec-ch-ua-mobile": "?0",
                "sec-ch-ua-platform": '"Windows"',
                "sec-fetch-dest": "document",
                "sec-fetch-mode": "navigate",
                "sec-fetch-site": "same-site",
                "sec-fetch-user": "?1",
                "Upgrade-Insecure-Requests": "1",
            }

            def _parse_eq_bytes(text: str) -> pd.DataFrame:
                import io as _io
                df = pd.read_csv(_io.StringIO(text), skipinitialspace=True)
                df.columns = [c.strip() for c in df.columns]
                if 'SERIES' in df.columns:
                    df = df[df['SERIES'].str.strip() == 'EQ'].copy()
                df['SYMBOL'] = df['SYMBOL'].str.strip().str.upper()
                return df.reset_index(drop=True)

            def _fetch_from_github() -> pd.DataFrame:
                """GitHub repo se cached EQUITY_L.csv fetch karo."""
                resp = requests.get(_GITHUB_EQ_URL, timeout=20)
                resp.raise_for_status()
                return _parse_eq_bytes(resp.text)

            def _fetch_from_nse() -> pd.DataFrame:
                """NSE direct fetch — cloud pe block hota hai, locally kaam karta hai."""
                session = requests.Session()
                for wu in ["https://www.nseindia.com",
                           "https://www.nseindia.com/market-data/securities-available-for-trading"]:
                    try:
                        session.get(wu, headers=_NSE_HDR, timeout=15)
                        time.sleep(0.8)
                    except Exception:
                        pass
                last_err = None
                for url in _NSE_FALLBACK_URLS:
                    try:
                        resp = session.get(url, headers=_NSE_HDR, timeout=30)
                        resp.raise_for_status()
                        return _parse_eq_bytes(resp.text)
                    except Exception as e:
                        last_err = e
                        time.sleep(1)
                raise RuntimeError(str(last_err))

            # ── Status display if already loaded ──────────────────
            if st.session_state.symbols and st.session_state.universe == "AllNSE":
                n = len(st.session_state.symbols)
                st.markdown(f"""<div class="metric-row">
                    {metric_card("Loaded Symbols", f"{n:,}", "green")}
                    {metric_card("Source", st.session_state.get("allnse_source","—"), "blue")}
                </div>""", unsafe_allow_html=True)
                if st.button("🔄 Reload Symbol List", type="secondary"):
                    st.session_state.symbols = None
                    st.session_state.eq_df   = None
                    st.rerun()
            else:
                # ── Info box ──────────────────────────────────────
                st.markdown("""
                <div class="nse-link-box">
                  <div>📥</div>
                  <div>
                    <b>NSE — Securities Available for Trading</b>
                    <div class="hint">
                      GitHub cache se auto-load hoga &nbsp;|&nbsp;
                      NSE direct (local only) &nbsp;|&nbsp; Ya manually CSV upload karo
                    </div>
                  </div>
                </div>
                """, unsafe_allow_html=True)

                # ── Button row: GitHub cache + NSE direct ─────────
                btn_col1, btn_col2, _ = st.columns([1, 1, 1])
                with btn_col1:
                    github_btn = st.button("☁️ Load from GitHub Cache", type="primary",
                                           use_container_width=True,
                                           help="Repo mein committed EQUITY_L.csv se load karo (cloud pe kaam karta hai)")
                with btn_col2:
                    nse_btn = st.button("📡 NSE Direct Fetch", type="secondary",
                                        use_container_width=True,
                                        help="NSE se live fetch — cloud pe block hota hai, locally try karo")

                # ── GitHub cache fetch ────────────────────────────
                if github_btn:
                    with st.spinner("☁️ GitHub se EQUITY_L.csv fetch ho raha hai..."):
                        try:
                            eq_df = _fetch_from_github()
                            syms_ns = [s + ".NS" for s in eq_df["SYMBOL"].tolist()]
                            syms_ns = add_extra_symbols(syms_ns)
                            st.session_state.eq_df           = eq_df
                            st.session_state.symbols          = syms_ns
                            st.session_state.universe_label   = f"AllNSE (GitHub — {len(syms_ns):,} stocks)"
                            st.session_state["allnse_source"] = "GitHub Cache"
                            st.success(f"✅ GitHub cache se {len(syms_ns):,} EQ stocks loaded!")
                            st.rerun()
                        except Exception as e:
                            st.warning(
                                f"⚠️ GitHub fetch failed: `{e}`\n\n"
                                "💡 `EQUITY_L.csv` repo mein commit karo ya manually upload karo."
                            )

                # ── NSE direct fetch ──────────────────────────────
                if nse_btn:
                    with st.spinner("🌐 NSE se EQUITY_L.csv fetch ho raha hai... (cloud pe block hoga)"):
                        try:
                            eq_df = _fetch_from_nse()
                            syms_ns = [s + ".NS" for s in eq_df["SYMBOL"].tolist()]
                            syms_ns = add_extra_symbols(syms_ns)
                            st.session_state.eq_df           = eq_df
                            st.session_state.symbols          = syms_ns
                            st.session_state.universe_label   = f"AllNSE (NSE Live — {len(syms_ns):,} stocks)"
                            st.session_state["allnse_source"] = "NSE Live"
                            st.success(f"✅ NSE se {len(syms_ns):,} EQ stocks loaded!")
                            st.rerun()
                        except Exception as e:
                            st.warning(
                                f"⚠️ NSE direct fetch failed: `{e}`\n\n"
                                "💡 Cloud environment mein NSE IP-block karta hai — "
                                "**GitHub Cache** ya **Manual Upload** use karo."
                            )

                # ── Manual CSV upload — ALWAYS VISIBLE ────────────
                st.markdown("---")
                st.markdown(
                    '<a href="https://www.nseindia.com/static/market-data/securities-available-for-trading" '
                    'target="_blank" style="font-size:12px;color:var(--blue);font-weight:600;">'
                    '📥 NSE se manually EQUITY_L.csv download karein</a>',
                    unsafe_allow_html=True
                )
                uploaded = st.file_uploader(
                    "📂 EQUITY_L.csv upload karein (manual — always works)",
                    type=["csv"], key="equity_csv"
                )
                if uploaded:
                    try:
                        eq_df = parse_equity_csv(uploaded)
                        syms_ns = [s + ".NS" for s in eq_df["SYMBOL"].tolist()]
                        syms_ns = add_extra_symbols(syms_ns)
                        st.session_state.eq_df           = eq_df
                        st.session_state.symbols          = syms_ns
                        st.session_state.universe_label   = f"AllNSE (CSV — {len(syms_ns):,} stocks)"
                        st.session_state["allnse_source"] = "CSV Upload"
                        st.success(f"✅ CSV loaded: **{len(syms_ns):,}** EQ stocks (incl. GOLDBEES & SILVERBEES)")
                        st.rerun()
                    except Exception as e:
                        st.error(f"CSV parse error: {e}")

                if not st.session_state.symbols:
                    st.info("💡 Symbol list load nahi hua — GitHub fallback (NSE_EQ_ALL.csv) screener run pe use hoga.")
                    st.markdown("""
                    <div style="background:var(--amber-bg);border:1px solid #fcd34d;border-radius:var(--radius-md);
                                padding:10px 16px;font-size:12.5px;color:#92400e;margin-top:6px;">
                    ➕ <b>Auto-included:</b> &nbsp;
                    <span style="background:white;border:1px solid #fcd34d;border-radius:12px;padding:2px 10px;font-weight:700;">🥇 GOLDBEES</span>
                    &nbsp;
                    <span style="background:white;border:1px solid #fcd34d;border-radius:12px;padding:2px 10px;font-weight:700;">🥈 SILVERBEES</span>
                    &nbsp; — har universe ke saath automatically add honge
                    </div>
                    """, unsafe_allow_html=True)

        # ── Other universes: auto-fetch info ─────────────────────
        else:
            st.info(f"📡 **{chosen_u}** ki symbol list screener run pe GitHub se auto-load hogi. CSV upload ki zaroorat nahi hai.")
            st.markdown("""
            <div style="background:var(--amber-bg);border:1px solid #fcd34d;border-radius:var(--radius-md);
                        padding:10px 16px;font-size:12.5px;color:#92400e;margin-top:6px;">
            ➕ <b>Auto-included:</b> &nbsp;
            <span style="background:white;border:1px solid #fcd34d;border-radius:12px;padding:2px 10px;font-weight:700;">🥇 GOLDBEES</span>
            &nbsp;
            <span style="background:white;border:1px solid #fcd34d;border-radius:12px;padding:2px 10px;font-weight:700;">🥈 SILVERBEES</span>
            &nbsp; — har universe ke saath automatically add honge
            </div>
            """, unsafe_allow_html=True)
            # Pre-load symbols when user confirms
            if st.button("✅ Load Symbol List", type="primary"):
                with st.spinner(f"Loading {chosen_u} from GitHub…"):
                    try:
                        syms_ns = load_symbols_from_github(chosen_u)
                        syms_ns = add_extra_symbols(syms_ns)
                        st.session_state.symbols = syms_ns
                        st.session_state.universe_label = f"{chosen_u} ({len(syms_ns)} stocks)"
                        st.success(f"✅ {chosen_u}: **{len(syms_ns)}** symbols loaded (incl. GOLDBEES & SILVERBEES)")
                    except Exception as e:
                        st.error(f"Symbol load failed: {e}")

            if st.session_state.symbols and st.session_state.universe == chosen_u:
                n = len(st.session_state.symbols)
                st.markdown(f"""<div class="metric-row">
                    {metric_card("Loaded Symbols", f"{n:,}", "green")}
                    {metric_card("Universe", chosen_u, "blue")}
                </div>""", unsafe_allow_html=True)

        st.divider()

        # ── Cache status (Pre-cached options ke liye info) ────────
        if _CACHE_AVAILABLE:
            _cd = st.session_state.get("cache_selected_date", None)
            st.markdown(get_cache_status_html(cache_date=_cd), unsafe_allow_html=True)
        if _CACHE_UPSTOX_AVAILABLE:
            _cd = st.session_state.get("cache_selected_date", None)
            st.markdown(get_cache_status_html_upstox(cache_date=_cd), unsafe_allow_html=True)
        if _CACHE_ANGEL_AVAILABLE:
            _cd = st.session_state.get("cache_selected_date", None)
            st.markdown(get_cache_status_html_angel(cache_date=_cd), unsafe_allow_html=True)
        if _CACHE_FYERS_AVAILABLE:
            st.markdown(get_cache_status_html_fyers(), unsafe_allow_html=True)

        # ── Next step button ──────────────────────────────────────
        if st.session_state.symbols or chosen_u != "AllNSE":
            if st.button("▶ Next: Run Screener →", type="primary"):
                if st.session_state.symbols is None and chosen_u != "AllNSE":
                    # Will load during screener run
                    pass
                st.session_state.current_step = 2; st.rerun()
        elif chosen_u == "AllNSE" and not st.session_state.symbols:
            if st.button("▶ Next: Run Screener → (GitHub fallback)", type="secondary"):
                st.session_state.current_step = 2; st.rerun()

    # ═══════════════════════════════════════════════════════════════
    # STEP 2 — RUN SCREENER
    # ═══════════════════════════════════════════════════════════════
    elif st.session_state.current_step == 2:
        st.markdown('<div class="section-hdr">📊 Step 2 — Run Momentum Screener</div>', unsafe_allow_html=True)

        if not _CALCS_AVAILABLE:
            st.error("❌ `calculations.py` not found. Project folder mein rakh kar dobara run karo.")
            st.stop()
        if not _DS_AVAILABLE:
            st.warning(
                f"⚠️ `data_service.py` import failed (`{_DS_IMPORT_ERR[:100]}`). "
                "**YFinance** inline fallback use hoga. "
                "Upstox/Angel One ke liye `pyotp` + `smartapi-python` `requirements.txt` mein add karo."
            )

        # ── Filter settings ───────────────────────────────────────
        # Lock filter panel during download so accidental slider touch
        # doesn't reset the page and restart the download.
        _downloading = st.session_state.get("_downloading", False)

        with st.expander("🔧 Filter Settings", expanded=not _downloading):
            if _downloading:
                st.info("⏳ Data download chal raha hai — filters locked hain.")
                # Read-only display of current values during download
                _fp = st.session_state.get("_last_filter_params", {})
                st.markdown(
                    f"Close > 200-DMA: **{_fp.get('use_dma200',True)}** &nbsp;|&nbsp; "
                    f"12M ROC > 5.5%: **{_fp.get('use_roc12',True)}** &nbsp;|&nbsp; "
                    f"Avg Vol: **{_fp.get('volm_cr_min',1.0)} Cr** &nbsp;|&nbsp; "
                    f"Min CMP: **₹{_fp.get('close_min',30.0)}**"
                )
                # Use last saved filter params during download
                filter_params = _fp
            else:
                # ── 2-column layout: checkboxes left, sliders right ─
                f_left, f_right = st.columns([1, 1])
                with f_left:
                    st.markdown("**✅ Filters**")
                    use_dma200 = st.checkbox("Close > 200-day DMA",    value=True)
                    use_roc12  = st.checkbox("12M ROC > 5.5%",         value=True)
                    use_roc_cap= st.checkbox("12M return < 1000x",     value=True)
                    use_ath    = st.checkbox("Within 25% of ATH",      value=True)
                with f_right:
                    st.markdown("**📊 Thresholds**")
                    volm_min    = st.slider("Avg Vol (Cr) >",   0.0, 10.0, 1.0, 0.1)
                    close_min   = st.slider("Min CMP ₹",        0.0, 500.0, 30.0, 5.0)
                    circuit_max = st.slider("Circuit hits/yr <", 1, 100, 20, 1)
                    circuit5    = st.slider("5% circuit 3M ≤",  0, 30, 10, 1)

                filter_params = {
                    "use_dma200": use_dma200, "use_roc12": use_roc12, "use_roc_cap": use_roc_cap,
                    "volm_cr_min": volm_min, "circuit_max": circuit_max, "circuit5_max": circuit5,
                    "use_away_ath": use_ath, "close_min": close_min,
                }
                # Save for use during locked state
                st.session_state["_last_filter_params"] = filter_params

        U          = st.session_state.universe
        api_source = st.session_state.data_source
        end_date   = st.session_state.lookback_date

        col_run, col_info = st.columns([1, 2])
        with col_run:
            _running = st.session_state.get("_run_download", False)
            run_clicked = st.button(
                "⏳ Downloading..." if _running else "▶ Start Data Download",
                type="primary",
                use_container_width=True,
                disabled=_running,
            )
        with col_info:
            n_loaded = len(st.session_state.symbols) if st.session_state.symbols else "—"
            st.markdown(f"""
            <div style="background:var(--bg-white);border:1px solid var(--border);border-radius:var(--radius-md);
                        padding:10px 16px;font-size:12.5px;color:var(--text-sub);line-height:2;
                        box-shadow:var(--shadow-sm);">
            🌐 Universe: <b style="color:var(--text-main)">{U}</b> &nbsp;|&nbsp;
            📋 Symbols: <b style="color:var(--blue)">{n_loaded}</b> &nbsp;|&nbsp;
            📅 End: <b style="color:var(--text-main)">{end_date.strftime('%d-%m-%Y')}</b><br>
            📐 Method: <b style="color:var(--violet)">{st.session_state.ranking_method}</b> &nbsp;|&nbsp;
            📡 Source: <b style="color:var(--teal)">{api_source}</b>
            </div>""", unsafe_allow_html=True)

        # ── Button click: set persistent trigger, save params, rerun ─
        # We DON'T run the download inside `if run_clicked:` because
        # Streamlit can rerun during a long download (WebSocket reconnect),
        # at which point run_clicked=False and _downloading stays True forever.
        # Instead: button sets a session-state flag; download runs on next render
        # via `if _run_download:` — survives reconnects, always cleans up.
        if run_clicked:
            st.session_state["_run_download"]      = True
            st.session_state["_run_filter_params"] = filter_params
            st.session_state["_run_api_source"]    = api_source
            st.session_state["_run_end_date"]      = end_date
            st.session_state["_run_u"]             = U
            # Reset cross-source review for fresh run
            st.session_state["_cross_review_done"]      = False
            st.session_state["_cross_review_overrides"] = {}
            st.session_state["_pending_topup"]          = False
            st.session_state["_cross_diff_df"]          = None
            st.session_state["_cross_error"]            = None
            st.session_state["_cross_error_detail"]     = None
            st.rerun()

        if st.session_state.get("_run_download", False):
            # Restore params saved at click-time
            _filter_params = st.session_state.get("_run_filter_params", filter_params)
            _api_source    = st.session_state.get("_run_api_source",    api_source)
            _end_date      = st.session_state.get("_run_end_date",      end_date)
            _U             = st.session_state.get("_run_u",             U)

            dates     = build_dates(_end_date)
            prog_bar  = st.progress(0)
            status_tx = st.empty()
            _download_ok = False
            try:

                # ══════════════════════════════════════════════════════
                # BRANCH A — Pre-cached (Instant load from GitHub)
                # Supports both YFinance cache and Upstox cache
                # ══════════════════════════════════════════════════════
                if _api_source in (
                    "📦 Pre-cached YFinance",
                    "📦 Pre-cached Upstox",
                    "📦 Pre-cached Angel One",
                    "📦 Pre-cached Fyers",
                ):
                    _is_upstox_cache = (_api_source == "📦 Pre-cached Upstox")
                    _is_angel_cache  = (_api_source == "📦 Pre-cached Angel One")
                    _is_fyers_cache  = (_api_source == "📦 Pre-cached Fyers")

                    # ── Select correct loader ─────────────────────────
                    if _is_fyers_cache:
                        if not _CACHE_FYERS_AVAILABLE:
                            st.error(
                                "❌ cache_loader_fyers.py nahi mila. "
                                "Repo mein add karo + daily_cache_fyers.yml workflow run karo."
                            )
                            st.stop()
                        _loader    = load_cache_fyers
                        _meta_fn   = get_cache_meta_fyers
                        _age_fn    = get_cache_age_days_fyers
                        _cache_lbl = "Fyers"
                    elif _is_angel_cache:
                        if not _CACHE_ANGEL_AVAILABLE:
                            st.error(
                                "❌ cache_loader_angelone.py nahi mila. "
                                "Repo mein add karo + daily_cache_angelone.yml workflow run karo."
                            )
                            st.stop()
                        _loader    = load_cache_angel
                        _meta_fn   = get_cache_meta_angel
                        _age_fn    = get_cache_age_days_angel
                        _cache_lbl = "Angel One"
                    elif _is_upstox_cache:
                        if not _CACHE_UPSTOX_AVAILABLE:
                            st.error("❌ cache_loader_upstox.py nahi mila. Repo mein add karo.")
                            st.stop()
                        _loader    = load_cache_upstox
                        _meta_fn   = get_cache_meta_upstox
                        _age_fn    = get_cache_age_days_upstox
                        _cache_lbl = "Upstox"
                    else:
                        if not _CACHE_AVAILABLE:
                            st.error("❌ cache_loader.py nahi mila. cache_loader.py repo mein add karo.")
                            st.stop()
                        _loader    = load_cache
                        _meta_fn   = get_cache_meta
                        _age_fn    = get_cache_age_days
                        _cache_lbl = "YFinance"

                    _cache_date = st.session_state.get("cache_selected_date", None)
                    status_tx.markdown(
                        f"⚡ **{_cache_lbl} cache ({_cache_date or 'latest'}) load ho raha hai...**"
                    )
                    prog_bar.progress(0.1)
                    try:
                        close, high, volume = _loader(cache_date=_cache_date)
                        prog_bar.progress(0.85)
                        status_tx.markdown("✅ **Cache loaded!** Calculations shuru ho rahi hain...")

                        # ── Meta & failed symbols (Issues 3 & 4) ─────
                        meta = _meta_fn(cache_date=_cache_date)

                        # Issue 3: meta se ALL failed symbols lo (not just volume blank)
                        failed_from_meta = meta.get("failed_symbols", [])
                        failed_from_meta = [s.replace(".NS", "") for s in failed_from_meta]

                        # Volume-based blank detection (symbols fetched but all NaN data)
                        volume12M_check = volume.loc[dates['date12M']:].copy() if not volume.empty else pd.DataFrame()
                        median_volume   = volume12M_check.median() if not volume12M_check.empty else pd.Series()
                        vol_blank       = median_volume[median_volume.isna()].index.tolist()
                        vol_blank       = [t.replace('.NS', '') for t in vol_blank]

                        # Merge both lists (dedup, preserve order)
                        failed_blank = list(dict.fromkeys(failed_from_meta + vol_blank))
                        st.session_state.failed_blank = failed_blank

                        age = _age_fn()
                        _cache_last_date_str = meta.get("last_date_in_cache", "?")
                        _today_flag          = meta.get("today_data_present", False)
                        _fresh_icon          = "✅" if _today_flag else "⚠️"
                        if age > 3:
                            st.warning(
                                f"⚠️ {_cache_lbl} Cache {int(age)} din purana hai "
                                f"(build: {meta.get('build_date','?')}). "
                                "Data slightly stale ho sakta hai."
                            )
                        else:
                            st.success(
                                f"✅ {_cache_lbl} Cache loaded! "
                                f"{meta.get('symbols_fetched','?'):,} symbols | "
                                f"Build: {meta.get('build_date','?')} | "
                                f"Last date: {_fresh_icon} {_cache_last_date_str} | "
                                f"Age: {int(age)} din"
                            )

                        # ── Stale close check → pending_topup ────────────────
                        import datetime as _dtm2
                        _today_date  = _dtm2.date.today()
                        _wd          = _today_date.weekday()
                        if _wd == 5:   _last_trade  = _today_date - _dtm2.timedelta(days=1)
                        elif _wd == 6: _last_trade  = _today_date - _dtm2.timedelta(days=2)
                        else:          _last_trade  = _today_date
                        _cache_ld = None
                        try:    _cache_ld = _dtm2.date.fromisoformat(_cache_last_date_str)
                        except: pass
                        if _cache_ld is not None and _cache_ld < _last_trade:
                            st.session_state["_pending_topup"]       = True
                            st.session_state["_topup_close"]         = close
                            st.session_state["_topup_high"]          = high
                            st.session_state["_topup_volume"]        = volume
                            st.session_state["_topup_fp"]            = _filter_params
                            st.session_state["_topup_dates"]         = dates
                            st.session_state["_topup_cache_lbl"]     = _cache_lbl
                            st.session_state["_topup_api_source"]    = _api_source
                            st.session_state["_topup_last_date"]     = str(_cache_ld)
                            st.session_state["_topup_target_date"]   = str(_last_trade)
                            st.rerun()

                        # ── Issue 4: Missing stocks detection + YFinance top-up ──
                        _universe_syms = st.session_state.symbols or []
                        _cache_syms    = set(close.columns.str.replace('.NS', '', regex=False).str.upper())
                        _missing = [
                            s for s in _universe_syms
                            if s.replace('.NS', '').upper() not in _cache_syms
                        ]
                        if _missing:
                            # ── BUG FIX: Don't show a button inside the download
                            # flow — it gets skipped because code falls through to
                            # build_dfStats and st.rerun() fires before the user
                            # can click it.
                            # Instead: save the loaded cache data + params into
                            # session state, set a pending flag, and st.stop().
                            # The merge decision UI is rendered on the NEXT pass
                            # (below, in the "_pending_merge" block) where Streamlit
                            # can actually wait for a button click.
                            st.session_state["_cache_missing_syms"] = _missing
                            st.session_state["_pending_close"]      = close
                            st.session_state["_pending_high"]       = high
                            st.session_state["_pending_volume"]     = volume
                            st.session_state["_pending_fp"]         = _filter_params
                            st.session_state["_pending_dates"]      = dates
                            st.session_state["_pending_merge"]      = True
                            # Use st.rerun() NOT st.stop():
                            # st.stop() halts render but never triggers a new pass,
                            # so the pending merge UI never appears (hang/stuck).
                            # st.rerun() raises RerunException -> except ignores it
                            # -> finally clears _run_download -> fresh render shows
                            # the pending merge decision block correctly.
                            st.rerun()

                    except Exception as e:
                        st.error(f"❌ Cache load failed: {e}. YFinance select karke retry karo.")
                        st.stop()

                # ══════════════════════════════════════════════════════
                # BRANCH B — Live fetch (YFinance / Upstox / Angel One)
                # ══════════════════════════════════════════════════════
                else:
                    # ── Load symbols ──────────────────────────────────
                    if st.session_state.symbols is None or st.session_state.universe != U:
                        with st.spinner(f"Loading {_U} symbols…"):
                            try:
                                if _U == "AllNSE":
                                    url = f"{GITHUB_BASE}/NSE_EQ_ALL.csv"
                                    df_sym = pd.read_csv(url)
                                    df_sym['Yahoo_Symbol'] = df_sym['Symbol'].astype(str).str.strip() + '.NS'
                                    syms_ns = df_sym['Yahoo_Symbol'].tolist()
                                else:
                                    syms_ns = load_symbols_from_github(_U)
                                syms_ns = add_extra_symbols(syms_ns)
                                st.session_state.symbols = syms_ns
                            except Exception as e:
                                st.error(f"Symbol list load failed: {e}"); st.stop()

                    symbols = st.session_state.symbols
                    CHUNK   = 50 if _api_source == "Upstox" else (15 if _U == "AllNSE" else 50)
                    st.markdown(f"""
                    <div style="display:flex;gap:10px;flex-wrap:wrap;margin:8px 0 12px 0;">
                      <span style="background:var(--blue-bg);color:var(--blue);border:1px solid var(--blue-bdr);
                                   border-radius:20px;padding:3px 12px;font-size:12px;font-weight:700;">
                        📦 Chunk: {CHUNK}
                      </span>
                      <span style="background:var(--green-bg);color:var(--green);border:1px solid var(--green-bdr);
                                   border-radius:20px;padding:3px 12px;font-size:12px;font-weight:700;">
                        📋 Symbols: {len(symbols):,}
                      </span>
                      <span style="background:var(--violet-bg);color:var(--violet);border:1px solid #c4b5fd;
                                   border-radius:20px;padding:3px 12px;font-size:12px;font-weight:700;">
                        📡 Source: {_api_source}
                      </span>
                      <span style="background:var(--amber-bg);color:#92400e;border:1px solid #fcd34d;
                                   border-radius:20px;padding:3px 12px;font-size:12px;font-weight:700;">
                        🥇 GOLDBEES &amp; 🥈 SILVERBEES included
                      </span>
                    </div>
                    """, unsafe_allow_html=True)

                    # Use data_service if available AND source is not YFinance
                    _use_ds = _DS_AVAILABLE and _api_source in ("Upstox", "Angel One", "Fyers")
                    try:
                        if _use_ds:
                            close, high, volume, failed_symbols = fetch_data(
                                api_source   = _api_source,
                                symbols      = symbols,
                                start_date   = dates['startDate'],
                                end_date     = dates['endDate'],
                                chunk_size   = CHUNK,
                                progress_bar = prog_bar,
                                status_text  = status_tx,
                            )
                        else:
                            # YFinance inline fallback
                            close, high, volume, failed_symbols = _fetch_yfinance_inline(
                                symbols, dates['startDate'], dates['endDate'],
                                prog_bar, status_tx, chunk_size=CHUNK
                            )
                    except Exception as e:
                        st.error(f"Data fetch error: {e}"); st.stop()

                    if close is None or close.empty:
                        st.session_state["_downloading"] = False
                        st.error("❌ Data fetch hua nahi. Internet / token check karo."); st.stop()

                    # ── Failed symbols ────────────────────────────────
                    # 1. API se directly failed (instrument not found / network error)
                    api_failed = [t.replace('.NS','') for t in (failed_symbols or [])]
                    # 2. Volume-based blank detection (data fetched but all NaN)
                    volume12M_check = volume.loc[dates['date12M']:].copy() if not volume.empty else pd.DataFrame()
                    median_volume   = volume12M_check.median() if not volume12M_check.empty else pd.Series()
                    vol_blank       = median_volume[median_volume.isna()].index.tolist()
                    vol_blank       = [t.replace('.NS','') for t in vol_blank]
                    # Merge both lists (dedup, preserve order)
                    failed_blank = list(dict.fromkeys(api_failed + vol_blank))
                    st.session_state.failed_blank = failed_blank
                    # Note: failed stocks display results section mein hoga (rerun ke baad)
                    # Yahan show karne se wo 1 second ke liye dikha aur rerun pe gaayab ho jaata tha

                # ── Calculate metrics ─────────────────────────────────
                status_tx.markdown("⏳ **Calculating momentum metrics...**")
                prog_bar.progress(0.92)
                try:
                    dfStats   = build_dfStats(close, high, volume, dates, st.session_state.ranking_method)
                    dfFiltered = apply_filters(dfStats.copy(), _filter_params)
                    st.session_state.dfStats    = dfStats
                    st.session_state.dfFiltered = dfFiltered
                    st.session_state.screener_done = True
                    st.session_state["_cross_filter_params"] = _filter_params  # for cross-source review
                    st.session_state["_downloading"] = False   # unlock filters
                    prog_bar.progress(1.0)
                    status_tx.markdown("✅ **Screener complete!**")

                    _download_ok = True

                except Exception as e:
                    # StopException (st.stop) silently re-raises; all others show error.
                    _cls = type(e).__name__
                    if "StopException" not in _cls and "RerunException" not in _cls:
                        st.error(f"Download/Calculation error: {e}")
                    raise

            finally:
                # ALWAYS clear the flag — even on st.stop() or any error.
                # This prevents the permanently-disabled-button stuck state.
                st.session_state["_run_download"] = False

            if _download_ok:
                st.rerun()  # fresh render → shows results cleanly



        # ══════════════════════════════════════════════════════════
        # PENDING MERGE DECISION
        # Shown after pre-cached load when universe has stocks that
        # are missing from the cache. Rendered on a clean pass so
        # Streamlit can actually wait for a button click.
        # ══════════════════════════════════════════════════════════
        if st.session_state.get("_pending_merge", False):
            _missing_list  = st.session_state.get("_cache_missing_syms", [])
            _p_close       = st.session_state["_pending_close"]
            _p_high        = st.session_state["_pending_high"]
            _p_volume      = st.session_state["_pending_volume"]
            _p_fp          = st.session_state["_pending_fp"]
            _p_dates       = st.session_state["_pending_dates"]
            _p_failed      = list(st.session_state.failed_blank or [])

            _n_uni   = len(st.session_state.symbols or [])
            _n_cache = _p_close.shape[1]

            st.info(
                f"ℹ️ Cache mein **{len(_missing_list)} stocks missing** hain "
                f"(Universe: {_n_uni:,} symbols, Cache: {_n_cache:,} symbols). "
                f"YFinance se fetch karke merge kar sakte hain — ya seedha calculate karo."
            )

            _prog_pm  = st.progress(0)
            _stat_pm  = st.empty()

            col_fetch, col_skip = st.columns(2)

            with col_fetch:
                if st.button(
                    f"📡 Fetch {len(_missing_list)} missing stocks (YFinance) & merge",
                    key="fetch_missing_btn", type="secondary", use_container_width=True
                ):
                    with st.spinner(f"YFinance se {len(_missing_list)} missing stocks fetch ho rahi hain..."):
                        try:
                            _m_close, _m_high, _m_vol, _m_failed = _fetch_yfinance_inline(
                                _missing_list, _p_dates['startDate'], _p_dates['endDate'],
                                _prog_pm, _stat_pm, chunk_size=15
                            )
                            if _m_close is not None and not _m_close.empty:
                                # ── MERGE FIX ─────────────────────────────────────
                                # _m_close sirf 3 NEW columns hai (existing symbols nahi).
                                # reindex+combine_first galat tha — YFinance 2000→today
                                # fetch karta hai, to _m_close.index mein pre-40M dates
                                # aate hain. union() → _p_close.reindex() pe Upstox
                                # ke 2131 columns ke liye 2000-2022 ke rows = NaN.
                                # Result: Close = NaN (wohi pehle wala bug wapas aa jaata).
                                #
                                # Sahi approach: sirf nayi columns add karo,
                                # existing DataFrame ka index TOUCH MAT KARO.
                                # _m_close ko _p_close ke index pe align karo (ffill safety)
                                # aur concat(axis=1) se jodo.
                                _cache_idx = _p_close.index
                                _new_close  = _m_close.reindex(_cache_idx).ffill().bfill()
                                _new_high   = _m_high.reindex(_cache_idx).ffill().bfill()
                                _new_vol    = _m_vol.reindex(_cache_idx).ffill().bfill()
                                _p_close  = pd.concat([_p_close,  _new_close],  axis=1)
                                _p_high   = pd.concat([_p_high,   _new_high],   axis=1)
                                _p_volume = pd.concat([_p_volume, _new_vol],    axis=1)
                                # Duplicate columns remove (safety)
                                _p_close  = _p_close.loc[:,  ~_p_close.columns.duplicated()]
                                _p_high   = _p_high.loc[:,   ~_p_high.columns.duplicated()]
                                _p_volume = _p_volume.loc[:, ~_p_volume.columns.duplicated()]
                                st.success(
                                    f"✅ {_m_close.shape[1] - len(_m_failed)} missing stocks merged! "
                                    f"Total: {_p_close.shape[1]:,} symbols"
                                )
                            if _m_failed:
                                _p_failed = list(dict.fromkeys(
                                    _p_failed + [t.replace('.NS', '') for t in _m_failed]
                                ))
                        except Exception as _me:
                            st.warning(f"Missing stocks fetch failed: {_me}")

                    # Calculate after merge
                    _stat_pm.markdown("⏳ **Calculating momentum metrics...**")
                    _prog_pm.progress(0.92)
                    try:
                        _dfS  = build_dfStats(_p_close, _p_high, _p_volume, _p_dates,
                                              st.session_state.ranking_method)
                        _dfF  = apply_filters(_dfS.copy(), _p_fp)
                        st.session_state.dfStats         = _dfS
                        st.session_state.dfFiltered      = _dfF
                        st.session_state.failed_blank    = _p_failed
                        st.session_state.screener_done   = True
                        st.session_state["_cross_filter_params"] = _p_fp  # for cross-source review
                        st.session_state["_pending_merge"] = False
                        _prog_pm.progress(1.0)
                        _stat_pm.markdown("✅ **Screener complete!**")
                    except Exception as _ce:
                        st.error(f"Calculation error: {_ce}")
                        st.stop()
                    st.rerun()

            with col_skip:
                if st.button(
                    "⏭ Skip — Calculate without missing stocks",
                    key="skip_missing_btn", type="primary", use_container_width=True
                ):
                    # Calculate directly with what cache gave us
                    _stat_pm.markdown("⏳ **Calculating momentum metrics...**")
                    _prog_pm.progress(0.92)
                    try:
                        _dfS  = build_dfStats(_p_close, _p_high, _p_volume, _p_dates,
                                              st.session_state.ranking_method)
                        _dfF  = apply_filters(_dfS.copy(), _p_fp)
                        st.session_state.dfStats         = _dfS
                        st.session_state.dfFiltered      = _dfF
                        st.session_state.failed_blank    = _p_failed
                        st.session_state.screener_done   = True
                        st.session_state["_cross_filter_params"] = _p_fp  # for cross-source review
                        st.session_state["_pending_merge"] = False
                        _prog_pm.progress(1.0)
                        _stat_pm.markdown("✅ **Screener complete!**")
                    except Exception as _ce:
                        st.error(f"Calculation error: {_ce}")
                        st.stop()
                    st.rerun()

            # Block results display until user makes a decision
            st.stop()


        # ══════════════════════════════════════════════════════════
        # PENDING CLOSE TOP-UP
        # Shown when cache last_date < last trading day.
        # User selects source to fetch missing closing prices,
        # which are appended as a new row and calculations run fresh.
        # Source options depend on which cache is currently loaded:
        #   Upstox cache    → Pre-cached Angel One | Pre-cached YFinance | YFinance
        #   Angel One cache → Pre-cached Upstox    | Pre-cached YFinance | YFinance
        #   YFinance cache  → Pre-cached Upstox    | Pre-cached Angel One
        # ══════════════════════════════════════════════════════════
        if st.session_state.get("_pending_topup", False):
            _tp_close   = st.session_state["_topup_close"]
            _tp_high    = st.session_state["_topup_high"]
            _tp_volume  = st.session_state["_topup_volume"]
            _tp_fp      = st.session_state["_topup_fp"]
            _tp_dates   = st.session_state["_topup_dates"]
            _tp_lbl     = st.session_state["_topup_cache_lbl"]
            _tp_apisrc  = st.session_state["_topup_api_source"]
            _tp_last    = st.session_state["_topup_last_date"]
            _tp_target  = st.session_state["_topup_target_date"]
            _tp_prog    = st.progress(0)
            _tp_stat    = st.empty()

            st.markdown(
                f'''<div style="background:#fff7ed;border:1px solid #fed7aa;border-left:4px solid #f97316;
    border-radius:10px;padding:12px 16px;margin:8px 0;font-size:13px;color:#7c2d12;">
    ⚠️ <b>{_tp_lbl} cache last date: {_tp_last}</b> &nbsp;|&nbsp;
    Expected last trading day: <b>{_tp_target}</b><br>
    <span style="font-size:12px;">Cache mein latest closing prices nahi hain.
    Source choose karo aur Close top-up karo — ya skip karke cached data se hi calculate karo.</span>
    </div>''', unsafe_allow_html=True)

            # ── Source options based on which cache is loaded ─────
            _is_upstox_tp  = "Upstox"    in _tp_apisrc and "Angel" not in _tp_apisrc
            _is_angel_tp   = "Angel One" in _tp_apisrc
            _is_yf_tp      = "YFinance"  in _tp_apisrc

            if _is_upstox_tp:
                _topup_opts = [
                    "📦 Pre-cached Angel One (recommended)",
                    "📦 Pre-cached YFinance",
                    "🌐 YFinance (live)",
                ]
            elif _is_angel_tp:
                _topup_opts = [
                    "📦 Pre-cached Upstox (recommended)",
                    "📦 Pre-cached YFinance",
                    "🌐 YFinance (live)",
                ]
            else:  # YFinance cache
                _topup_opts = [
                    "📦 Pre-cached Upstox (recommended)",
                    "📦 Pre-cached Angel One",
                ]

            _tp_c1, _tp_c2, _tp_c3 = st.columns([2, 2, 1])
            with _tp_c1:
                _chosen_tp = st.selectbox("📡 Top-up source", _topup_opts, key="topup_src_sel")
            with _tp_c2:
                _do_topup  = st.button(f"🔄 Top-up Close", key="topup_btn", type="primary", use_container_width=True)
            with _tp_c3:
                _skip_tp   = st.button("⏭ Skip", key="skip_topup_btn", use_container_width=True)

            if _do_topup:
                import datetime as _dtt
                _syms_tp  = list(_tp_close.columns)
                _new_close = _new_high = _new_vol = None

                # ── Helper: merge new rows into cache DFs ─────────
                def _merge_topup(tp_df, new_df):
                    if new_df is None or new_df.empty:
                        return tp_df
                    _existing = set(tp_df.index)
                    _rows = new_df[~new_df.index.isin(_existing)]
                    if _rows.empty:
                        return tp_df
                    _nr = _rows.reindex(columns=tp_df.columns)
                    return pd.concat([tp_df, _nr], axis=0).sort_index()

                _tp_stat.text(f"Fetching missing close from {_chosen_tp}...")
                _tp_prog.progress(0.1)

                try:
                    if "Pre-cached Angel One" in _chosen_tp:
                        if not _CACHE_ANGEL_AVAILABLE:
                            st.error("Angel One cache loader nahi mila.")
                        else:
                            _ac, _ah, _av = load_cache_angel()
                            _ac.index = pd.to_datetime(_ac.index)
                            _ah.index = pd.to_datetime(_ah.index)
                            _av.index = pd.to_datetime(_av.index)
                            # Rows in angel cache that are NOT in topup cache
                            _ac.columns = _ac.columns.str.upper()
                            _ah.columns = _ah.columns.str.upper()
                            _av.columns = _av.columns.str.upper()
                            _tp_c_norm  = _tp_close.copy(); _tp_c_norm.columns = _tp_c_norm.columns.str.upper()
                            _new_close  = _ac[~_ac.index.isin(set(_tp_c_norm.index))]
                            _new_high   = _ah[~_ah.index.isin(set(_tp_c_norm.index))]
                            _new_vol    = _av[~_av.index.isin(set(_tp_c_norm.index))]
                            # Restore original column names
                            def _remap(df, ref):
                                _mp = {c.replace(".NS","").upper(): c for c in ref.columns}
                                df.columns = [_mp.get(c, c) for c in df.columns]
                                return df
                            if _new_close is not None: _new_close = _remap(_new_close, _tp_close)
                            if _new_high  is not None: _new_high  = _remap(_new_high,  _tp_high)
                            if _new_vol   is not None: _new_vol   = _remap(_new_vol,   _tp_volume)

                    elif "Pre-cached Upstox" in _chosen_tp:
                        if not _CACHE_UPSTOX_AVAILABLE:
                            st.error("Upstox cache loader nahi mila.")
                        else:
                            _uc, _uh, _uv = load_cache_upstox()
                            _uc.index = pd.to_datetime(_uc.index)
                            _uh.index = pd.to_datetime(_uh.index)
                            _uv.index = pd.to_datetime(_uv.index)
                            _uc.columns = _uc.columns.str.upper()
                            _uh.columns = _uh.columns.str.upper()
                            _uv.columns = _uv.columns.str.upper()
                            _tp_c_norm  = _tp_close.copy(); _tp_c_norm.columns = _tp_c_norm.columns.str.upper()
                            _new_close  = _uc[~_uc.index.isin(set(_tp_c_norm.index))]
                            _new_high   = _uh[~_uh.index.isin(set(_tp_c_norm.index))]
                            _new_vol    = _uv[~_uv.index.isin(set(_tp_c_norm.index))]
                            def _remap2(df, ref):
                                _mp = {c.replace(".NS","").upper(): c for c in ref.columns}
                                df.columns = [_mp.get(c, c) for c in df.columns]
                                return df
                            if _new_close is not None: _new_close = _remap2(_new_close, _tp_close)
                            if _new_high  is not None: _new_high  = _remap2(_new_high,  _tp_high)
                            if _new_vol   is not None: _new_vol   = _remap2(_new_vol,   _tp_volume)

                    else:  # YFinance live
                        _yf_end_tp  = datetime.datetime.combine(
                            _dtt.date.today() + _dtt.timedelta(days=1), datetime.time())
                        _yf_st_tp   = _tp_close.index[-1] + _dtt.timedelta(days=1)
                        _chunks_c, _chunks_h, _chunks_v = [], [], []
                        _csz = 50
                        for _ki in range(0, len(_syms_tp), _csz):
                            _ch = _syms_tp[_ki:_ki+_csz]
                            try:
                                _raw = yf.download(_ch, start=_yf_st_tp, end=_yf_end_tp,
                                                   progress=False, auto_adjust=True,
                                                   threads=True, multi_level_index=False)
                                if not _raw.empty:
                                    if "Close"  in _raw.columns: _chunks_c.append(_raw["Close"])
                                    if "High"   in _raw.columns: _chunks_h.append(_raw["High"])
                                    if "Close"  in _raw.columns and "Volume" in _raw.columns:
                                        _chunks_v.append(_raw["Close"] * _raw["Volume"])
                            except Exception: pass
                            _tp_prog.progress(min(0.1 + 0.7*(_ki+_csz)/len(_syms_tp), 0.80))
                        if _chunks_c:
                            _new_close = pd.concat(_chunks_c, axis=1)
                            if _chunks_h: _new_high = pd.concat(_chunks_h, axis=1)
                            if _chunks_v: _new_vol  = pd.concat(_chunks_v, axis=1)
                            for _df in [_new_close, _new_high, _new_vol]:
                                if _df is not None:
                                    _df.index = pd.to_datetime(_df.index)
                                    if hasattr(_df.index, 'tz') and _df.index.tz:
                                        _df.index = _df.index.tz_localize(None)

                except Exception as _te:
                    st.warning(f"Top-up fetch error: {_te}")

                _tp_prog.progress(0.85)
                # Merge
                _tp_close  = _merge_topup(_tp_close,  _new_close)
                _tp_high   = _merge_topup(_tp_high,   _new_high)
                _tp_volume = _merge_topup(_tp_volume, _new_vol)

                _new_last  = _tp_close.index[-1].strftime('%d-%b-%Y') if not _tp_close.empty else "?"
                if _new_close is not None and not _new_close.empty:
                    st.success(f"✅ Top-up done! New last date: **{_new_last}**")
                else:
                    st.info("ℹ️ No new rows added (market holiday, or data already current). Proceeding with cached data.")

                _tp_prog.progress(0.92)
                _tp_stat.text("Calculating momentum metrics...")
                try:
                    _dfS = build_dfStats(_tp_close, _tp_high, _tp_volume, _tp_dates,
                                         st.session_state.ranking_method)
                    _dfF = apply_filters(_dfS.copy(), _tp_fp)
                    st.session_state.dfStats       = _dfS
                    st.session_state.dfFiltered    = _dfF
                    st.session_state.screener_done = True
                    st.session_state["_cross_filter_params"] = _tp_fp
                    st.session_state["_pending_topup"] = False
                    _tp_prog.progress(1.0)
                    _tp_stat.text("✅ Screener complete!")
                except Exception as _te2:
                    st.error(f"Calculation error: {_te2}"); st.stop()
                st.rerun()

            if _skip_tp:
                _tp_stat.text("Calculating with cached data...")
                try:
                    _dfS = build_dfStats(_tp_close, _tp_high, _tp_volume, _tp_dates,
                                         st.session_state.ranking_method)
                    _dfF = apply_filters(_dfS.copy(), _tp_fp)
                    st.session_state.dfStats       = _dfS
                    st.session_state.dfFiltered    = _dfF
                    st.session_state.screener_done = True
                    st.session_state["_cross_filter_params"] = _tp_fp
                    st.session_state["_pending_topup"] = False
                except Exception as _te3:
                    st.error(f"Calculation error: {_te3}"); st.stop()
                st.rerun()

            st.stop()

        # ══════════════════════════════════════════════════════════
        # CROSS-SOURCE REVIEW BLOCK
        # Angel One vs Upstox ke beech top 400 stocks mein
        # Close / ATH major difference detect karke user ko
        # per-stock source select karne ka option deta hai.
        # ══════════════════════════════════════════════════════════
        _src = st.session_state.get("data_source", "")
        _cross_done = st.session_state.get("_cross_review_done", False)

        # Determine if source is a broker pre-cache (Angel One or Upstox)
        _is_angel_src  = "Angel One" in _src
        _is_upstox_src = "Upstox" in _src and "Angel" not in _src

        # Trigger: screener done, broker source selected, review not yet done
        if (
            st.session_state.screener_done
            and st.session_state.dfStats is not None
            and (_is_angel_src or _is_upstox_src)
            and not _cross_done
            and not st.session_state.get("_pending_merge", False)
        ):
            # ── Step A: Load secondary cache & compute diff (only once) ──
            if st.session_state.get("_cross_diff_df") is None:
                _TOP_N_COMPARE = st.session_state.get("_cross_top_n", 400)
                _primary_lbl   = "Angel One" if _is_angel_src else "Upstox"
                _secondary_lbl = "Upstox" if _is_angel_src else "Angel One"

                # Check secondary cache availability before trying
                _sec_avail = (_secondary_lbl == "Upstox" and _CACHE_UPSTOX_AVAILABLE) or                          (_secondary_lbl == "Angel One" and _CACHE_ANGEL_AVAILABLE)

                if not _sec_avail:
                    # Secondary cache not available — store error, show to user
                    st.session_state["_cross_error"] = (
                        f"{_secondary_lbl} cache loader ({_secondary_lbl.lower().replace(' ','')}"
                        f"_loader.py) available nahi hai. "
                        f"Dono broker caches GitHub pe hone chahiye cross-review ke liye."
                    )
                    st.session_state["_cross_diff_df"] = pd.DataFrame()  # empty = no diff

                else:
                    _cross_error = None
                    with st.spinner(f"🔄 {_secondary_lbl} cache load ho raha hai — top {_TOP_N_COMPARE} stocks ka comparison..."):
                        try:
                            # Load secondary cache (3 return values: close, high, volume)
                            if _secondary_lbl == "Upstox":
                                _sec_close, _sec_high, _ = load_cache_upstox()
                            else:
                                _sec_close, _sec_high, _ = load_cache_angel()

                            # Normalize column names (remove .NS suffix, uppercase)
                            def _norm_cols(df):
                                df = df.copy()
                                df.columns = df.columns.str.replace(".NS","",regex=False).str.upper()
                                return df
                            _sec_close = _norm_cols(_sec_close)
                            _sec_high  = _norm_cols(_sec_high)

                            # Get top N — use dfStats with robust Rank access
                            _dfS_cx = st.session_state.dfStats.copy()
                            # Ensure Ticker is a column (not index)
                            if "Ticker" not in _dfS_cx.columns:
                                _dfS_cx = _dfS_cx.reset_index()
                            # Ensure Rank is a column; if missing, synthesize from sort order
                            if "Rank" not in _dfS_cx.columns:
                                # Try one more reset (multi-index edge case)
                                _dfS_cx = _dfS_cx.reset_index()
                            if "Rank" not in _dfS_cx.columns:
                                # Fallback: assign rank by current sort order
                                _dfS_cx = _dfS_cx.reset_index(drop=True)
                                _dfS_cx["Rank"] = range(1, len(_dfS_cx) + 1)
                            _top_df = _dfS_cx[_dfS_cx["Rank"] <= _TOP_N_COMPARE].copy()

                            # Thresholds
                            _CLOSE_THRESH = 2.0   # %
                            _ATH_THRESH   = 10.0  # %
                            _AWAY_THRESH  = 8.0   # pp

                            _rows = []
                            for _, _row in _top_df.iterrows():
                                _tick = str(_row["Ticker"]).replace(".NS","").upper()
                                if _tick not in _sec_close.columns or _tick not in _sec_high.columns:
                                    continue
                                _sc_series = _sec_close[_tick].dropna()
                                _sh_series = _sec_high[_tick].dropna()
                                if _sc_series.empty or _sh_series.empty:
                                    continue
                                _sec_cl  = float(_sc_series.iloc[-1])
                                _sec_ath = float(_sh_series.max())
                                _sec_aw  = (_sec_cl - _sec_ath) / _sec_ath * 100

                                _pri_cl  = float(_row.get("Close",    0) or 0)
                                _pri_ath = float(_row.get("ATH",      0) or 0)
                                _pri_aw  = float(_row.get("AWAY_ATH", 0) or 0)
                                if _pri_cl == 0 or _pri_ath == 0:
                                    continue

                                _cl_diff  = abs(_pri_cl  - _sec_cl)  / _pri_cl  * 100
                                _ath_diff = abs(_pri_ath - _sec_ath) / max(_pri_ath, _sec_ath) * 100
                                _aw_diff  = abs(_pri_aw  - _sec_aw)

                                if _cl_diff > _CLOSE_THRESH or _ath_diff > _ATH_THRESH or _aw_diff > _AWAY_THRESH:
                                    _p3 = _primary_lbl[:3]
                                    _s3 = _secondary_lbl[:3]
                                    _rows.append({
                                        "Rank":              int(_row["Rank"]),
                                        "Ticker":            _tick,
                                        f"Close_{_p3}":      round(_pri_cl, 2),
                                        f"Close_{_s3}":      round(_sec_cl, 2),
                                        "Close_Diff%":       round(_cl_diff, 2),
                                        f"ATH_{_p3}":        round(_pri_ath, 2),
                                        f"ATH_{_s3}":        round(_sec_ath, 2),
                                        "ATH_Diff%":         round(_ath_diff, 2),
                                        f"Away_{_p3}%":      round(_pri_aw, 2),
                                        f"Away_{_s3}%":      round(_sec_aw, 2),
                                        "Away_Diff_pp":      round(_aw_diff, 2),
                                        "_sec_close":        round(_sec_cl,  4),
                                        "_sec_ath":          round(_sec_ath, 4),
                                        "_sec_away":         round(_sec_aw,  4),
                                    })

                            _diff_df = pd.DataFrame(_rows)
                            st.session_state["_cross_diff_df"]         = _diff_df
                            st.session_state["_cross_primary_label"]   = _primary_lbl
                            st.session_state["_cross_secondary_label"] = _secondary_lbl

                        except Exception as _cx_e:
                            import traceback
                            _cross_error = traceback.format_exc()
                            # Store error — show to user with Skip button (do NOT auto-skip)
                            st.session_state["_cross_error"] = str(_cx_e)
                            st.session_state["_cross_error_detail"] = _cross_error
                            # Set empty diff so Step B shows the error card
                            st.session_state["_cross_diff_df"] = pd.DataFrame()

                st.rerun()  # Fresh pass to render Step B (review UI or error card)

            # ── Step B: Show review UI ──
            _diff_df   = st.session_state.get("_cross_diff_df", pd.DataFrame())
            _pri_lbl   = st.session_state.get("_cross_primary_label", _src)
            _sec_lbl   = st.session_state.get("_cross_secondary_label", "")
            _cx_err    = st.session_state.get("_cross_error", None)

            # Error card — show prominently with Skip button (no auto-rerun!)
            if _cx_err:
                st.error(f"❌ Cross-source comparison error: {_cx_err}")
                _det = st.session_state.get("_cross_error_detail","")
                if _det:
                    with st.expander("🔍 Error detail (for debugging)"):
                        st.code(_det)
                if st.button("⏭ Skip cross-source review", key="cx_err_skip", type="primary"):
                    st.session_state["_cross_review_done"] = True
                    st.session_state["_cross_diff_df"]     = None
                    st.session_state["_cross_error"]       = None
                    st.rerun()
                st.stop()

            if _diff_df is not None and not _diff_df.empty:
                _n_diff = len(_diff_df)

                st.markdown(f"""
                <div style="background:linear-gradient(135deg,#1e3a5f,#0f2942);border-radius:12px;
                            padding:16px 20px;margin-bottom:16px;border-left:4px solid #3b82f6;">
                  <div style="color:#93c5fd;font-size:13px;font-weight:700;letter-spacing:.5px;
                              text-transform:uppercase;margin-bottom:6px;">
                    🔍 Cross-Source Review — {_pri_lbl} vs {_sec_lbl}
                  </div>
                  <div style="color:#f1f5f9;font-size:22px;font-weight:800;">
                    {_n_diff} stocks mein significant difference hai
                  </div>
                  <div style="color:#94a3b8;font-size:12px;margin-top:4px;">
                    Top {st.session_state.get('_cross_top_n',400)} ranked stocks mein se |
                    Close diff &gt; 2% ya ATH diff &gt; 10% ya AWAY_ATH diff &gt; 8pp
                  </div>
                </div>
                """, unsafe_allow_html=True)

                # ── Global override buttons ──
                _col_all_pri, _col_all_sec, _col_skip = st.columns([1,1,1])
                with _col_all_pri:
                    if st.button(f"✅ Sabhi {_pri_lbl} rakho", key="cx_all_pri", use_container_width=True):
                        st.session_state["_cross_review_overrides"] = {
                            r["Ticker"]: "primary" for _, r in _diff_df.iterrows()
                        }
                        st.rerun()
                with _col_all_sec:
                    if st.button(f"🔄 Sabhi {_sec_lbl} se override karo", key="cx_all_sec", use_container_width=True, type="secondary"):
                        st.session_state["_cross_review_overrides"] = {
                            r["Ticker"]: "secondary" for _, r in _diff_df.iterrows()
                        }
                        st.rerun()
                with _col_skip:
                    if st.button("⏭ Skip — Review nahi karna", key="cx_skip", use_container_width=True):
                        st.session_state["_cross_review_done"]      = True
                        st.session_state["_cross_diff_df"]          = None
                        st.session_state["_cross_review_overrides"] = {}
                        st.rerun()

                st.markdown("---")
                st.markdown(f"**Per-stock source select karo** (default: primary = {_pri_lbl})")

                # Initialize overrides dict if needed
                _overrides = st.session_state.get("_cross_review_overrides", {})
                # Load ATH memory from file exactly once per session
                if not st.session_state.get("_ath_memory_loaded", False):
                    _loaded = _load_ath_memory()
                    if _loaded:
                        st.session_state["_ath_memory"] = _loaded
                    st.session_state["_ath_memory_loaded"] = True
                _ath_mem = st.session_state.get("_ath_memory") or {}

                # ── Display table with per-stock radio ──
                # Show display columns (hide internal _sec_* cols)
                _show_cols = [c for c in _diff_df.columns if not c.startswith("_")]
                _display_df = _diff_df[_show_cols].copy()

                # Color coding helper for ATH diff
                _ATH_CRITICAL = 50.0  # > 50% diff = almost certainly split-adjusted error

                # Render as custom table with selectbox per row
                for _ridx, _drow in _diff_df.iterrows():
                    _tick = _drow["Ticker"]
                    _rank = int(_drow["Rank"])
                    _ath_d = float(_drow["ATH_Diff%"])
                    _close_d = float(_drow["Close_Diff%"])
                    _away_d = float(_drow["Away_Diff_pp"])

                    # Severity badge
                    if _ath_d > 50:
                        _badge = "🔴 ATH Critical"
                        _badge_color = "#ef4444"
                    elif _ath_d > _ATH_CRITICAL * 0.4 or _close_d > 5:
                        _badge = "🟠 High Diff"
                        _badge_color = "#f97316"
                    else:
                        _badge = "🟡 Moderate"
                        _badge_color = "#eab308"

                    _pri_ath_col = f"ATH_{_pri_lbl[:3]}"
                    _sec_ath_col = f"ATH_{_sec_lbl[:3]}"
                    _pri_cl_col  = f"Close_{_pri_lbl[:3]}"
                    _sec_cl_col  = f"Close_{_sec_lbl[:3]}"
                    _pri_aw_col  = f"Away_{_pri_lbl[:3]}%"
                    _sec_aw_col  = f"Away_{_sec_lbl[:3]}%"

                    # Suggestion: pick source with LOWER ATH (more likely split-adjusted)
                    _suggested = "secondary" if float(_drow[_sec_ath_col]) < float(_drow[_pri_ath_col]) else "primary"
                    _suggested_lbl = _sec_lbl if _suggested == "secondary" else _pri_lbl

                    # ── ATH Memory lookup for this ticker ──────────────────
                    _mem_entry = _ath_mem.get(_tick)
                    _mem_hint_html = ""
                    _mem_role = None  # role (primary/secondary) from memory in THIS session's context
                    if _mem_entry:
                        _m_lbl  = _mem_entry.get("chosen_lbl", "")
                        _m_date = _mem_entry.get("reviewed_date", "?")
                        _m_ath  = _mem_entry.get("chosen_ath", 0)
                        # Map memory's chosen label to current session's primary/secondary roles
                        if _m_lbl == _pri_lbl:
                            _mem_role = "primary"
                        elif _m_lbl == _sec_lbl:
                            _mem_role = "secondary"
                        # else: sources swapped or different — show hint but don't force default
                        _mem_hint_html = (
                            f'&nbsp;&nbsp;&nbsp;'
                            f'<span style="background:#0f3460;border:1px solid #38bdf8;border-radius:10px;'
                            f'padding:1px 8px;font-size:11px;color:#38bdf8;">'
                            f'📚 Prev ({_m_date}): <b>{_m_lbl}</b>'
                            f'&nbsp;<span style="color:#7dd3fc;font-weight:400;">ATH {_m_ath:,.0f}</span>'
                            f'</span>'
                        )

                    # Radio default: memory > algorithmic suggestion (only if not already set this session)
                    if _tick not in _overrides:
                        _cur_sel = _mem_role if _mem_role is not None else _suggested
                    else:
                        _cur_sel = _overrides[_tick]

                    with st.container():
                        _c1, _c2 = st.columns([3, 1])
                        with _c1:
                            # st.components.v1.html — proper iframe, onclick JS executes fully
                            _card_html = f"""
                            <style>
                              body {{ margin:0; padding:0; background:transparent; font-family:'Segoe UI',sans-serif; }}
                              .cpybtn {{
                                cursor:pointer; background:transparent;
                                border:1px solid #334155; border-radius:6px;
                                padding:1px 7px; font-size:11px; color:#64748b;
                                font-family:inherit; vertical-align:middle;
                                transition:color .15s, border-color .15s;
                              }}
                              .cpybtn:hover {{ border-color:#64748b; color:#94a3b8; }}
                            </style>
                            <div style="background:#1e293b;border-radius:8px;padding:10px 14px;
                                        border:1px solid {_badge_color};border-left:3px solid {_badge_color};">
                              <span style="color:#f1f5f9;font-weight:700;font-size:15px;">
                                #{_rank} &nbsp; {_tick}
                              </span>
                              &nbsp;
                              <button class="cpybtn" onclick="
                                var sym='{_tick}';
                                navigator.clipboard.writeText(sym).then(function(){{
                                  this.innerText='\\u2713 Copied';
                                  this.style.color='#4ade80';
                                  this.style.borderColor='#4ade80';
                                  var b=this;
                                  setTimeout(function(){{
                                    b.innerText='\\u29c9 Copy';
                                    b.style.color='#64748b';
                                    b.style.borderColor='#334155';
                                  }}, 1500);
                                }}.bind(this)).catch(function(){{
                                  var ta=document.createElement('textarea');
                                  ta.value=sym; ta.style.position='fixed'; ta.style.opacity='0';
                                  document.body.appendChild(ta); ta.focus(); ta.select();
                                  document.execCommand('copy'); document.body.removeChild(ta);
                                  this.innerText='\\u2713 Copied';
                                  this.style.color='#4ade80';
                                  this.style.borderColor='#4ade80';
                                  var b=this;
                                  setTimeout(function(){{
                                    b.innerText='\\u29c9 Copy';
                                    b.style.color='#64748b';
                                    b.style.borderColor='#334155';
                                  }}, 1500);
                                }}.bind(this));
                              ">&#x29c9; Copy</button>
                              &nbsp;
                              <span style="background:{_badge_color}20;color:{_badge_color};border-radius:10px;
                                           padding:1px 8px;font-size:11px;font-weight:600;">{_badge}</span>
                              &nbsp;
                              <span style="color:#64748b;font-size:11px;">
                                &#x1f4a1; Suggested: <b style="color:#a3e635;">{_suggested_lbl}</b>
                              </span>
                              {_mem_hint_html}
                              <div style="display:flex;gap:24px;margin-top:8px;flex-wrap:wrap;">
                                <span style="color:#94a3b8;font-size:12px;">
                                  Close: <b style="color:#f1f5f9;">{_drow[_pri_cl_col]}</b> ({_pri_lbl[:3]})
                                  vs <b style="color:#f1f5f9;">{_drow[_sec_cl_col]}</b> ({_sec_lbl[:3]})
                                  &nbsp;<span style="color:#fbbf24;font-weight:700;">&Delta; {_close_d:.1f}%</span>
                                </span>
                                <span style="color:#94a3b8;font-size:12px;">
                                  ATH: <b style="color:#f1f5f9;">{_drow[_pri_ath_col]:,.0f}</b> ({_pri_lbl[:3]})
                                  vs <b style="color:#f1f5f9;">{_drow[_sec_ath_col]:,.0f}</b> ({_sec_lbl[:3]})
                                  &nbsp;<span style="color:{_badge_color};font-weight:700;">&Delta; {_ath_d:.1f}%</span>
                                </span>
                                <span style="color:#94a3b8;font-size:12px;">
                                  Away ATH: <b style="color:#f1f5f9;">{_drow[_pri_aw_col]:.1f}%</b> ({_pri_lbl[:3]})
                                  vs <b style="color:#f1f5f9;">{_drow[_sec_aw_col]:.1f}%</b> ({_sec_lbl[:3]})
                                  &nbsp;<span style="color:#a78bfa;font-weight:700;">&Delta; {_away_d:.1f}pp</span>
                                </span>
                              </div>
                            </div>"""
                            import streamlit.components.v1 as _stc
                            _stc.html(_card_html, height=115)
                        with _c2:
                            _opts    = [f"✅ {_pri_lbl}", f"🔄 {_sec_lbl}"]
                            _sel_idx = 0 if _cur_sel == "primary" else 1
                            _chosen  = st.radio(
                                f"Source for {_tick}",
                                _opts,
                                index=_sel_idx,
                                key=f"cx_radio_{_tick}",
                                label_visibility="collapsed"
                            )
                            _overrides[_tick] = "primary" if _chosen == _opts[0] else "secondary"

                st.session_state["_cross_review_overrides"] = _overrides

                st.markdown("---")
                _apply_col, _skip2_col = st.columns([1,1])
                with _apply_col:
                    if st.button("✅ Apply Overrides & Continue →", key="cx_apply",
                                 type="primary", use_container_width=True):
                        # Apply overrides to dfStats
                        _dfS_mod = st.session_state.dfStats.copy()
                        # Ensure Ticker is accessible (might be index)
                        if "Ticker" not in _dfS_mod.columns:
                            _dfS_mod = _dfS_mod.reset_index()
                        _was_indexed = "Ticker" not in st.session_state.dfStats.columns

                        _n_applied = 0
                        for _, _orow in _diff_df.iterrows():
                            _tick = _orow["Ticker"]
                            if _overrides.get(_tick, "primary") == "secondary":
                                _mask = _dfS_mod["Ticker"] == _tick
                                if _mask.any():
                                    _dfS_mod.loc[_mask, "Close"]    = float(_orow["_sec_close"])
                                    _dfS_mod.loc[_mask, "ATH"]      = float(_orow["_sec_ath"])
                                    _dfS_mod.loc[_mask, "AWAY_ATH"] = float(_orow["_sec_away"])
                                    _n_applied += 1

                        if _was_indexed:
                            _dfS_mod = _dfS_mod.set_index("Ticker")

                        # ── Save ATH override decisions to memory ──────────
                        _today_str = datetime.date.today().strftime("%d-%b-%y")
                        _ath_mem_upd = st.session_state.get("_ath_memory") or {}
                        for _, _orow in _diff_df.iterrows():
                            _tick = _orow["Ticker"]
                            _chosen_role = _overrides.get(_tick, "primary")
                            _chosen_lbl  = _pri_lbl if _chosen_role == "primary" else _sec_lbl
                            _ath_col     = f"ATH_{_pri_lbl[:3]}" if _chosen_role == "primary" else f"ATH_{_sec_lbl[:3]}"
                            _ath_mem_upd[_tick] = {
                                "chosen_lbl":   _chosen_lbl,
                                "chosen_role":  _chosen_role,
                                "chosen_ath":   float(_orow[_ath_col]),
                                "reviewed_date": _today_str,
                                "pri_lbl":      _pri_lbl,
                                "sec_lbl":      _sec_lbl,
                            }
                        st.session_state["_ath_memory"] = _ath_mem_upd
                        _mem_saved = _save_ath_memory(_ath_mem_upd)

                        # Re-apply filters with corrected data
                        _fp_reapply = st.session_state.get("_cross_filter_params") or {}
                        try:
                            _fp_reapply = st.session_state.get("_cross_filter_params") or {}
                            # If filter params empty, try to reconstruct from current dfFiltered
                            # (safe fallback: just update dfStats & rerun apply_filters with existing params)
                            _dfF_new = apply_filters(_dfS_mod.copy(), _fp_reapply)
                            st.session_state.dfStats    = _dfS_mod
                            st.session_state.dfFiltered = _dfF_new
                            _mem_info = f" | ATH memory updated ({len(_ath_mem_upd)} stocks)" if _mem_saved else ""
                            st.success(f"✅ {_n_applied} stocks ke liye {_sec_lbl} data apply hua! dfStats updated.{_mem_info}")
                        except Exception as _oe:
                            # Fallback: just save modified dfStats, keep existing dfFiltered
                            st.session_state.dfStats = _dfS_mod
                            st.warning(f"Filter re-apply failed ({_oe}), dfStats updated only.")

                        st.session_state["_cross_review_done"]      = True
                        st.session_state["_cross_diff_df"]          = None
                        st.session_state["_cross_review_overrides"] = {}
                        st.rerun()

                with _skip2_col:
                    if st.button("⏭ Skip — Primary source hi rakhna hai", key="cx_skip2",
                                 use_container_width=True):
                        st.session_state["_cross_review_done"]      = True
                        st.session_state["_cross_diff_df"]          = None
                        st.session_state["_cross_review_overrides"] = {}
                        st.rerun()

                st.stop()  # Block results until decision

            else:
                # No significant differences found
                st.success(f"✅ Cross-source check complete: {_pri_lbl} vs {_sec_lbl} — Top {st.session_state.get('_cross_top_n',400)} stocks mein koi major difference nahi mila. Primary source data use ho raha hai.")
                st.session_state["_cross_review_done"] = True
                st.session_state["_cross_diff_df"]     = None
                # Auto-proceed
                import time; time.sleep(1.5)
                st.rerun()

        # ── Display results ───────────────────────────────────────
        if st.session_state.screener_done and st.session_state.dfFiltered is not None:
            dfF = st.session_state.dfFiltered
            dfU = st.session_state.dfStats
            n_f = len(dfF); n_u = len(dfU) if dfU is not None else 0
            top_n = st.session_state.top_n_rank
            rank_col = st.session_state.ranking_method

            st.markdown(f"""<div class="metric-row">
                {metric_card("Total Screened", f"{n_u:,}")}
                {metric_card("Passed Filters", f"{n_f:,}", "green")}
                {metric_card("Top-N Universe", f"Top {top_n}", "blue")}
                {metric_card("End Date", st.session_state.lookback_date.strftime('%d %b %Y'), "amber")}
            </div>""", unsafe_allow_html=True)

            # ── Failed stocks — persistent (rerun ke baad bhi dikhta hai) ─
            _fb = st.session_state.failed_blank or []
            if _fb:
                with st.expander(f"⚠️ {len(_fb)} stocks failed to download — click to view", expanded=False):
                    st.dataframe(
                        pd.DataFrame({'S.No.': range(1, len(_fb)+1),
                                      'Failed Stock': _fb}).set_index('S.No.'),
                        use_container_width=False
                    )
            else:
                st.success("✅ All stocks downloaded successfully!")

            tab1, tab2 = st.tabs(["✅ Filtered (Top Ranked)", "📊 All Stocks (Unfiltered)"])
            with tab1:
                top_view = dfF.head(top_n).reset_index()
                dcols = ["Rank","Ticker","Close",rank_col,"roc12M","roc6M","roc3M","vol12M","volm_cr","AWAY_ATH","circuit","dma200d"]
                dcols = [c for c in dcols if c in top_view.columns]
                st.dataframe(top_view[dcols].style.format(precision=2),
                             use_container_width=True, height=440)
            with tab2:
                if dfU is not None:
                    st.dataframe(dfU.reset_index().head(300).style.format(precision=2),
                                 use_container_width=True, height=440)

            st.divider()
            if st.button("▶ Next: Plan Rebalance →", type="primary"):
                st.session_state.current_step = 3; st.rerun()

        elif not st.session_state.screener_done:
            st.info("⬆️ Upar se 'Start Data Download' click karo.")

    # ═══════════════════════════════════════════════════════════════
    # STEP 3 — PLAN REBALANCE
    # ═══════════════════════════════════════════════════════════════
    elif st.session_state.current_step == 3:
        st.markdown('<div class="section-hdr">⚖️ Step 3 — Plan Rebalance</div>', unsafe_allow_html=True)

        if not st.session_state.screener_done:
            st.warning("⚠️ Pehle Step 2 mein screener run karo.")
            if st.button("← Step 2 par jao"): st.session_state.current_step = 2; st.rerun()
            st.stop()

        # ── Portfolio source ──────────────────────────────────────
        port_source = st.radio("Portfolio data source",
                               ["📊 Google Sheet (auto)", "📂 CSV manually upload"],
                               horizontal=True)

        if "📊" in port_source:
            col_load, _ = st.columns([1, 2])
            with col_load:
                if st.button("🔄 Fetch from Google Sheet", type="primary"):
                    with st.spinner("Fetching portfolio..."):
                        try:
                            pdf = pd.read_csv(PORTFOLIO_CSV_URL)
                            if 'Current Portfolio' in pdf.columns:
                                st.session_state.reb_portfolio = [
                                    str(x).strip().upper() for x in pdf['Current Portfolio'].dropna()
                                    if str(x).strip() and str(x).strip().lower() not in ('nan','current portfolio','')
                                ]
                                st.success(f"✅ Portfolio loaded: **{len(st.session_state.reb_portfolio)}** stocks")
                            else:
                                st.error("Column 'Current Portfolio' not found in sheet.")
                        except Exception as e:
                            st.error(f"Sheet fetch failed: {e}")
        else:
            up_reb = st.file_uploader("📂 Upload Portfolio CSV", type=["csv"], key="reb_csv")
            if up_reb:
                try:
                    df_reb = pd.read_csv(up_reb)
                    df_reb.columns = [c.strip() for c in df_reb.columns]
                    col_b = df_reb.columns[1] if len(df_reb.columns) > 1 else df_reb.columns[0]
                    st.session_state.reb_portfolio = [
                        str(x).strip().upper() for x in df_reb[col_b].dropna()
                        if str(x).strip() and len(str(x).strip()) > 1
                    ]
                    st.success(f"✅ Portfolio loaded: {len(st.session_state.reb_portfolio)} stocks")
                except Exception as e:
                    st.error(f"CSV parse error: {e}")

        # ── Manual override ───────────────────────────────────────
        with st.expander("✏️ Manual Edit (comma-separated)", expanded=False):
            port_text = st.text_area("Current Portfolio",
                                      value=", ".join(st.session_state.reb_portfolio or []), height=100)
            if st.button("Apply Manual Edit"):
                st.session_state.reb_portfolio = [
                    s.strip().upper() for s in port_text.split(",") if s.strip()
                ]
                st.success("Updated!")

        portfolio = st.session_state.reb_portfolio or []
        # ── Step 3 sub-tabs ───────────────────────────────────────────
        _s3_tab_a, _s3_tab_b = st.tabs([
            "🌡️  Regime & Allocation",
            "⚖️  Rebalancer & Orders",
        ])

        with _s3_tab_a:
        # ══════════════════════════════════════════════════════════════
        # REGIME PANEL — Market Regime & Multi-Asset Allocation
        # Inserted here so dfStats is available (screener already run)
        # ══════════════════════════════════════════════════════════════

            import datetime as _dt_regime, math as _math_regime
            from calculations import get_regime_score, get_next_rebalance_dates, get_weekly_deployment_plan
            import streamlit.components.v1 as _stc_regime

            st.markdown('<div class="section-hdr">🌡️ Market Regime & Multi-Asset Allocation</div>',
                        unsafe_allow_html=True)

            # ── NAV + VIX — same approach as Regime Tab (direct fetch, no Dashboard API) ──
            _nav_series   = st.session_state.get("_regime_nav_series", [])
            _vix_curr     = st.session_state.get("_regime_vix", None)
            _weekly_nav_r = st.session_state.get("_regime_weekly_ret", None)
            _dash_loaded  = st.session_state.get("_regime_dash_loaded", False)

            _dash_col1, _dash_col2 = st.columns([3, 1])
            with _dash_col2:
                _fetch_nav = st.button("📡 Refresh NAV", key="refresh_nav_btn",
                                       help="NAV CSV + India VIX yfinance se fresh fetch karo")

            # Auto-fetch on first open OR when button pressed
            if _fetch_nav or not _dash_loaded:
                with st.spinner("📡 VIX + NAV fetch ho raha hai..."):
                    try:
                        _vix_new  = _fetch_vix_yf()
                        _nav_new  = _fetch_nav_from_sheet(_NAV_SHEET_CSV)
                        _wret_new = (round((_nav_new[-1] / _nav_new[-6] - 1) * 100, 2)
                                     if len(_nav_new) >= 6 else None)
                        if _nav_new:
                            _nav_series   = _nav_new
                            _vix_curr     = _vix_new
                            _weekly_nav_r = _wret_new
                            st.session_state["_regime_nav_series"]  = _nav_series
                            st.session_state["_regime_vix"]         = _vix_curr
                            st.session_state["_regime_weekly_ret"]  = _weekly_nav_r
                            st.session_state["_regime_dash_loaded"] = True
                            _vix_msg = f" | VIX: {_vix_curr:.1f}" if _vix_curr else " | VIX: N/A"
                            st.success(f"✅ NAV: {len(_nav_series)} pts · Latest: {_nav_series[-1]:.2f}{_vix_msg}")
                        else:
                            st.warning("⚠️ NAV fetch nahi hua — CSV URL check karo.")
                    except Exception as _ef:
                        st.warning(f"⚠️ Fetch error: {_ef}")
            else:
                _lbl_v = f"VIX {_vix_curr:.1f}" if _vix_curr else "VIX N/A"
                st.caption(f"📊 NAV {len(_nav_series)} pts | {_lbl_v} · Refresh button se update karo")

            # ── Regime score (7-signal QFSM v2026.08) ────────────────────
            from calculations import get_full_regime_result, RegimeState
            _dfS_rg   = st.session_state.get("dfStats")
            _prev_rs  = st.session_state.get("_regime_prev_state")
            _rt_nc_rg = st.session_state.get("_rt_nifty_close")
            _rt_nd_rg = st.session_state.get("_rt_nifty_dma200")
            _rt_rk_rg = st.session_state.get("_rt_rank_history", [])

            _rs_rg = get_full_regime_result(
                dfStats=_dfS_rg,
                equity_nav_series=_nav_series or None,
                vix_value=_vix_curr,
                nifty_close=_rt_nc_rg,
                nifty_dma200=_rt_nd_rg,
                rank_history=_rt_rk_rg or None,
                fii_score=0.5,
                prev_state=_prev_rs,
                total_capital=0.0,
                dd_pct=0.0,
            )
            st.session_state["_regime_state"] = _rs_rg

            _sc   = _rs_rg.effective_band
            _lbl  = _rs_rg.label()
            _eq   = _rs_rg.equity
            _gd   = _rs_rg.gold
            _cs   = _rs_rg.cash
            _sigs = _rs_rg.signals
            _smeta= _rs_rg.signal_meta

            _fc,_em = {3:("#00d09e","🟢"),2:("#38bdf8","🔵"),
                       1:("#f59e0b","🟡"),0:("#f87171","🔴")}[_sc]

            # ── Next dates + status banner ────────────────────────────────
            _dates_rg = get_next_rebalance_dates()
            _nxt_fri  = _dates_rg["next_friday"]
            _nxt_rb   = _dates_rg["next_monthly_rb"]
            _days_fri = (_nxt_fri - _dt_regime.date.today()).days

            _st_c  = {"STABLE":"#15803d","PENDING":"#d97706","CONFIRMED":"#1d4ed8",
                      "DD_OVERRIDE":"#dc2626"}.get(_rs_rg.status,"#6b7280")
            _st_bg = {"STABLE":"#dcfce7","PENDING":"#fef3c7","CONFIRMED":"#dbeafe",
                      "DD_OVERRIDE":"#fee2e2"}.get(_rs_rg.status,"#f1f5f9")

            _vix_div_rg = ""
            if _vix_curr:
                _vc_rg = "#dc2626" if _vix_curr > 20 else "#15803d"
                _vix_div_rg = (f'<div style="background:#fef3c7;border:1px solid #fcd34d;border-radius:8px;'
                               f'padding:7px 14px;font-size:12px;color:{_vc_rg};font-family:DM Mono,monospace;">'
                               f'VIX: <b>{round(_vix_curr,1)}</b>{"  🔴" if _vix_curr>20 else ""}</div>')

            st.markdown(
                f'<div style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:12px;">'
                f'<div style="background:#dbeafe;border:1px solid #93c5fd;border-radius:8px;padding:7px 14px;font-size:12px;color:#1d4ed8;font-family:DM Mono,monospace;">📅 <b>Next Friday Check:</b> {_nxt_fri.strftime("%d %b %Y")} ({_days_fri}d)</div>'
                f'<div style="background:#dcfce7;border:1px solid #86efac;border-radius:8px;padding:7px 14px;font-size:12px;color:#15803d;font-family:DM Mono,monospace;">📆 <b>Monthly RB:</b> {_nxt_rb.strftime("%d %b %Y")}</div>'
                f'<div style="background:{_st_bg};border:1px solid;border-radius:8px;padding:7px 14px;font-size:12px;font-weight:700;color:{_st_c};font-family:DM Mono,monospace;">⚡ {_rs_rg.status} ({_rs_rg.confirmation_count}/2)</div>'
                f'{_vix_div_rg}'
                f'</div>',
                unsafe_allow_html=True
            )

            # ── 7-Signal cards ─────────────────────────────────────────────
            _s1v = _sigs.get("s1_nav",0);     _s2v = _sigs.get("s2_breadth",0)
            _s3v = _sigs.get("s3_roc",0);     _s4v = _sigs.get("s4_vix",0)
            _s5v = _sigs.get("s5_nifty",0);   _s6v = _sigs.get("s6_ad",0)
            _s7v = _sigs.get("s7_rank",0)

            _nav_txt = (f'NAV {_smeta["nav_current"]:.2f} vs DMA {_smeta["nav_dma200"]:.2f} ({_smeta.get("gap_pct",0):+.1f}%)'
                        if _smeta.get("nav_current") else "NAV data loading...")
            _nif_txt = (f'Nifty {_rt_nc_rg:,.0f} vs DMA {_rt_nd_rg:,.0f} (ratio {_smeta.get("nifty_ratio",1):.3f})'
                        if _rt_nc_rg else "Nifty: loading (Market Regime tab refresh karo)")

            def _mk_sig_rg(icon, title, sub, val_txt, score_val, weight, ok, partial=False):
                c  = "#f59e0b" if partial else ("#00d09e" if ok else "#f87171")
                bg = "#fef3c7" if partial else ("#e8fdf2" if ok else "#fef2f2")
                bd = "#fcd34d" if partial else ("#86efac" if ok else "#fca5a5")
                return (f'<div class="sig" style="background:{bg};border-color:{bd}">' +
                        f'<div class="si">{icon}</div>' +
                        f'<div class="sb"><div class="st" style="color:{c}">{title}</div>' +
                        f'<div class="sc_t" style="color:#374151">{sub}</div>' +
                        f'<div class="sv" style="color:{c}">{val_txt}</div></div>' +
                        f'<div class="sbg" style="background:{c};color:white">{score_val:.2g}/{weight}pt</div>' +
                        '</div>')

            _date_str_rg = _dt_regime.date.today().strftime('%d %b %Y')
            _gauge_html  = _build_mmi_gauge(_sc, _fc, _lbl, _em, "", _date_str_rg)
            _sig_html_rg = (
                '<style>body{margin:0;padding:0;background:transparent;font-family:"Segoe UI",sans-serif;}' +
                '.sigs{display:flex;flex-direction:column;gap:8px;}' +
                '.sig{border-radius:10px;padding:11px 14px;border:1.5px solid;display:flex;align-items:center;gap:12px;}' +
                '.si{font-size:20px;flex-shrink:0;}.sb{flex:1;}' +
                '.st{font-size:11px;font-weight:700;letter-spacing:.4px;text-transform:uppercase;margin-bottom:2px;}' +
                '.sc_t{font-size:11px;margin-bottom:3px;}.sv{font-size:13px;font-weight:800;}' +
                '.sbg{padding:3px 10px;border-radius:20px;font-size:11px;font-weight:700;flex-shrink:0;}' +
                '</style><div class="sigs">' +
                _mk_sig_rg("✅" if _s1v>0 else "❌","S1 — Equity Curve Trend","NAV > 200-Day Moving Average",_nav_txt,_s1v,1.5,_s1v>0) +
                _mk_sig_rg("✅" if _s2v>0 else "❌","S2 — Market Breadth","% Stocks > 200DMA > 50%",f'{_smeta.get("breadth_pct",0):.1f}% above DMA',_s2v,1.0,_s2v>0) +
                _mk_sig_rg("✅" if _s3v>0 else "❌","S3 — Universe Momentum","Median 3M ROC > 0%",f'{_smeta.get("median_roc3m",0):+.1f}% median 3M ROC',_s3v,1.0,_s3v>0) +
                _mk_sig_rg("⚠️" if 0<_s4v<1.5 else ("✅" if _s4v>0 else "❌"),"S4 — India VIX ★","VIX ≤ 20=PASS | 20-25=Partial | >25=FAIL",f'VIX {round(_vix_curr,1) if _vix_curr else "N/A"}',_s4v,1.5,_s4v>0,0<_s4v<1.5) +
                _mk_sig_rg("⚠️" if 0<_s5v<1.5 else ("✅" if _s5v>0 else "❌"),"S5 — Nifty 200DMA ★","^NSEI Close > 200-Day Moving Average",_nif_txt,_s5v,1.5,_s5v>0,0<_s5v<1.5) +
                _mk_sig_rg("✅" if _s6v>0 else "❌","S6 — A-D Ratio ★","Advances/Total > 45% (1M ROC)",f'A-D: {_smeta.get("ad_ratio",0.5)*100:.0f}%',_s6v,1.0,_s6v>0) +
                _mk_sig_rg("✅" if _s7v>0 else "❌","S7 — Rank Stability ★","Top-50 overlap > 60% (4-week)",f'Overlap: {_smeta.get("rank_overlap_pct",65):.0f}%',_s7v,0.5,_s7v>0) +
                '</div>'
            )

            _g_col, _s_col = st.columns([1, 1.6])
            with _g_col:
                _stc_regime.html(_gauge_html, height=360)
                # Score X.XX/8.5 below gauge
                _scbg2 = {3:"#dcfce7",2:"#dbeafe",1:"#fef3c7",0:"#fee2e2"}[_sc]
                _scfc2 = {3:"#15803d",2:"#1d4ed8",1:"#d97706",0:"#dc2626"}[_sc]
                st.markdown(
                    f'<div style="background:{_scbg2};border-radius:8px;padding:8px;text-align:center;margin-top:4px;">' +
                    f'<div style="font-size:10px;color:{_scfc2};font-weight:700;text-transform:uppercase;">Weighted Score</div>' +
                    f'<div style="font-size:32px;font-weight:900;color:{_scfc2};line-height:1.1;">{_rs_rg.raw_score:.2f}</div>' +
                    f'<div style="font-size:10px;color:{_scfc2};">/ 8.5 pts · {_lbl}</div>' +
                    '</div>', unsafe_allow_html=True)

                # QFSM mode badge
                _qc2  = "#7c3aed" if _rs_rg.qfsm_mode=="BLEND" else "#15803d"
                _qbg2 = "#ede9fe" if _rs_rg.qfsm_mode=="BLEND" else "#dcfce7"
                st.markdown(
                    f'<div style="background:{_qbg2};border:1px solid {_qc2};border-radius:6px;padding:5px 10px;font-size:11px;color:{_qc2};font-weight:700;margin-top:6px;text-align:center;">' +
                    f'{"⚛ QFSM BLEND" if _rs_rg.qfsm_mode=="BLEND" else "✅ Standard Band"} {_rs_rg.effective_band}' +
                    '</div>', unsafe_allow_html=True)
            with _s_col:
                _stc_regime.html(_sig_html_rg, height=580)

            # VIX overlay notice
            if _vix_curr is not None and _rs_rg.vix_overlay_pct > 0:
                _vc3 = "#dc2626" if _vix_curr > 30 else "#d97706"
                st.markdown(
                    f'<div style="background:{"#fef2f2" if _vix_curr>30 else "#fef3c7"};border:1.5px solid {_vc3};border-left:4px solid {_vc3};border-radius:8px;padding:8px 14px;font-size:12px;color:{_vc3};margin-bottom:8px;">' +
                    f'⚡ <b>VIX Overlay (VIX {_vix_curr:.1f}):</b> +{_rs_rg.vix_overlay_pct:.1f}pp Gold (Liquid → Gold) | Equity UNTOUCHED' +
                    '</div>', unsafe_allow_html=True)

            st.markdown("---")

            # ── Portfolio value input ─────────────────────────────────────
            _pv1, _pv2 = st.columns(2)
            with _pv1:
                _total_pf = st.number_input("💼 Total Portfolio Value ₹ (Equity + Gold + Cash)",
                                             min_value=0, value=int(st.session_state.get("regime_pf_val",1000000)),
                                             step=10000, key="regime_pf_val")
            with _pv2:
                _prev_sc_input = st.number_input("📅 Pichle Mahine Ka Band (0-3)",
                                                  min_value=0, max_value=3,
                                                  value=int(st.session_state.get("regime_prev_score", _sc)),
                                                  step=1, key="regime_prev_score")
                _prev_sc = _prev_sc_input

            # ── Allocation cards (QFSM values) ────────────────────────────
            _a1,_a2,_a3 = st.columns(3)
            for col,(lbl,pct,fc,bg) in zip([_a1,_a2,_a3],[
                ("📈 Equity",_eq,"#2563eb","#dbeafe"),
                ("🥇 GOLDBEES",_gd,"#b45309","#fef3c7"),
                ("💵 Liquid Fund",_cs,"#475569","#f1f5f9")]):
                with col:
                    st.markdown(f'<div style="background:{bg};border:1px solid {fc};border-radius:8px;' +
                                f'padding:12px;text-align:center;margin-bottom:8px;">' +
                                f'<div style="font-size:11px;color:{fc};margin-bottom:4px">{lbl}</div>' +
                                f'<div style="font-size:28px;font-weight:800;color:{fc}">{pct*100:.1f}%</div>' +
                                f'<div style="font-size:12px;color:{fc};opacity:.8">₹{_total_pf*pct:,.0f}</div>' +
                                '</div>', unsafe_allow_html=True)

            # ── Shift message ─────────────────────────────────────────────
            _sc_diff = _sc - _prev_sc

        # Persist for Order Calculator (compatible dict from RegimeState)
        st.session_state["_regime_result"]   = {
            "score":  _rs_rg.effective_band,
            "label":  _rs_rg.label(),
            "equity": _rs_rg.equity,
            "gold":   _rs_rg.gold,
            "cash":   _rs_rg.cash,
        }
        st.session_state["_regime_prev_sc"]  = int(_prev_sc)
        st.session_state["_regime_total_pf"] = float(_total_pf)
        if _sc_diff == 0:
            _smsg,_sfc,_sbg = "✅ Score same — normal equity rebalance karo. GOLDBEES/Liquid drift ±7% check karo. New entries at new target weight (Eq Budget ÷ 30).","#15803d","#dcfce7"
        elif abs(_sc_diff) == 1:
            _smsg,_sfc,_sbg = f"🔄 Minor shift ({_prev_sc}→{_sc}) — exits se Gold/Liquid fund karo, new entries at new target weight. Existing stocks drift band mein rahenge.","#1d4ed8","#dbeafe"
        else:
            _smsg,_sfc,_sbg = f"⚠️ Major shift ({_prev_sc}→{_sc}) — phased 2-month plan. Monthly exits se Gold/Liquid fund karo. Weekly plan neeche dekho.","#b45309","#fef3c7"
            st.markdown(f"""<div style="background:{_sbg};border:1px solid {_sfc};border-left:4px solid {_sfc};
                        border-radius:8px;padding:10px 14px;font-size:13px;color:{_sfc};margin:8px 0">
              {_smsg}</div>""", unsafe_allow_html=True)

            # ── GOLDBEES + Liquid actions — aligned layout ───────────────
            # Section header
            st.markdown("""<div style="font-size:14px;font-weight:700;color:var(--text-color);
                            border-left:4px solid #0ea5e9;padding:6px 0 6px 12px;
                            background:linear-gradient(90deg,rgba(14,165,233,.06) 0%,transparent 60%);
                            border-radius:0 6px 6px 0;margin:1rem 0 .8rem;">
                🏦 Asset Actions
            </div>""", unsafe_allow_html=True)

            _act_col1, _act_col2 = st.columns(2)

            # ─── GOLDBEES ───────────────────────────────────────────────────
            with _act_col1:
                st.markdown('<div style="font-size:13px;font-weight:700;color:#b45309;letter-spacing:.3px;margin-bottom:10px;padding-bottom:6px;border-bottom:2px solid #fcd34d">🥇 GOLDBEES Action</div>', unsafe_allow_html=True)
                _gb_curr = st.number_input("Current GOLDBEES ₹", min_value=0,
                                            value=0, step=1000, key="goldbees_curr_val",
                                            label_visibility="visible")
                _gb_cmp  = st.number_input("GOLDBEES CMP ₹", min_value=0.0,
                                            value=0.0, step=0.5, key="goldbees_cmp",
                                            label_visibility="visible")
                if _total_pf > 0:
                    _gd_tgt = _total_pf * _gd
                    _gd_dif = _gd_tgt - _gb_curr
                    _gdok   = abs(_gd_dif) / _total_pf < 0.07
                    _gdc    = "#15803d" if _gdok else ("#b45309" if abs(_gd_dif/_total_pf) < 0.15 else "#dc2626")
                    _gd_bg  = "#dcfce7" if _gdok else ("#fef3c7" if abs(_gd_dif/_total_pf) < 0.15 else "#fee2e2")
                    _gu_txt = f" (~{int(abs(_gd_dif)/_gb_cmp)} units)" if not _gdok and _gb_cmp > 0 else ""
                    _gact_icon = "✅" if _gdok else ("🔺" if _gd_dif > 0 else "🔻")
                    # ₹15K min guardrail — sub-₹15K transaction not worth brokerage
                    _gd_min_ok = abs(_gd_dif) >= 15000
                    _gact_txt  = "Hold (within ±7%)" if _gdok else (
                        f"BUY ₹{abs(_gd_dif):,.0f}{_gu_txt}" if _gd_dif > 0 else f"SELL ₹{abs(_gd_dif):,.0f}{_gu_txt}"
                    )
                    if not _gdok and not _gd_min_ok:
                        _gact_txt = f"⏭ Skip (< ₹15K threshold) — drift ₹{abs(_gd_dif):,.0f}"
                        _gdc, _gd_bg = "#64748b", "#f1f5f9"
                    # VIX overlay adjusted gold target
                    _vix_ovl_pct = 0
                    if _vix_curr and _vix_curr > 30: _vix_ovl_pct = 5
                    elif _vix_curr and _vix_curr > 20: _vix_ovl_pct = 3
                    _eff_gd_pct = min(_gd + _vix_ovl_pct/100, 0.30)
                    _gd_tgt_eff = _total_pf * _eff_gd_pct
                    _ovl_note   = f" (VIX +{_vix_ovl_pct}%)" if _vix_ovl_pct > 0 else ""
                    st.markdown(f"""<div style="background:{_gd_bg};border:1px solid {_gdc};
                            border-radius:10px;padding:13px 15px;margin-top:6px;">
                      <div style="display:flex;justify-content:space-between;align-items:center;
                                  margin-bottom:10px;font-size:13px;font-weight:600;color:{_gdc};">
                        <span>Current: <b style="font-size:14px;">₹{_gb_curr:,.0f}</b></span>
                        <span style="font-size:16px;opacity:.4">→</span>
                        <span>Target: <b style="font-size:14px;">₹{_gd_tgt_eff:,.0f}</b>{_ovl_note}</span>
                      </div>
                      <div style="font-size:17px;font-weight:800;color:{_gdc};text-align:center;">
                        {_gact_icon} {_gact_txt}
                      </div>
                      <div style="font-size:10.5px;color:#64748b;margin-top:6px;text-align:center;">
                        Band: ±7% of PF · Min txn ₹15K · {'Drift ₹' + f"{abs(_gd_dif):,.0f}" if not _gdok else 'Within band'}
                      </div>
                    </div>""", unsafe_allow_html=True)

            # ─── LIQUID FUND ─────────────────────────────────────────────────
            with _act_col2:
                st.markdown('<div style="font-size:13px;font-weight:700;color:#475569;letter-spacing:.3px;margin-bottom:10px;padding-bottom:6px;border-bottom:2px solid #cbd5e1">💵 Liquid Fund Action</div>', unsafe_allow_html=True)
                _lf_curr = st.number_input("Current Liquid Fund ₹", min_value=0,
                                            value=0, step=1000, key="liquid_curr_val",
                                            label_visibility="visible")
                if _total_pf > 0:
                    _cs_tgt = _total_pf * _cs
                    _cs_dif = _cs_tgt - _lf_curr
                    _csok   = abs(_cs_dif) / _total_pf < 0.07
                    _csc    = "#15803d" if _csok else ("#1d4ed8" if _cs_dif > 0 else "#b45309")
                    _cs_bg  = "#dcfce7" if _csok else ("#dbeafe" if _cs_dif > 0 else "#fef3c7")
                    _cs_icon = "✅" if _csok else ("🔺" if _cs_dif > 0 else "🔻")
                    # ₹15K min guardrail
                    _cs_min_ok = abs(_cs_dif) >= 15000
                    _cs_txt  = "Hold (within ±7%)" if _csok else (
                        f"ADD ₹{abs(_cs_dif):,.0f}" if _cs_dif > 0 else f"REDEEM ₹{abs(_cs_dif):,.0f}"
                    )
                    if not _csok and not _cs_min_ok:
                        _cs_txt = f"⏭ Skip (< ₹15K threshold) — drift ₹{abs(_cs_dif):,.0f}"
                        _csc, _cs_bg = "#64748b", "#f1f5f9"
                    st.markdown(f"""<div style="background:{_cs_bg};border:1px solid {_csc};
                            border-radius:10px;padding:13px 15px;margin-top:89px;">
                      <div style="display:flex;justify-content:space-between;align-items:center;
                                  margin-bottom:10px;font-size:13px;font-weight:600;color:{_csc};">
                        <span>Current: <b style="font-size:14px;">₹{_lf_curr:,.0f}</b></span>
                        <span style="font-size:16px;opacity:.4">→</span>
                        <span>Target: <b style="font-size:14px;">₹{_cs_tgt:,.0f} ({_cs*100:.0f}%)</b></span>
                      </div>
                      <div style="font-size:17px;font-weight:800;color:{_csc};text-align:center;">
                        {_cs_icon} {_cs_txt}
                      </div>
                      <div style="font-size:10.5px;color:#64748b;margin-top:6px;text-align:center;">
                        Band: ±7% of PF · Min txn ₹15K · {'Drift ₹' + f"{abs(_cs_dif):,.0f}" if not _csok else 'Within band'}
                      </div>
                    </div>""", unsafe_allow_html=True)

            # ── Equity budget ─────────────────────────────────────────────
            _eq_budget = _total_pf * _eq
            _per_stock_target = _eq_budget / 30 if _eq_budget > 0 else 0
            _drift_band_rs = 20000  # ±₹20K per stock drift band (SOP 9.5.1)
            st.markdown(f"""<div style="background:#dbeafe;border:1px solid #93c5fd;border-left:4px solid #2563eb;
                        border-radius:8px;padding:10px 16px;font-size:13px;margin:10px 0;">
              <b style="color:#1d4ed8">📈 Equity Budget:</b>
              <span style="color:#1e3a5f;margin-left:8px;">₹{_total_pf:,.0f} × {_eq*100:.0f}% =
                <b style="font-size:16px;color:#1d4ed8"> ₹{_eq_budget:,.0f}</b>
              </span>
              &nbsp;&nbsp;
              <span style="color:#475569;font-size:12px;">
                | Per stock target: <b style="color:#1d4ed8">₹{_per_stock_target:,.0f}</b>
                &nbsp;| Drift band: <b>±₹{_drift_band_rs:,}</b>
                &nbsp;| Band Low: ₹{max(0,_per_stock_target-_drift_band_rs):,.0f}
                — High: ₹{_per_stock_target+_drift_band_rs:,.0f}
              </span>
            </div>""", unsafe_allow_html=True)

            # ── VIX Overlay Panel (SOP Section 7.5) ──────────────────────
            if _vix_curr is not None and _total_pf > 0:
                _base_gold_pct = _gd * 100
                _vix_overlay_pct = 0
                _vix_overlay_src = ""
                if _vix_curr > 30:
                    _vix_overlay_pct = 5
                    _vix_overlay_src = "VIX > 30"
                elif _vix_curr > 20:
                    _vix_overlay_pct = 3
                    _vix_overlay_src = "VIX 20-30"

                if _vix_overlay_pct > 0:
                    _eff_gold_pct = min(_base_gold_pct + _vix_overlay_pct, 30)  # hard cap 30%
                    _actual_overlay = _eff_gold_pct - _base_gold_pct
                    _eff_cash_pct   = (_cs * 100) - _actual_overlay  # liquid funds the shift
                    _eff_gold_rs    = _total_pf * _eff_gold_pct / 100
                    _eff_cash_rs    = _total_pf * _eff_cash_pct / 100
                    _overlay_rs     = _total_pf * _actual_overlay / 100
                    _vix_col = "#dc2626" if _vix_curr > 30 else "#d97706"
                    _vix_bg  = "#fef2f2" if _vix_curr > 30 else "#fef3c7"
                    st.markdown(f"""
                    <div style="background:{_vix_bg};border:1.5px solid {_vix_col};border-left:4px solid {_vix_col};
                                border-radius:8px;padding:12px 16px;margin:8px 0;">
                      <div style="font-size:13px;font-weight:700;color:{_vix_col};margin-bottom:6px;">
                        ⚡ VIX Overlay Active — {_vix_overlay_src} (+{_actual_overlay:.0f}% Gold from Liquid)
                      </div>
                      <div style="display:flex;gap:20px;flex-wrap:wrap;font-size:12px;color:#374151;">
                        <span>Base Gold: <b>{_base_gold_pct:.0f}%</b></span>
                        <span style="color:{_vix_col};">→ Effective Gold: <b>{_eff_gold_pct:.0f}%</b> (₹{_eff_gold_rs:,.0f})</span>
                        <span>Effective Cash: <b>{_eff_cash_pct:.0f}%</b> (₹{_eff_cash_rs:,.0f})</span>
                        <span style="color:#6d28d9;font-weight:600;">Move ₹{_overlay_rs:,.0f} from Liquid → GOLDBEES</span>
                      </div>
                      <div style="font-size:11px;color:#6b7280;margin-top:6px;">
                        ⚠️ Equity UNTOUCHED — only Liquid → Gold shift. Apply at monthly RB (VIX 20-30) or this Friday (VIX > 30).
                        Normalize hone pe (VIX ≤ 20) → excess Gold wapas Liquid mein.
                      </div>
                    </div>
                    """, unsafe_allow_html=True)
                else:
                    st.markdown(f"""
                    <div style="background:#f0fdf4;border:1px solid #86efac;border-radius:8px;
                                padding:8px 14px;font-size:12px;color:#15803d;margin:4px 0;">
                      ✅ VIX {_vix_curr:.1f} ≤ 20 — No VIX Overlay. Base allocation applies.
                    </div>""", unsafe_allow_html=True)

            # ── Drawdown Protocol (SOP Section 8.5) ──────────────────────
            st.markdown("""<div style="font-size:14px;font-weight:700;color:var(--text-color);
                            border-left:4px solid #dc2626;padding:6px 0 6px 12px;
                            background:linear-gradient(90deg,rgba(220,38,38,.06) 0%,transparent 60%);
                            border-radius:0 6px 6px 0;margin:1rem 0 .6rem;">
                📉 Portfolio Drawdown Protocol
            </div>""", unsafe_allow_html=True)
            _dd_c1, _dd_c2 = st.columns(2)
            with _dd_c1:
                _pf_ath = st.number_input("📈 Portfolio ATH Value ₹ (All-Time High)",
                                           min_value=0, value=int(st.session_state.get("_pf_ath_val", _total_pf or 1000000)),
                                           step=10000, key="_pf_ath_val",
                                           help="Apne portfolio ka highest value — NAV sheet se dekho")
            with _dd_c2:
                _pf_curr_dd = st.number_input("💼 Current Portfolio Value ₹",
                                               min_value=0, value=int(_total_pf),
                                               step=10000, key="_pf_curr_dd_val",
                                               help="Aaj ki total value (Equity + Gold + Cash)")

            if _pf_ath > 0 and _pf_curr_dd > 0:
                _dd_pct = (_pf_curr_dd / _pf_ath - 1) * 100
                _dd_abs  = abs(_dd_pct)
                if _dd_pct >= 0:
                    st.markdown(f"""<div style="background:#dcfce7;border:1px solid #86efac;border-radius:8px;
                                padding:8px 14px;font-size:13px;color:#15803d;">
                      ✅ Portfolio ATH pe ya upar hai — DD: <b>{_dd_pct:+.1f}%</b>. No override needed.
                    </div>""", unsafe_allow_html=True)
                elif _dd_abs < 15:
                    st.markdown(f"""<div style="background:#f0fdf4;border:1px solid #86efac;border-radius:8px;
                                padding:8px 14px;font-size:13px;color:#15803d;">
                      ✅ DD: <b>{_dd_pct:.1f}%</b> — Normal range (< 15%). Strategy as usual.
                    </div>""", unsafe_allow_html=True)
                elif _dd_abs < 20:
                    st.warning(f"⚠️ DD: **{_dd_pct:.1f}%** — DD ≥ 15% zone. Weekly check mandatory. Capital additions pause karo.")
                elif _dd_abs < 30:
                    st.error(f"🚨 DD Override TRIGGERED — DD: **{_dd_pct:.1f}%** ≥ 20% from ATH!")
                    st.markdown(f"""<div style="background:#fef2f2;border:1.5px solid #dc2626;border-radius:8px;
                                padding:12px 16px;font-size:13px;color:#dc2626;margin:4px 0;">
                      <b>⚠️ DD Override Active:</b> Current signal Score {_sc} ignored.
                      Treat as Score 0 (Bear). Target: Equity 25% | Gold 30% | Cash 45%.<br>
                      <span style="font-size:12px;color:#7f1d1d;">
                      4-week defensive shift shuru karo. Equity ₹{_total_pf*0.25:,.0f} | Gold ₹{_total_pf*0.30:,.0f} | Cash ₹{_total_pf*0.45:,.0f}
                      </span>
                    </div>""", unsafe_allow_html=True)
                else:
                    st.error(f"🚨🚨 EMERGENCY — DD: **{_dd_pct:.1f}%** ≥ 30% from ATH!")
                    st.markdown(f"""<div style="background:#fef2f2;border:2px solid #7f1d1d;border-radius:8px;
                                padding:12px 16px;font-size:13px;color:#7f1d1d;">
                      <b>🆘 Emergency Protocol:</b> Single-week move. Target: Equity 20% | Gold 30% | Cash 50%.<br>
                      Equity ₹{_total_pf*0.20:,.0f} | Gold ₹{_total_pf*0.30:,.0f} | Cash ₹{_total_pf*0.50:,.0f}
                    </div>""", unsafe_allow_html=True)

            # ── Equiweight Maintenance (SOP Section 9.5) ──────────────────
            st.markdown("""<div style="font-size:14px;font-weight:700;color:var(--text-color);
                            border-left:4px solid #7c3aed;padding:6px 0 6px 12px;
                            background:linear-gradient(90deg,rgba(124,58,237,.06) 0%,transparent 60%);
                            border-radius:0 6px 6px 0;margin:1rem 0 .6rem;">
                ⚖️ Equiweight Maintenance — Exit-Funded Regime Shift
            </div>""", unsafe_allow_html=True)
            _ew_c1, _ew_c2, _ew_c3 = st.columns(3)
            with _ew_c1:
                _exit_proceeds = st.number_input("💰 Exit Proceeds ₹ (sells se mila)",
                                                  min_value=0, value=0, step=1000, key="_exit_proceeds_val",
                                                  help="Is mahine ke exits ki total sell value")
            with _ew_c2:
                _n_new_entries = st.number_input("🟢 New Entries Count", min_value=0, max_value=30,
                                                  value=0, step=1, key="_n_new_entries_val",
                                                  help="Kitne naye stocks buy karne hain")
            with _ew_c3:
                st.markdown(f"""<div style="background:#ede9fe;border:1px solid #a78bfa;border-radius:8px;
                            padding:10px 12px;text-align:center;margin-top:4px;">
                  <div style="font-size:10px;color:#6d28d9;font-weight:600;text-transform:uppercase;">Per Stock Target</div>
                  <div style="font-size:22px;font-weight:800;color:#6d28d9;">₹{_per_stock_target:,.0f}</div>
                  <div style="font-size:10px;color:#7c3aed;">Eq Budget ÷ 30</div>
                </div>""", unsafe_allow_html=True)

            if _exit_proceeds > 0 and _total_pf > 0:
                # VIX overlay adjusted gold gap
                _vix_adj_gold_pct = min(_gd + (_vix_overlay_pct/100 if '_vix_overlay_pct' in dir() else 0), 0.30)
                _gold_gap    = max(0, _total_pf * _vix_adj_gold_pct - (_gb_curr if '_gb_curr' in dir() else 0))
                _liquid_gap  = max(0, _total_pf * _cs - (_lf_curr if '_lf_curr' in dir() else 0))
                _proceeds_after_gold  = max(0, _exit_proceeds - _gold_gap)
                _proceeds_after_liq   = max(0, _proceeds_after_gold - _liquid_gap)
                _new_entry_cost       = _n_new_entries * _per_stock_target if _per_stock_target > 0 else 0
                _surplus              = _proceeds_after_liq - _new_entry_cost

                st.markdown(f"""
                <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:10px;
                            padding:14px 16px;margin-top:8px;">
                  <div style="font-size:13px;font-weight:700;color:#0f172a;margin-bottom:10px;">
                    📊 Proceeds Allocation Plan
                  </div>
                  <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:10px;">
                    <div style="background:#dbeafe;border-radius:8px;padding:10px;text-align:center;">
                      <div style="font-size:10px;color:#1d4ed8;font-weight:600">EXIT PROCEEDS</div>
                      <div style="font-size:18px;font-weight:800;color:#1d4ed8">₹{_exit_proceeds:,.0f}</div>
                    </div>
                    <div style="background:#{'fef3c7' if _gold_gap>0 else 'f0fdf4'};border-radius:8px;padding:10px;text-align:center;">
                      <div style="font-size:10px;color:#{'b45309' if _gold_gap>0 else '15803d'};font-weight:600">1. GOLD GAP (first)</div>
                      <div style="font-size:18px;font-weight:800;color:#{'b45309' if _gold_gap>0 else '15803d'}">₹{_gold_gap:,.0f}</div>
                    </div>
                    <div style="background:#{'dbeafe' if _liquid_gap>0 else 'f0fdf4'};border-radius:8px;padding:10px;text-align:center;">
                      <div style="font-size:10px;color:#{'1d4ed8' if _liquid_gap>0 else '15803d'};font-weight:600">2. LIQUID GAP</div>
                      <div style="font-size:18px;font-weight:800;color:#{'1d4ed8' if _liquid_gap>0 else '15803d'}">₹{_liquid_gap:,.0f}</div>
                    </div>
                    <div style="background:#dcfce7;border-radius:8px;padding:10px;text-align:center;">
                      <div style="font-size:10px;color:#15803d;font-weight:600">3. EQUITY ENTRIES ({_n_new_entries} × ₹{_per_stock_target:,.0f})</div>
                      <div style="font-size:18px;font-weight:800;color:#15803d">₹{_new_entry_cost:,.0f}</div>
                    </div>
                    <div style="background:#{'dcfce7' if _surplus>=0 else 'fee2e2'};border-radius:8px;padding:10px;text-align:center;">
                      <div style="font-size:10px;color:#{'15803d' if _surplus>=0 else 'dc2626'};font-weight:600">SURPLUS / SHORTFALL</div>
                      <div style="font-size:18px;font-weight:800;color:#{'15803d' if _surplus>=0 else 'dc2626'}">₹{_surplus:+,.0f}</div>
                    </div>
                  </div>
                  <div style="margin-top:10px;font-size:12px;color:#64748b;background:#f1f5f9;border-radius:6px;padding:8px 12px;">
                    {'✅ Surplus ₹' + f"{_surplus:,.0f}" + ' → most underweight existing stock mein daalo (1 extra transaction max).'
                     if _surplus > 15000 else
                     ('⚠️ Shortfall ₹' + f"{abs(_surplus):,.0f}" + ' → partial entry ya ek stock entry defer karo. Agla month complete hoga.'
                     if _surplus < -5000 else
                     '✅ Proceeds approximately match. Clean rebalance.')}
                  </div>
                </div>
                """, unsafe_allow_html=True)
                st.caption("Rule: Gold/Liquid gap fund karo FIRST. Existing stocks drift band ±₹20K — sirf surplus ho tabhi correct karo. Standalone sell for equiweight KABHI nahi.")

            st.markdown("---")

            # ── Weekly Deployment Plan ────────────────────────────────────
            if _sc_diff != 0 and _total_pf > 0:
                st.markdown("### 📅 Weekly Deployment Plan")
                _plan = get_weekly_deployment_plan(
                    prev_score=_prev_sc, curr_score=_sc, total_pf=_total_pf,
                    goldbees_curr=_gb_curr if "_gb_curr" in dir() else 0,
                    liquid_curr=_lf_curr if "_lf_curr" in dir() else 0,
                    weekly_nav_ret=_weekly_nav_r, vix_curr=_vix_curr
                )
                if _plan["paused"]:
                    st.error(f"⏸ Week 1 PAUSED — VIX {_vix_curr:.1f} > 30 AND Weekly return {_weekly_nav_r:.1f}% < -5%. Next Friday check karo.")
                elif _plan["is_recovery"]:
                    st.info(f"🔺 Recovery mode — {_plan['n_weeks']} weeks. Faster deploy.")
                else:
                    st.info(f"🔻 Defensive mode — {_plan['n_weeks']} weeks. Gradual reduce.")
                st.markdown(f"""<div style="background:#fef3c7;border:1px solid #fcd34d;border-left:3px solid #b45309;
                            border-radius:6px;padding:8px 14px;font-size:12px;color:#78350f;margin-bottom:8px;">
                  {_plan["accelerate_msg"]}</div>""", unsafe_allow_html=True)
                _fri_lst = _dates_rg["upcoming_fridays"]
                _prev_eq = _plan["weeks"][0]["eq_pct"]  # will compare against previous week
                _prev_gd = _plan["weeks"][0]["gd_pct"]
                _prev_cs = _plan["weeks"][0]["cs_pct"]
                # Starting point = prev_score allocation
                _alloc_start = {3:(80,15,5),2:(65,20,15),1:(45,25,30),0:(25,30,45)}
                _st_e,_st_g,_st_c = _alloc_start.get(_prev_sc,(65,20,15))

                _wk_rows = []
                for _wi, wd in enumerate(_plan["weeks"]):
                    fri_lbl = _fri_lst[wd["week"]-1].strftime("%d %b") if wd["week"]-1 < len(_fri_lst) else f"Wk{wd['week']}"
                    # Compare vs starting allocation (not previous week)
                    _de = wd["eq_pct"] - _st_e
                    _dg = wd["gd_pct"] - _st_g
                    _dc = wd["cs_pct"] - _st_c
                    # Arrow indicators
                    def _arrow(delta): return "" if abs(delta)<0.1 else ("▲ " if delta>0 else "▼ ")
                    def _sign(v): return f"+{v:.1f}" if v>0 else f"{v:.1f}" if v<0 else "0"
                    _wk_rows.append({
                        "Check" : fri_lbl,
                        "Eq %"  : f"{_arrow(_de)}{wd['eq_pct']}% ({_sign(_de)}pp)",
                        "Gold %": f"{_arrow(_dg)}{wd['gd_pct']}% ({_sign(_dg)}pp)",
                        "Cash %": f"{_arrow(_dc)}{wd['cs_pct']}% ({_sign(_dc)}pp)",
                        "Eq ₹"  : f"₹{wd['eq_val']:,.0f}",
                        "Gold ₹" : f"₹{wd['gd_val']:,.0f}",
                        "Cash ₹" : f"₹{wd['cs_val']:,.0f}",
                        "Action" : wd["action"],
                    })
                st.dataframe(pd.DataFrame(_wk_rows), use_container_width=True, hide_index=True)
                st.markdown("---")

            # ── Weekly Switch Alerts ──────────────────────────────────────
            _w_alerts = []
            if _weekly_nav_r is not None and _weekly_nav_r < -5.0:
                _w_alerts.append(f"📉 Weekly NAV return = **{_weekly_nav_r:.1f}%** (< -5%)")
            if _vix_curr is not None and _vix_curr > 30:
                _w_alerts.append(f"😱 India VIX = **{_vix_curr:.1f}** (> 30)")
            _dfF_check = st.session_state.get("dfFiltered")
            if _dfF_check is not None and len(_dfF_check) < 30:
                _w_alerts.append(f"📊 Qualifying stocks = **{len(_dfF_check)}** (< 30)")
            if _w_alerts:
                st.markdown("### ⚡ Weekly Switch Alert")
                for _wa in _w_alerts: st.warning(_wa)
                if len(_w_alerts) >= 2:
                    st.error("🚨 **Weekly Switch TRIGGERED** — Weekly rotation mode activate karo (5 weeks).")
                else:
                    st.info("⚠️ Single trigger — monitor karo. Both VIX>30 AND weekly<-5% chahiye for switch.")

            st.divider()
            # ══════════════════════════════════════════════════════════════

        with _s3_tab_b:
            # ── Compute rebalance ─────────────────────────────────────
            if portfolio and st.session_state.dfFiltered is not None:
                dfFiltered      = st.session_state.dfFiltered
                dfStats         = st.session_state.dfStats
                top_n           = st.session_state.top_n_rank
                rank_threshold  = top_n

                top_rank_tickers = dfFiltered.reset_index()
                top_rank_tickers = top_rank_tickers[top_rank_tickers['Rank'] <= rank_threshold]['Ticker']

                current_portfolio_tickers = pd.Series(portfolio)
                entry_stocks = top_rank_tickers[~top_rank_tickers.isin(current_portfolio_tickers)]
                exit_stocks  = current_portfolio_tickers[~current_portfolio_tickers.isin(top_rank_tickers)]
                hold_stocks  = current_portfolio_tickers[current_portfolio_tickers.isin(top_rank_tickers)]

                num_sells = len(exit_stocks)
                entry_stocks = entry_stocks.head(num_sells)

                if len(entry_stocks) < num_sells:
                    entry_stocks = pd.concat([
                        entry_stocks,
                        pd.Series([None] * (num_sells - len(entry_stocks)))
                    ])

                # ── Reasons for exit (v10 logic) ──────────────────────
                reasons_for_exit = []
                for ticker in exit_stocks:
                    if pd.isna(ticker) or ticker == "":
                        reasons_for_exit.append(""); continue
                    reasons    = []
                    stock_data = dfStats[dfStats['Ticker'] == ticker] if dfStats is not None else pd.DataFrame()
                    if len(stock_data) > 0:
                        if stock_data.index[0] > rank_threshold:          reasons.append(f"Rank > {rank_threshold}")
                        if stock_data['volm_cr'].values[0] <= 1:           reasons.append("Volume ≤ 1 Cr")
                        if stock_data['Close'].values[0] <= stock_data['dma200d'].values[0]:
                                                                           reasons.append("Close ≤ 200-DMA")
                        if stock_data['roc12M'].values[0] <= 5.5:          reasons.append("12M ROC ≤ 5.5%")
                        if stock_data['circuit'].values[0] >= 20:          reasons.append("Circuit ≥ 20")
                        if stock_data['AWAY_ATH'].values[0] <= -25:        reasons.append("Away ATH ≤ -25%")
                        if stock_data['roc12M'].values[0] >= 1000:         reasons.append("12M ROC ≥ 1000%")
                        if stock_data['Close'].values[0] <= 30:            reasons.append("Close ≤ ₹30")
                        if stock_data['circuit5'].values[0] > 10:          reasons.append("5% Circuit > 10")
                    else:
                        reasons.append("Not in selected universe")
                    reasons_for_exit.append(", ".join(reasons) if reasons else "Rank dropped")

                reasons_for_exit.extend([""] * (len(entry_stocks) - len(reasons_for_exit)))

                rebalance_table = pd.DataFrame({
                    'S.No.':           range(1, num_sells + 1),
                    'Sell Stocks':     exit_stocks.tolist(),
                    'Buy Stocks':      entry_stocks.tolist(),
                    'Reason for Exit': reasons_for_exit,
                })
                rebalance_table = rebalance_table[
                    ~(rebalance_table['Sell Stocks'].isna() & rebalance_table['Buy Stocks'].isna())
                ]
                rebalance_table.set_index('S.No.', inplace=True)
                st.session_state.sell_list = exit_stocks.dropna().tolist()
                st.session_state.buy_list  = entry_stocks.dropna().tolist()
                st.session_state.rebalance_table = rebalance_table
                st.session_state.rebalance_done  = True

                # ── Summary strip ──────────────────────────────────────
                st.markdown(f"""<div class="reb-strip">
                  <div class="reb-stat"><div class="label">Portfolio</div><div class="val b">{len(portfolio)}</div></div>
                  <div class="reb-stat"><div class="label">Top-{rank_threshold} Screener</div><div class="val b">{len(top_rank_tickers)}</div></div>
                  <div class="reb-stat"><div class="label">SELL (Exit)</div><div class="val r">{len(exit_stocks)}</div></div>
                  <div class="reb-stat"><div class="label">BUY (Entry)</div><div class="val g">{len(entry_stocks.dropna())}</div></div>
                  <div class="reb-stat"><div class="label">HOLD</div><div class="val p">{len(hold_stocks)}</div></div>
                </div>""", unsafe_allow_html=True)

                # ── Sell / Buy / Hold columns ──────────────────────────
                col_sell, col_buy, col_hold = st.columns(3)
                with col_sell:
                    st.markdown('<div class="section-hdr" style="border-left-color:var(--red)">🔴 SELL List</div>', unsafe_allow_html=True)
                    sell_list = exit_stocks.dropna().tolist()
                    if sell_list:
                        chips = " ".join([f'<span class="chip chip-sell">{s}</span>' for s in sell_list])
                        st.markdown(chips, unsafe_allow_html=True)
                        cmp_map = {}
                        if dfStats is not None:
                            cmp_map = dict(zip(dfStats['Ticker'], dfStats['Close']))
                        # Also try dfFiltered for CMP (in case stock is in filtered but not dfStats)
                        if dfFiltered is not None:
                            for t, c in zip(dfFiltered.reset_index()['Ticker'], dfFiltered.reset_index()['Close']):
                                if t not in cmp_map:
                                    cmp_map[t] = c
                        sell_df = pd.DataFrame({
                            "Stock": sell_list,
                            "CMP ₹": [
                                round(cmp_map[s], 2) if s in cmp_map and cmp_map[s] > 0
                                else "N/A *"
                                for s in sell_list
                            ],
                            "Reason": reasons_for_exit[:len(sell_list)]
                        })
                        st.dataframe(sell_df, hide_index=True, use_container_width=True)
                        missing_cmp = [s for s in sell_list if s not in cmp_map or cmp_map.get(s, 0) == 0]
                        if missing_cmp:
                            st.caption(
                                f"* {', '.join(missing_cmp)} — CMP unavailable "
                                f"(stock selected universe ({st.session_state.universe}) mein nahi hai). "
                                "Broker app se manually CMP check karo."
                            )
                    else:
                        st.success("Koi sell nahi hai!")

                with col_buy:
                    st.markdown('<div class="section-hdr" style="border-left-color:var(--green)">🟢 BUY List (New Entry)</div>', unsafe_allow_html=True)
                    buy_list = entry_stocks.dropna().tolist()
                    if buy_list:
                        chips = " ".join([f'<span class="chip chip-buy">{s}</span>' for s in buy_list])
                        st.markdown(chips, unsafe_allow_html=True)
                        rank_map = dict(zip(dfFiltered.reset_index()['Ticker'], dfFiltered.reset_index()['Rank']))
                        cmp_map2 = {}
                        if dfStats is not None:
                            cmp_map2 = dict(zip(dfStats['Ticker'], dfStats['Close']))
                        buy_df = pd.DataFrame({
                            "Stock":        buy_list,
                            "Screener Rank":[rank_map.get(s, "—") for s in buy_list],
                            "CMP ₹":        [round(cmp_map2.get(s, 0), 2) for s in buy_list],
                        })
                        st.dataframe(buy_df, hide_index=True, use_container_width=True)
                    else:
                        st.info("Koi buy nahi hai.")

                with col_hold:
                    st.markdown('<div class="section-hdr" style="border-left-color:var(--violet)">🔵 HOLD (Retain)</div>', unsafe_allow_html=True)
                    if not hold_stocks.empty:
                        chips = " ".join([f'<span class="chip chip-hold">{s}</span>' for s in hold_stocks.tolist()])
                        st.markdown(chips, unsafe_allow_html=True)

                # ── Rebalance table ────────────────────────────────────
                st.markdown('<div class="section-hdr">📋 Rebalance Table (Sell → Buy mapping)</div>', unsafe_allow_html=True)
                if not rebalance_table.empty:
                    st.dataframe(rebalance_table, use_container_width=True)

                st.divider()

                # ══════════════════════════════════════════════════════════
                # WORKFLOW PANEL — Screener → Rebalancer → Order Calculator
                # ══════════════════════════════════════════════════════════
                st.markdown('<div class="section-hdr">🔄 Rebalancer Workflow</div>', unsafe_allow_html=True)

                # ── Step A: Copy Top-N screener list → Google Sheet "Worst Rank Held"
                sell_list_local = exit_stocks.dropna().tolist()
                buy_list_local  = entry_stocks.dropna().tolist()

                cmp_map3 = {}
                if dfStats is not None:
                    cmp_map3 = dict(zip(dfStats['Ticker'], dfStats['Close']))

                # Top-N screener tickers — Worst Rank Held column ke liye
                # Sirf wahi stocks jo filter pass kiye AND rank <= top_n_rank
                # (Excel "Filtered Stocks" sheet ke same 48 stocks)
                _df_sorted = dfFiltered.reset_index()
                if 'Rank' in _df_sorted.columns:
                    _df_sorted = _df_sorted.sort_values('Rank', ascending=True)
                    _top_filtered = _df_sorted[_df_sorted['Rank'] <= st.session_state.top_n_rank]
                else:
                    _top_filtered = _df_sorted.head(st.session_state.top_n_rank)
                top_n_tickers = _top_filtered["Ticker"].tolist()
                worst_rank_text = "\n".join(top_n_tickers) if top_n_tickers else "(no data)"

                st.markdown("""
                <div class="workflow-box">
                <b>📋 Workflow Steps:</b><br>
                <span class="step-tag">1</span> Neeche <b>Top-N Screener list</b> copy karo → Google Sheet ke <b>"Worst Rank Held"</b> column mein paste karo
                  <span style="color:#64748b;font-size:12px;">(ye list rebalancer ko batati hai ki kaun good rank mein hai)</span><br>
                <span class="step-tag">2</span> <b>"Open Portfolio Rebalancer"</b> button dabao → Sell stocks select karo → actual sell value note karo<br>
                <span class="step-tag">3</span> <i>(Optional)</i> Neeche <b>"Buy/Sell order calculate karna chahte hain?"</b> checkbox enable karo → Sell Value enter karo → Buy orders auto-calculate honge
                </div>
                """, unsafe_allow_html=True)

                wa1, wa2 = st.columns([1, 1])
                with wa1:
                    n_top = len(top_n_tickers)
                    st.markdown(f"**📋 Top-{st.session_state.top_n_rank} Screener List — Google Sheet 'Worst Rank Held' column mein paste karo:**")
                    st.caption(f"✅ {n_top} filtered & ranked stocks | Rank 1 se Rank {n_top} tak")
                    st.text_area(
                        "Top-N list — Google Sheet Worst Rank Held column mein paste karo",
                        value=worst_rank_text,
                        height=min(160, max(80, len(top_n_tickers) * 6 + 60)),
                        key="sell_copy_area",
                        label_visibility="collapsed",
                        help="Yeh Top-N screener stocks Google Sheet ke Worst Rank Held column mein paste karo"
                    )
                    # Clipboard copy — uses execCommand fallback for Streamlit iframe sandbox
                    import streamlit.components.v1 as _components
                    _safe_text = worst_rank_text.replace("`", "'").replace("\\", "/")
                    _copy_html = f"""
                    <textarea id="cpytxt" style="position:absolute;left:-9999px;">{_safe_text}</textarea>
                    <button id="cpybtn"
                      onclick="
                        var t=document.getElementById('cpytxt');
                        t.select(); t.setSelectionRange(0,99999);
                        var ok=false;
                        try{{ok=document.execCommand('copy');}}catch(e){{}}
                        if(!ok && navigator.clipboard){{
                          navigator.clipboard.writeText(t.value).then(function(){{
                            document.getElementById('cpybtn').innerHTML='✅ Copied!';
                            document.getElementById('cpybtn').style.background='#16a34a';
                          }});
                        }} else if(ok) {{
                          document.getElementById('cpybtn').innerHTML='✅ Copied!';
                          document.getElementById('cpybtn').style.background='#16a34a';
                        }} else {{
                          alert('Manually select text above aur Ctrl+C / Cmd+C dabao');
                        }}
                      "
                      style="background:#2563eb;color:white;border:none;padding:9px 22px;
                             border-radius:8px;font-weight:700;cursor:pointer;font-size:13px;
                             margin-top:6px;letter-spacing:.2px;
                             box-shadow:0 2px 8px rgba(37,99,235,.3);
                             transition:background .2s;">
                      📋 Copy to Clipboard
                    </button>
                    """
                    _components.html(_copy_html, height=50)

                with wa2:
                    st.markdown("**⚖️ Portfolio Rebalancer:**")
                    st.markdown(f"""
                    <a href="{APPS_SCRIPT_URL}" target="_blank" class="qlink-btn qlink-rebalancer"
                       style="display:block;text-decoration:none;font-weight:700;font-size:14px;
                              color:#ffffff !important;
                              padding:13px 20px;border-radius:10px;text-align:center;margin:4px 0;">
                      ⚖️ Open Portfolio Rebalancer
                    </a>
                    <div style="font-size:11.5px;color:var(--muted);margin-top:8px;line-height:1.7;
                                padding:8px 10px;background:var(--bg);border-radius:6px;border:1px solid var(--border);">
                      📌 Wahan se sell karke <b>actual sell value</b> note karo.<br>
                      ↩️ Phir neeche woh value enter karo.
                    </div>
                    """, unsafe_allow_html=True)

                st.divider()

                # ── Order Calculator (optional) ────────────────────────────
                show_order_calc = st.checkbox(
                    "⚡ Buy/Sell order calculate karna chahte hain?",
                    value=False, key="show_order_calc",
                    help="Sell value enter karke buy order quantities auto-calculate honge"
                )

                if show_order_calc:
                    st.markdown('<div class="section-hdr">⚡ Order Calculator</div>', unsafe_allow_html=True)

                    _oc_mode = st.radio(
                        "Calculation Mode",
                        options=["📈 Only Equity", "🏦 Multi-Asset (Equity + Gold + Liquid)"],
                        index=0, horizontal=True, key="oc_mode_radio",
                        help="Only Equity: sell proceeds sirf equity mein.\nMulti-Asset: Gold/Liquid gap pehle, baaki equity mein."
                    )
                    _multi_asset_mode = "Multi-Asset" in _oc_mode

                    qr1, qr2, qr3, qr4 = st.columns(4)
                    with qr1:
                        capital_add = st.number_input("💰 Capital Addition ₹", min_value=0, value=0, step=5000, key="qr_cap")
                    with qr2:
                        brokerage = st.number_input("🏦 Brokerage/Stock ₹", min_value=0, value=0, step=10, key="qr_brk")
                    with qr3:
                        sell_val_input = st.number_input("💸 Sell Value ₹", min_value=0, value=0, step=1000, key="qr_sell",
                                                         help="Portfolio Rebalancer mein jo actual sell value mili")

                    sell_brk   = len(sell_list_local) * brokerage
                    buy_brk    = len(buy_list_local)  * brokerage
                    gross_pool = sell_val_input + capital_add - sell_brk

                    if not _multi_asset_mode:
                        # ── ONLY EQUITY — original behavior ──────────
                        net_pool  = gross_pool - buy_brk
                        per_stock = net_pool / len(buy_list_local) if buy_list_local else 0
                        with qr4:
                            st.markdown(f"""<div class="metric-card green">
                              <div class="metric-label">Net Pool / Stock</div>
                              <div class="metric-value green">{fmt_inr(per_stock)}</div>
                            </div>""", unsafe_allow_html=True)
                        st.markdown(f"""<div class="reb-strip">
                          <div class="reb-stat"><div class="label">Sell Value</div><div class="val b">₹{sell_val_input:,.0f}</div></div>
                          <div class="reb-stat"><div class="label">+ Capital</div><div class="val g">₹{capital_add:,.0f}</div></div>
                          <div class="reb-stat"><div class="label">- Sell Brok</div><div class="val r">₹{sell_brk:,.0f}</div></div>
                          <div class="reb-stat"><div class="label">- Buy Brok</div><div class="val r">₹{buy_brk:,.0f}</div></div>
                          <div class="reb-stat"><div class="label">Net Pool</div><div class="val g">₹{net_pool:,.0f}</div></div>
                          <div class="reb-stat"><div class="label">Per Stock</div><div class="val g">{fmt_inr(per_stock)}</div></div>
                        </div>""", unsafe_allow_html=True)
                        if sell_val_input == 0 and not capital_add:
                            st.info("💡 Sell Value enter karo → Buy orders auto-calculate honge.")
                        if buy_list_local and per_stock > 0:
                            st.markdown('<div class="section-hdr">📋 Buy Orders (Estimated)</div>', unsafe_allow_html=True)
                            orders = []; total_invested = 0
                            for i, stock in enumerate(buy_list_local, 1):
                                cmp = cmp_map3.get(stock, 0)
                                if cmp > 0:
                                    qty = int(per_stock / cmp); val = qty * cmp; total_invested += val
                                    orders.append({"#": i, "Stock": stock, "CMP ₹": round(cmp, 2),
                                        "Gross Alloc": round(per_stock + brokerage), "Brok ₹": brokerage,
                                        "Net Alloc": round(per_stock), "Qty": qty, "Value ₹": round(val)})
                            if orders:
                                st.dataframe(pd.DataFrame(orders).style.format(
                                    {"CMP ₹": "{:.2f}", "Gross Alloc": "{:,.0f}", "Net Alloc": "{:,.0f}", "Value ₹": "{:,.0f}"}),
                                    use_container_width=True, hide_index=True, height=300)
                                leftover = net_pool - total_invested
                                st.markdown(f"""<div class="reb-strip">
                                  <div class="reb-stat"><div class="label">Total Invested</div><div class="val g">₹{total_invested:,.0f}</div></div>
                                  <div class="reb-stat"><div class="label">Leftover</div><div class="val p">₹{leftover:,.0f}</div></div>
                                  <div class="reb-stat"><div class="label">Buy Orders</div><div class="val b">{len(orders)}</div></div>
                                </div>""", unsafe_allow_html=True)

                    else:
                        # ── MULTI-ASSET MODE — SOP v2026.06 ──────────
                        _oc_rg       = st.session_state.get("_regime_result", {})
                        _oc_sc       = int(_oc_rg.get("score",  _sc))
                        _oc_lbl      = _oc_rg.get("label",  _lbl)
                        _oc_eq_pct   = float(_oc_rg.get("equity", _eq))
                        _oc_gd_pct   = float(_oc_rg.get("gold",   _gd))
                        _oc_cs_pct   = float(_oc_rg.get("cash",   _cs))
                        _oc_total_pf = float(st.session_state.get("_regime_total_pf", _total_pf))
                        _oc_prev_sc  = int(st.session_state.get("_regime_prev_sc",   _prev_sc))
                        if _oc_total_pf == 0:
                            st.warning("⚠️ 'Regime & Allocation' tab mein Total Portfolio Value enter karo.")
                        _ma_c1, _ma_c2 = st.columns(2)
                        with _ma_c1:
                            _oc_gd_curr = st.number_input("🥇 Current GOLDBEES ₹", min_value=0,
                                value=int(st.session_state.get("_gb_curr_val", 0)), step=1000, key="oc_gd_curr")
                        with _ma_c2:
                            _oc_lf_curr = st.number_input("💵 Current Liquid Fund ₹", min_value=0,
                                value=int(st.session_state.get("_lf_curr_val", 0)), step=1000, key="oc_lf_curr")
                        _oc_gd_gap     = max(0.0, _oc_total_pf * _oc_gd_pct - _oc_gd_curr)
                        _oc_cs_gap     = max(0.0, _oc_total_pf * _oc_cs_pct - _oc_lf_curr)
                        _oc_for_gold   = min(_oc_gd_gap, gross_pool)
                        _oc_rem1       = gross_pool - _oc_for_gold
                        _oc_for_liquid = min(_oc_cs_gap, _oc_rem1)
                        _oc_for_equity = max(0.0, _oc_rem1 - _oc_for_liquid - buy_brk)
                        _oc_eq_budget  = _oc_total_pf * _oc_eq_pct
                        _oc_per_stock_tgt = _oc_eq_budget / 30 if _oc_eq_budget > 0 else 0
                        _oc_final_per_stock = _oc_per_stock_tgt
                        _oc_final_stocks    = buy_list_local
                        _oc_final_pool      = _oc_for_equity
                        _wdp_mode = False

                        st.markdown("---")
                        _use_wdp = st.radio(
                            "📅 Weekly Deployment Plan ke hisab se?",
                            options=["✅ Haan — Weekly plan se (Regime shift ho raha hai)",
                                     "❌ Nahi — Normal monthly RB"],
                            index=1, horizontal=True, key="oc_wdp_toggle"
                        )
                        _wdp_mode = "Haan" in _use_wdp

                        if _wdp_mode:
                            from calculations import get_weekly_deployment_plan, get_next_rebalance_dates
                            _wdp_plan  = get_weekly_deployment_plan(prev_score=_oc_prev_sc, curr_score=_oc_sc,
                                total_pf=_oc_total_pf, goldbees_curr=_oc_gd_curr, liquid_curr=_oc_lf_curr,
                                weekly_nav_ret=st.session_state.get("_regime_weekly_ret"),
                                vix_curr=st.session_state.get("_regime_vix"))
                            _wdp_weeks = _wdp_plan.get("weeks", [])
                            _fri_lst   = get_next_rebalance_dates().get("upcoming_fridays", [])
                            if not _wdp_weeks:
                                st.info(f"ℹ️ Score same (Prev {_oc_prev_sc} = Curr {_oc_sc}) — 'Nahi' select karo.")
                            else:
                                _week_labels = [
                                    f"Week {wd['week']} ({_fri_lst[wd['week']-1].strftime('%d %b') if wd['week']-1 < len(_fri_lst) else '?'})"
                                    f" — Eq {wd['eq_pct']}% ₹{wd['eq_val']:,.0f}"
                                    f" | Gold {wd['gd_pct']}% ₹{wd['gd_val']:,.0f}"
                                    f" | Cash {wd['cs_pct']}% ₹{wd['cs_val']:,.0f}"
                                    for wd in _wdp_weeks
                                ]
                                _sel_lbl = st.selectbox("Kaunsa week?", options=_week_labels, key="oc_wk_sel")
                                _sel_wk  = _wdp_weeks[_week_labels.index(_sel_lbl)]
                                _wdp_eq  = float(_sel_wk["eq_val"]); _wdp_gd = float(_sel_wk["gd_val"]); _wdp_cs = float(_sel_wk["cs_val"])
                                _wdp_ps  = _wdp_eq / 30 if _wdp_eq > 0 else 0
                                _wdp_nb  = min(int(_wdp_eq / _wdp_ps) if _wdp_ps > 0 else 0, len(buy_list_local))
                                st.markdown(f"""<div class="reb-strip">
                                  <div class="reb-stat"><div class="label">Wk{_sel_wk['week']} Equity Pool</div><div class="val b">₹{_wdp_eq:,.0f}</div></div>
                                  <div class="reb-stat"><div class="label">Gold Target</div><div class="val" style="color:#b45309">₹{_wdp_gd:,.0f}</div></div>
                                  <div class="reb-stat"><div class="label">Liquid Target</div><div class="val" style="color:#475569">₹{_wdp_cs:,.0f}</div></div>
                                  <div class="reb-stat"><div class="label">Per Stock Tgt</div><div class="val g">₹{_wdp_ps:,.0f}</div></div>
                                  <div class="reb-stat"><div class="label">Buy Entries</div><div class="val b">{_wdp_nb}</div></div>
                                </div>""", unsafe_allow_html=True)
                                _gd_gap_wk = max(0.0, _wdp_gd - _oc_gd_curr); _cs_gap_wk = max(0.0, _wdp_cs - _oc_lf_curr)
                                st.markdown(f'<div style="background:#fef3c7;border-left:3px solid #b45309;border-radius:6px;padding:8px 14px;font-size:12px;color:#78350f;margin-bottom:8px;">{"🥇 GOLDBEES BUY ₹"+f"{_gd_gap_wk:,.0f}" if _gd_gap_wk>=15000 else "🥇 GOLDBEES: within band"} &nbsp;·&nbsp; {"💵 Liquid ADD ₹"+f"{_cs_gap_wk:,.0f}" if _cs_gap_wk>=15000 else "💵 Liquid: within band"}</div>', unsafe_allow_html=True)
                                if _wdp_plan.get("paused"): st.error("⏸ Week 1 PAUSED — VIX > 30 AND weekly return < -5%.")
                                _oc_final_per_stock = _wdp_ps; _oc_final_stocks = buy_list_local[:_wdp_nb]; _oc_final_pool = _wdp_eq
                        else:
                            st.markdown(f"""<div class="reb-strip">
                              <div class="reb-stat"><div class="label">Sell Value</div><div class="val b">₹{sell_val_input:,.0f}</div></div>
                              <div class="reb-stat"><div class="label">→ Gold BUY</div><div class="val" style="color:#b45309">₹{_oc_for_gold:,.0f}</div></div>
                              <div class="reb-stat"><div class="label">→ Liquid ADD</div><div class="val" style="color:#475569">₹{_oc_for_liquid:,.0f}</div></div>
                              <div class="reb-stat"><div class="label">Equity Pool</div><div class="val g">₹{_oc_for_equity:,.0f}</div></div>
                              <div class="reb-stat"><div class="label">Per Stock Tgt</div><div class="val b">₹{_oc_per_stock_tgt:,.0f}</div></div>
                            </div>""", unsafe_allow_html=True)
                            st.caption(f"Regime: {_oc_lbl} · Eq {_oc_eq_pct*100:.0f}% / Gold {_oc_gd_pct*100:.0f}% / Liquid {_oc_cs_pct*100:.0f}% · Per-stock = Eq Budget ÷ 30 = ₹{_oc_per_stock_tgt:,.0f}")

                        if sell_val_input == 0 and not capital_add and not _wdp_mode:
                            st.info("💡 Sell Value enter karo → Buy orders auto-calculate honge.")
                        if buy_list_local and _oc_final_per_stock > 0:
                            st.markdown('<div class="section-hdr">📋 Buy Orders (Estimated)</div>', unsafe_allow_html=True)
                            _ma_orders = []; _ma_invested = 0
                            for _ii, _stk in enumerate(_oc_final_stocks, 1):
                                _cmp = cmp_map3.get(_stk, 0)
                                if _cmp > 0:
                                    _qty = int(_oc_final_per_stock / _cmp); _val = _qty * _cmp; _ma_invested += _val
                                    _ma_orders.append({"#": _ii, "Stock": _stk, "CMP ₹": round(_cmp, 2),
                                        "Gross Alloc": round(_oc_final_per_stock + brokerage), "Brok ₹": brokerage,
                                        "Net Alloc": round(_oc_final_per_stock), "Qty": _qty, "Value ₹": round(_val)})
                            if _ma_orders:
                                st.dataframe(pd.DataFrame(_ma_orders).style.format(
                                    {"CMP ₹": "{:.2f}", "Gross Alloc": "{:,.0f}", "Net Alloc": "{:,.0f}", "Value ₹": "{:,.0f}"}),
                                    use_container_width=True, hide_index=True, height=300)
                                st.markdown(f"""<div class="reb-strip">
                                  <div class="reb-stat"><div class="label">Total Invested</div><div class="val g">₹{_ma_invested:,.0f}</div></div>
                                  <div class="reb-stat"><div class="label">Leftover</div><div class="val p">₹{_oc_final_pool-_ma_invested:,.0f}</div></div>
                                  <div class="reb-stat"><div class="label">Buy Orders</div><div class="val b">{len(_ma_orders)}</div></div>
                                </div>""", unsafe_allow_html=True)

                st.divider()
                if st.button("▶ Next: Apply & Export →", type="primary"):
                    st.session_state.current_step = 4; st.rerun()

            elif not portfolio:
                st.info("⬆️ Upar se portfolio data load karo (Google Sheet ya CSV).")


    # ═══════════════════════════════════════════════════════════════
    # STEP 4 — APPLY & EXPORT
    # ═══════════════════════════════════════════════════════════════
    elif st.session_state.current_step == 4:
        st.markdown('<div class="section-hdr">💾 Step 4 — Apply & Export</div>', unsafe_allow_html=True)

        sell        = st.session_state.sell_list or []
        buy         = st.session_state.buy_list  or []
        portfolio   = st.session_state.reb_portfolio or []
        dfStats     = st.session_state.dfStats
        dfFiltered  = st.session_state.dfFiltered
        reb_table   = st.session_state.rebalance_table
        failed_blank= st.session_state.failed_blank or []
        U           = st.session_state.universe
        top_n       = st.session_state.top_n_rank
        rank_method = st.session_state.ranking_method
        api_source  = st.session_state.data_source
        end_date    = st.session_state.lookback_date

        # ── Summary ───────────────────────────────────────────────
        st.markdown(f"""<div class="reb-strip">
          <div class="reb-stat"><div class="label">Exits (SELL)</div><div class="val r">{len(sell)}</div></div>
          <div class="reb-stat"><div class="label">New Entries (BUY)</div><div class="val g">{len(buy)}</div></div>
          <div class="reb-stat"><div class="label">Retained (HOLD)</div><div class="val p">{len(portfolio) - len(sell)}</div></div>
          <div class="reb-stat"><div class="label">New Portfolio Size</div><div class="val b">{len(portfolio) - len(sell) + len(buy)}</div></div>
        </div>""", unsafe_allow_html=True)
        if dfFiltered is not None and dfStats is not None:
            st.markdown('<div class="section-hdr">💾 Excel Export (v10 Format — 4 Sheets)</div>', unsafe_allow_html=True)

            # ── Failed Downloads DF ──────────────────────────────
            if failed_blank:
                df_failed = pd.DataFrame({
                    'S.No.':        range(1, len(failed_blank)+1),
                    'Failed Stock': failed_blank
                }).set_index('S.No.')
            else:
                df_failed = pd.DataFrame(columns=['Failed Stock'])
                df_failed.index.name = 'S.No.'

            # ── Rebalance Table ──────────────────────────────────
            if reb_table is None or reb_table.empty:
                reb_table = pd.DataFrame(columns=['Sell Stocks','Buy Stocks','Reason for Exit'])
                reb_table.index.name = 'S.No.'

            filtered = dfFiltered.copy()

            excel_file = f"{end_date.strftime('%Y-%m-%d')}_{U}_{rank_method}_{api_source}_lookback.xlsx"

            with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
                dfStats.to_excel(   writer, sheet_name="Unfiltered Stocks",     index=True)
                filtered.to_excel(  writer, sheet_name="Filtered Stocks",       index=True)
                df_failed.to_excel( writer, sheet_name="Failed Downloads",      index=True)
                reb_table.to_excel( writer, sheet_name="Portfolio Rebalancing", index=True)

            # Apply v10 formatting
            try:
                format_excel_unfiltered(excel_file, U, top_n)
                format_excel_filtered(excel_file, U, top_n)
                format_simple_sheet(excel_file, "Failed Downloads")
                format_simple_sheet(excel_file, "Portfolio Rebalancing")
            except Exception as e:
                st.warning(f"Excel formatting partial error (file still usable): {e}")

            with open(excel_file, "rb") as f:
                st.download_button(
                    label     = "📥 Download Excel (4 Sheets)",
                    data      = f.read(),
                    file_name = excel_file,
                    mime      = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type      = "primary",
                )

            st.success(f"✅ Excel ready: `{excel_file}`")
            st.markdown(f"""
            <div style="background:var(--green-bg);border:1px solid var(--green-bdr);border-radius:var(--radius-md);
                        padding:10px 16px;font-size:12px;color:#15803d;margin:4px 0 12px 0;">
            📄 <b>4 Sheets:</b> &nbsp;
            <span style="background:white;padding:2px 8px;border-radius:12px;margin:0 3px;border:1px solid var(--green-bdr);">Unfiltered Stocks</span>
            <span style="background:white;padding:2px 8px;border-radius:12px;margin:0 3px;border:1px solid var(--green-bdr);">Filtered Stocks</span>
            <span style="background:white;padding:2px 8px;border-radius:12px;margin:0 3px;border:1px solid var(--green-bdr);">Failed Downloads</span>
            <span style="background:white;padding:2px 8px;border-radius:12px;margin:0 3px;border:1px solid var(--green-bdr);">Portfolio Rebalancing</span>
            </div>
            """, unsafe_allow_html=True)

        # ── Apps Script / Rebalance Sheet links ───────────────────
        st.markdown('<div class="section-hdr">📊 Apps Script Workflow — Quick Links</div>', unsafe_allow_html=True)

        col_links = st.columns(2)
        with col_links[0]:
            st.markdown(f"""
            <a href="{APPS_SCRIPT_URL}" target="_blank" class="qlink-btn qlink-rebalancer">
            ⚖️ Portfolio Rebalancer
            </a>
            """, unsafe_allow_html=True)
        with col_links[1]:
            st.markdown("""
            <a href="https://prayan2702.github.io/momn-dashboard/" target="_blank" class="qlink-btn qlink-dashboard">
            📈 Portfolio Dashboard
            </a>
            """, unsafe_allow_html=True)

        # ── Rebalancing table on screen ───────────────────────────
        if reb_table is not None and not reb_table.empty:
            st.markdown('<div class="section-hdr">📋 Portfolio Rebalancing</div>', unsafe_allow_html=True)
            st.dataframe(reb_table, use_container_width=True)



        st.divider()
        col_r1, col_r2 = st.columns(2)
        with col_r1:
            if st.button("🔄 New Month — Restart from Step 1", use_container_width=True):
                for k in list(st.session_state.keys()):
                    del st.session_state[k]
                st.rerun()
        with col_r2:
            if st.button("← Step 3 — Edit Rebalance"):
                st.session_state.current_step = 3; st.rerun()
